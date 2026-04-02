
# src/main.py
from __future__ import annotations

import os
import base64
import json
import hashlib
import zipfile
import unicodedata
import csv
import openpyxl
from datetime import datetime, timedelta, date
from typing import Optional
from io import BytesIO

import openpyxl
from PIL import Image

from sqlalchemy import text

from fastapi import FastAPI, Request, UploadFile, Form, Depends, HTTPException, File
from fastapi.responses import HTMLResponse, RedirectResponse, Response, PlainTextResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates

from sqlmodel import SQLModel, Session, create_engine, select

from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm

from passlib.hash import bcrypt, pbkdf2_sha256

from openai import OpenAI
import pypdfium2 as pdfium

PDF_TEXT_READER = None
try:
    from pypdf import PdfReader as _PdfReader  # type: ignore
    PDF_TEXT_READER = _PdfReader
except Exception:
    try:
        from PyPDF2 import PdfReader as _PdfReader  # type: ignore
        PDF_TEXT_READER = _PdfReader
    except Exception:
        PDF_TEXT_READER = None

# PDF merge (pypdf/PyPDF2) - import robuste
PDF_MERGER_AVAILABLE = False
PdfMerger = None
try:
    from pypdf import PdfMerger as _PdfMerger  # type: ignore
    PdfMerger = _PdfMerger
    PDF_MERGER_AVAILABLE = True
except Exception:
    try:
        from PyPDF2 import PdfMerger as _PdfMerger  # type: ignore
        PdfMerger = _PdfMerger
        PDF_MERGER_AVAILABLE = True
    except Exception:
        PDF_MERGER_AVAILABLE = False

from .pdf_orders import generate_order_pdf, generate_orders_pdf

from .models import (
    # auth / core
    User,
    # trésorerie
    BankTxn,
    # délais
    PayDelayRow,
    # docs
    Document,
    Invoice,
    InvoicePaymentMatch,
    InvoiceDocument,
    # paiements + fournisseurs
    Payment,
    Supplier,
    SupplierBankAccount,
    # cashflow master
    CashflowRubrique,
    CashflowActual,
    # comptes société
    CompanyBankAccount,
    # banques templates
    Bank,
    # règlements batch
    PaymentBatch,
    PaymentLine,
    # pilotage
    BudgetLine,
    Post,
    AllocationLine,
)

# ---------------- PATHS / DB ----------------

BASE_DIR = os.path.dirname(__file__)
DATA_DIR = os.path.join(BASE_DIR, "..", "data")
UPLOAD_DIR = os.path.join(DATA_DIR, "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

DB_PATH = os.path.join(DATA_DIR, "app.db")


def _db_url() -> str:
    """
    Render sets DATABASE_URL for Postgres (recommended).
    If not provided, fallback to SQLite (local).
    """
    url = os.environ.get("DATABASE_URL")
    if url:
        if url.startswith("postgres://"):
            url = url.replace("postgres://", "postgresql+psycopg2://", 1)
        if url.startswith("postgresql://"):
            url = url.replace("postgresql://", "postgresql+psycopg2://", 1)
        return url
    return f"sqlite:///{DB_PATH}"


engine = create_engine(_db_url(), echo=False)


def init_db():
    SQLModel.metadata.create_all(engine)


def get_session():
    with Session(engine) as session:
        yield session


def _table_exists_sqlite(conn, table_name: str) -> bool:
    r = conn.execute(
        text("SELECT name FROM sqlite_master WHERE type='table' AND name=:t"),
        {"t": table_name},
    ).fetchone()
    return bool(r)


def _table_columns_pg(conn, table_name: str) -> set[str]:
    rows = conn.execute(
        text(
            """
            SELECT column_name
            FROM information_schema.columns
            WHERE table_name=:t
            """
        ),
        {"t": table_name},
    ).fetchall()
    return {r[0] for r in rows}


def ensure_invoice_document_columns():
    cols = [
        ("ref_no", "VARCHAR(80)"),
        ("ref_date", "TIMESTAMP"),
    ]

    with engine.begin() as conn:
        if conn.dialect.name == "sqlite":
            if not _table_exists_sqlite(conn, "invoicedocument"):
                return
            existing = {r[1] for r in conn.execute(text("PRAGMA table_info(invoicedocument)")).fetchall()}
        else:
            existing = _table_columns_pg(conn, "invoicedocument")

        for name, sqltype in cols:
            if name not in existing:
                conn.execute(text(f"ALTER TABLE invoicedocument ADD COLUMN {name} {sqltype}"))


def ensure_payment_tables_columns():
    """
    Micro-migration: ajoute colonnes si manquantes (sécurise Postgres/SQLite).
    """
    batch_cols = [
        ("company_account_id", "INTEGER"),
        ("bank_name", "VARCHAR(60)"),
        ("debit_account", "VARCHAR(80)"),
        ("payment_date", "TIMESTAMP"),
        ("total_amount", "FLOAT"),
        ("status", "VARCHAR(20)"),
        ("created_at", "TIMESTAMP"),
    ]

    line_cols = [
        ("batch_id", "INTEGER"),
        ("invoice_id", "INTEGER"),
        ("amount", "FLOAT"),
        ("created_at", "TIMESTAMP"),
    ]

    with engine.begin() as conn:
        # paymentbatch
        if conn.dialect.name == "sqlite":
            existing_batch = (
                {r[1] for r in conn.execute(text("PRAGMA table_info(paymentbatch)")).fetchall()}
                if _table_exists_sqlite(conn, "paymentbatch")
                else set()
            )
        else:
            existing_batch = _table_columns_pg(conn, "paymentbatch")

        for name, sqltype in batch_cols:
            if existing_batch and name not in existing_batch:
                conn.execute(text(f"ALTER TABLE paymentbatch ADD COLUMN {name} {sqltype}"))

        # paymentline
        if conn.dialect.name == "sqlite":
            existing_line = (
                {r[1] for r in conn.execute(text("PRAGMA table_info(paymentline)")).fetchall()}
                if _table_exists_sqlite(conn, "paymentline")
                else set()
            )
        else:
            existing_line = _table_columns_pg(conn, "paymentline")

        for name, sqltype in line_cols:
            if existing_line and name not in existing_line:
                conn.execute(text(f"ALTER TABLE paymentline ADD COLUMN {name} {sqltype}"))


def ensure_company_bank_columns():
    """
    Micro-migration: ajoute colonnes d'attestation RIB dans companybankaccount
    (SQLite ou Postgres) si elles n'existent pas.
    """
    cols = [
        ("attestation_filename", "VARCHAR(255)"),
        ("attestation_path", "VARCHAR(255)"),
    ]

    with engine.begin() as conn:
        if conn.dialect.name == "sqlite":
            if not _table_exists_sqlite(conn, "companybankaccount"):
                return
            existing = {r[1] for r in conn.execute(text("PRAGMA table_info(companybankaccount)")).fetchall()}
        else:
            existing = _table_columns_pg(conn, "companybankaccount")

        for name, sqltype in cols:
            if name not in existing:
                conn.execute(text(f"ALTER TABLE companybankaccount ADD COLUMN {name} {sqltype}"))

def ensure_supplier_columns():
    """
    Micro-migration: ajoute les colonnes manquantes dans la table supplier (SQLite/Postgres).
    """
    cols = [
        ("name", "VARCHAR(200)"),
        ("ice", "VARCHAR(30)"),
        ("if_code", "VARCHAR(30)"),
        ("rc", "VARCHAR(30)"),
        ("rc_city", "VARCHAR(60)"),
        ("address", "VARCHAR(250)"),
        ("is_foreign", "BOOLEAN"),
        ("country_code", "VARCHAR(2)"),
        ("created_at", "TIMESTAMP"),
    ]

    with engine.begin() as conn:
        if conn.dialect.name == "sqlite":
            if not _table_exists_sqlite(conn, "supplier"):
                return
            existing = {r[1] for r in conn.execute(text("PRAGMA table_info(supplier)")).fetchall()}
        else:
            existing = _table_columns_pg(conn, "supplier")

        for name, sqltype in cols:
            if name not in existing:
                conn.execute(text(f"ALTER TABLE supplier ADD COLUMN {name} {sqltype}"))


def ensure_supplier_bank_columns():
    """
    Micro-migration: ajoute les colonnes manquantes dans la table supplierbankaccount (SQLite/Postgres).
    """
    cols = [
        ("supplier_id", "INTEGER"),
        ("bank_name", "VARCHAR(80)"),
        ("agency_name", "VARCHAR(80)"),
        ("rib_or_iban", "VARCHAR(50)"),
        ("swift", "VARCHAR(20)"),
        ("attestation_filename", "VARCHAR(255)"),
        ("attestation_path", "VARCHAR(255)"),
        ("created_at", "TIMESTAMP"),
    ]

    with engine.begin() as conn:
        if conn.dialect.name == "sqlite":
            if not _table_exists_sqlite(conn, "supplierbankaccount"):
                return
            existing = {r[1] for r in conn.execute(text("PRAGMA table_info(supplierbankaccount)")).fetchall()}
        else:
            existing = _table_columns_pg(conn, "supplierbankaccount")

        for name, sqltype in cols:
            if name not in existing:
                conn.execute(text(f"ALTER TABLE supplierbankaccount ADD COLUMN {name} {sqltype}"))
                


def ensure_banktxn_columns():
    """Micro-migration: ajoute colonnes manquantes dans banktxn (SQLite/Postgres)."""
    cols = [
        ("bank_name", "VARCHAR(80)"),
        ("account_no", "VARCHAR(80)"),
        ("currency", "VARCHAR(10)"),
        ("label_norm", "VARCHAR(400)"),
        ("dedup_key", "VARCHAR(200)"),
        ("created_at", "TIMESTAMP"),
        ("value_date", "TIMESTAMP"),
        ("balance", "FLOAT"),
        ("debit", "FLOAT"),
        ("credit", "FLOAT"),
        ("label", "VARCHAR(400)"),
        ("date", "TIMESTAMP"),
    ]
    with engine.begin() as conn:
        if conn.dialect.name == "sqlite":
            if not _table_exists_sqlite(conn, "banktxn"):
                return
            existing = {r[1] for r in conn.execute(text("PRAGMA table_info(banktxn)")).fetchall()}
        else:
            existing = _table_columns_pg(conn, "banktxn")

        for name, sqltype in cols:
            if name not in existing:
                conn.execute(text(f"ALTER TABLE banktxn ADD COLUMN {name} {sqltype}"))


def ensure_invoice_payment_match_columns():
    """Micro-migration: ajoute colonnes manquantes dans invoicepaymentmatch."""
    cols = [
        ("invoice_id", "INTEGER"),
        ("banktxn_id", "INTEGER"),
        ("matched_amount", "FLOAT"),
        ("method", "VARCHAR(12)"),
        ("notes", "VARCHAR(200)"),
        ("created_at", "TIMESTAMP"),
    ]
    with engine.begin() as conn:
        if conn.dialect.name == "sqlite":
            if not _table_exists_sqlite(conn, "invoicepaymentmatch"):
                return
            existing = {r[1] for r in conn.execute(text("PRAGMA table_info(invoicepaymentmatch)")).fetchall()}
        else:
            existing = _table_columns_pg(conn, "invoicepaymentmatch")

        for name, sqltype in cols:
            if name not in existing:
                conn.execute(text(f"ALTER TABLE invoicepaymentmatch ADD COLUMN {name} {sqltype}"))

def ensure_invoice_columns():
    """
    Micro-migration sans Alembic (SQLite/Postgres).
    Ajoute les colonnes manquantes dans la table invoice.
    """
    cols = [
        ("supplier_name", "VARCHAR(200)"),
        ("invoice_no", "VARCHAR(80)"),
        ("invoice_date", "TIMESTAMP"),
        ("reception_date", "TIMESTAMP"),
        ("amount", "FLOAT"),
        ("currency", "VARCHAR(10)"),
        ("department", "VARCHAR(80)"),
        ("analytic", "VARCHAR(80)"),
        ("category", "VARCHAR(60)"),
        ("due_date", "TIMESTAMP"),
        ("due_date_planned", "TIMESTAMP"),
        ("due_date_agreed", "TIMESTAMP"),
        ("status", "VARCHAR(20)"),
        ("payment_date", "TIMESTAMP"),
        ("payment_mode", "VARCHAR(60)"),
        ("dedup_key", "VARCHAR(140)"),
        ("file_path", "VARCHAR(255)"),
        ("supplier_id", "INTEGER"),
        ("supplier_if", "VARCHAR(30)"),
        ("supplier_ice", "VARCHAR(30)"),
        ("supplier_rc", "VARCHAR(30)"),
        ("supplier_rc_city", "VARCHAR(60)"),
        ("supplier_address", "VARCHAR(250)"),
        ("service_date", "TIMESTAMP"),
        ("nature_operation", "VARCHAR(20)"),
        ("payment_terms_days", "INTEGER"),
        ("derogation_sector", "BOOLEAN"),
        ("derogation_days", "INTEGER"),
        ("derogation_ref", "VARCHAR(120)"),
        ("calc_start_date", "TIMESTAMP"),
        ("calc_start_rule", "VARCHAR(30)"),
        ("applied_terms_days", "INTEGER"),
        ("legal_due_date", "TIMESTAMP"),
        ("amount_ht", "FLOAT"),
        ("vat_rate", "FLOAT"),
        ("amount_vat", "FLOAT"),
        ("amount_ttc", "FLOAT"),
        ("amount_paid", "FLOAT"),
        ("cashflow_category", "VARCHAR(60)"),
        ("cashflow_rubrique", "VARCHAR(80)"),
        ("reporting_groupe", "VARCHAR(80)"),
        ("impact_budget", "BOOLEAN"),
        ("site", "VARCHAR(30)"),
        ("cost_center", "VARCHAR(60)"),
        ("project", "VARCHAR(80)"),
        ("gl_account", "VARCHAR(20)"),
        ("expense_nature", "VARCHAR(60)"),
        ("is_disputed", "BOOLEAN"),
        ("dispute_reason", "VARCHAR(200)"),
        ("disputed_amount", "FLOAT"),
        ("bc_no", "VARCHAR(80)"),
        ("bc_date", "TIMESTAMP"),
        ("br_no", "VARCHAR(80)"),
        ("br_date", "TIMESTAMP"),
    ]

    with engine.begin() as conn:
        if conn.dialect.name == "sqlite":
            if not _table_exists_sqlite(conn, "invoice"):
                return
            existing = {r[1] for r in conn.execute(text("PRAGMA table_info(invoice)")).fetchall()}
        else:
            existing = _table_columns_pg(conn, "invoice")

        for name, sqltype in cols:
            if name not in existing:
                conn.execute(text(f"ALTER TABLE invoice ADD COLUMN {name} {sqltype}"))


# ---------------- APP ----------------

app = FastAPI(title="Mini-ERP Trésorerie")
APP_VERSION = "v-auth-final-vision-7"  # bump

templates = Jinja2Templates(directory=os.path.join(BASE_DIR, "templates"))
app.mount("/static", StaticFiles(directory=os.path.join(BASE_DIR, "static")), name="static")
app.mount("/uploads", StaticFiles(directory=UPLOAD_DIR), name="uploads")

os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(os.path.join(BASE_DIR, "static"), exist_ok=True)

OPENAI_API_KEY = (os.environ.get("OPENAI_API_KEY") or "").strip()
openai_client = OpenAI(api_key=OPENAI_API_KEY) if OPENAI_API_KEY else None
COOKIE_SECURE = os.environ.get("COOKIE_SECURE", "1") == "1"


# ---------------- HELPERS ----------------

def format_mad(value: Optional[float]) -> str:
    if value is None:
        return ""
    try:
        s = f"{float(value):,.2f}"
        s = s.replace(",", " ").replace(".", ",")
        return s
    except Exception:
        return str(value)


templates.env.filters["mad"] = format_mad


def _openai_enabled() -> bool:
    return bool(openai_client and OPENAI_API_KEY)


def _json_loads_safe(raw: Optional[str]) -> Optional[dict]:
    if not raw:
        return None
    try:
        return json.loads(raw)
    except Exception:
        return None


def _extract_output_text(resp) -> Optional[str]:
    raw = getattr(resp, "output_text", None)
    if raw:
        return raw
    parts = []
    for item in (getattr(resp, "output", None) or []):
        for c in (getattr(item, "content", None) or []):
            if getattr(c, "type", None) in ("output_text", "text") and getattr(c, "text", None):
                parts.append(c.text)
    joined = "\n".join(parts).strip()
    return joined or None


def _call_openai_json(*, model: str, input_payload: list, schema_name: str, schema: dict, max_output_tokens: int = 1200) -> dict:
    if not _openai_enabled():
        raise RuntimeError("OPENAI indisponible")
    resp = openai_client.responses.create(
        model=model,
        input=input_payload,
        text={"format": {"type": "json_schema", "name": schema_name, "schema": schema, "strict": True}},
        max_output_tokens=max_output_tokens,
    )
    raw = _extract_output_text(resp)
    data = _json_loads_safe(raw)
    if not data:
        raise RuntimeError("Réponse OpenAI vide ou invalide")
    return data


def _read_pdf_text_safe(file_path: str) -> str:
    if not PDF_TEXT_READER:
        return ""
    try:
        reader = PDF_TEXT_READER(file_path)
        parts = []
        for page in getattr(reader, "pages", [])[:5]:
            try:
                parts.append(page.extract_text() or "")
            except Exception:
                pass
        return "\n".join(parts).strip()
    except Exception:
        return ""


def _read_text_hints(file_path: str, original_mime: str) -> str:
    mime = (original_mime or "").lower()
    path = file_path.lower()
    if mime == "application/pdf" or path.endswith(".pdf"):
        txt = _read_pdf_text_safe(file_path)
        if txt:
            return txt
    try:
        with open(file_path, "rb") as f:
            blob = f.read(200000)
        return blob.decode("utf-8", "ignore")
    except Exception:
        return ""


def _first_match(text_value: str, patterns: list[str]) -> Optional[str]:
    for pat in patterns:
        m = re.search(pat, text_value, re.I | re.M)
        if m:
            for g in m.groups():
                if g:
                    return g.strip()
            return m.group(0).strip()
    return None


def _parse_amount_guess(value: Optional[str]) -> Optional[float]:
    if not value:
        return None
    s = str(value).strip()
    s = re.sub(r"[^0-9,\.\-]", "", s)
    if not s:
        return None
    if s.count(",") and s.count("."):
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif s.count(",") and not s.count("."):
        s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None


def _extract_amount_triplet(text_value: str) -> dict:
    """Essaie d'extraire TTC / HT / TVA / taux depuis du texte brut (OCR/PDF/bytes)."""
    def pick(patterns: list[str]) -> Optional[float]:
        return _parse_amount_guess(_first_match(text_value, patterns))

    ttc = pick([
        r"(?:TOTAL\s*TTC|NET\s*[ÀA]\s*PAYER|GRAND\s*TOTAL|TOTAL\s*FACTURE)\s*[: ]*([0-9\s.,]+)",
        r"(?:MONTANT\s*TTC)\s*[: ]*([0-9\s.,]+)",
    ])
    ht = pick([
        r"(?:TOTAL\s*HT|MONTANT\s*HT)\s*[: ]*([0-9\s.,]+)",
        r"(?:SOUS\s*TOTAL|TOTAL\s*HORS\s*TAXE)\s*[: ]*([0-9\s.,]+)",
    ])
    vat = pick([
        r"(?:TOTAL\s*TVA|MONTANT\s*TVA|TVA)\s*[: ]*([0-9\s.,]+)",
    ])
    rate = pick([
        r"(?:TVA|TAXE)\s*[: ]*([0-9]{1,2}(?:[.,][0-9]{1,2})?)\s*%",
        r"([0-9]{1,2}(?:[.,][0-9]{1,2})?)\s*%",
    ])

    # Reconstitution intelligente si partiel
    if rate is not None and rate <= 1:
        rate = rate * 100
    r = (rate / 100.0) if rate is not None else None

    if ttc is None and ht is not None and vat is not None:
        ttc = round(ht + vat, 2)
    if vat is None and ht is not None and ttc is not None:
        vat = round(ttc - ht, 2)
    if ht is None and ttc is not None and r is not None:
        ht = round(ttc / (1 + r), 2)
    if vat is None and ht is not None and r is not None:
        vat = round(ht * r, 2)
    if ttc is None and ht is not None and r is not None:
        ttc = round(ht * (1 + r), 2)
    if rate is None and ht not in (None, 0) and vat is not None:
        rate = round((vat / ht) * 100.0, 2)

    return {
        "amount_ht": ht,
        "amount_vat": vat,
        "amount_ttc": ttc,
        "vat_rate": rate,
    }


def _guess_currency(text_value: str) -> str:
    t = (text_value or "").upper()
    if "EUR" in t or "€" in t:
        return "EUR"
    if "USD" in t or "US$" in t or "$" in t:
        return "USD"
    return "MAD"


def _guess_supplier_name(text_value: str, file_path: str) -> Optional[str]:
    lines = [ln.strip() for ln in (text_value or "").splitlines() if ln.strip()]
    for ln in lines[:12]:
        if len(ln) > 3 and len(ln) < 120 and not re.search(r"(facture|invoice|bon de|relev|page|date)", ln, re.I):
            if re.search(r"[A-Za-zÀ-ÿ]", ln):
                return ln[:200]
    stem = pathlib.Path(file_path).stem.replace("_", " ").replace("-", " ").strip()
    return stem[:200] if stem else None


def _fallback_extract_doc_data(file_path: str, original_mime: str, *, detailed_invoice: bool = False, rib_mode: bool = False) -> dict:
    text_value = _read_text_hints(file_path, original_mime)
    upper = text_value.upper()
    stem = pathlib.Path(file_path).stem

    invoice_no = _first_match(text_value, [
        r"(?:FACTURE|INVOICE|FACT|N[°ºO]?\s*FACTURE)\s*[:#-]?\s*([A-Z0-9\-_/]{3,})",
        r"\b(?:FAC|INV)[-_ ]?([A-Z0-9\-_/]{3,})\b",
    ])
    bc_no = _first_match(text_value, [
        r"(?:BON\s+DE\s+COMMANDE|\bBC\b)\s*[:#-]?\s*([A-Z0-9\-_/]{2,})",
    ])
    br_no = _first_match(text_value, [
        r"(?:BON\s+DE\s+R[ÉE]CEPTION|BON\s+DE\s+LIVRAISON|\bBR\b|\bBL\b)\s*[:#-]?\s*([A-Z0-9\-_/]{2,})",
    ])
    date_guess = _first_match(text_value, [
        r"\b(20\d{2}[/-]\d{2}[/-]\d{2})\b",
        r"\b(\d{2}[/-]\d{2}[/-]20\d{2})\b",
    ])
    amount_bits = _extract_amount_triplet(text_value)
    amount_guess = amount_bits.get("amount_ttc")
    terms_guess = _parse_amount_guess(_first_match(text_value, [r"(\d{1,3})\s*(?:jours|j)\b"]))
    supplier_guess = _guess_supplier_name(text_value, file_path)
    currency_guess = _guess_currency(text_value)

    doc_type = "AUTRE"
    if invoice_no or re.search(r"(FACTURE|INVOICE)", upper):
        doc_type = "FACTURE"
    elif bc_no or re.search(r"(BON\s+DE\s+COMMANDE|\bBC\b)", upper):
        doc_type = "BC"
    elif br_no or re.search(r"(BON\s+DE\s+R[ÉE]CEPTION|BON\s+DE\s+LIVRAISON|\bBR\b|\bBL\b)", upper):
        doc_type = "BR"

    if rib_mode:
        return {
            "beneficiary_name": supplier_guess,
            "bank_name": _first_match(text_value, [r"(?:BANQUE|BANK)\s*[: -]?\s*([A-ZÀ-ÿ0-9 \-]{3,60})"]),
            "rib_or_iban": _first_match(text_value, [r"\b([A-Z]{2}\d{2}[A-Z0-9]{11,30})\b", r"\b(\d{20,30})\b"]),
            "swift": _first_match(text_value, [r"\b([A-Z]{6}[A-Z0-9]{2}(?:[A-Z0-9]{3})?)\b"]),
        }

    base_data = {
        "doc_type": doc_type,
        "supplier_name": supplier_guess,
        "supplier_ice": _first_match(text_value, [r"\bICE\s*[:#-]?\s*(\d{6,20})\b"]),
        "supplier_if": _first_match(text_value, [r"\bIF\s*[:#-]?\s*(\d{3,20})\b"]),
        "supplier_rc": _first_match(text_value, [r"\bRC\s*[:#-]?\s*([A-Z0-9\-_/]{2,30})\b"]),
        "supplier_rc_city": None,
        "supplier_address": None,
        "invoice_no": invoice_no,
        "invoice_date": date_guess,
        "amount_ttc": amount_guess,
        "currency": currency_guess,
        "payment_terms_days": int(terms_guess) if terms_guess is not None else None,
        "bc_no": bc_no,
        "bc_date": date_guess if doc_type == "BC" else None,
        "br_no": br_no,
        "br_date": date_guess if doc_type == "BR" else None,
        "service_date": date_guess if doc_type == "BR" else None,
    }
    if detailed_invoice:
        base_data.update({
            "amount_ht": amount_bits.get("amount_ht"),
            "vat_rate": amount_bits.get("vat_rate"),
            "amount_vat": amount_bits.get("amount_vat"),
            "department": None,
            "analytic": None,
            "nature_operation": "SERVICES" if re.search(r"SERVICE", upper) else "BIENS",
            "due_date": None,
            "due_date_planned": None,
            "due_date_agreed": None,
        })
    return base_data


def _invoice_completion_snapshot(inv: Invoice) -> dict:
    flags = _invoice_quality_flags(inv)
    score = max(0, 100 - 20 * len(flags))
    return {"score": score, "missing": flags}


def _slugify_filename(s: str, maxlen: int = 80) -> str:
    """Convertit un texte en nom de fichier safe (ASCII, sans caractères spéciaux)."""
    s = (s or "").strip()
    if not s:
        return "document"
    s = unicodedata.normalize("NFKD", s)
    s = s.encode("ascii", "ignore").decode("ascii")
    s = "".join(ch if (ch.isalnum() or ch in ("-", "_")) else "_" for ch in s)
    while "__" in s:
        s = s.replace("__", "_")
    s = s.strip("_-")
    return (s or "document")[:maxlen]


def hash_password(p: str) -> str:
    return pbkdf2_sha256.hash(p)


def verify_password(p: str, stored_hash: str) -> bool:
    try:
        if stored_hash.startswith("$2"):
            return bcrypt.verify(p, stored_hash)
        return pbkdf2_sha256.verify(p, stored_hash)
    except Exception:
        return False


def get_current_user(request: Request, session: Session) -> User | None:
    uid = request.cookies.get("uid")
    if not uid:
        return None
    try:
        uid_int = int(uid)
    except Exception:
        return None
    return session.get(User, uid_int)



def ensure_bootstrap_users(session: Session):
    u1 = (os.environ.get("BOOTSTRAP_USER1") or "").strip()
    p1 = (os.environ.get("BOOTSTRAP_PASS1") or "")
    u2 = (os.environ.get("BOOTSTRAP_USER2") or "").strip()
    p2 = (os.environ.get("BOOTSTRAP_PASS2") or "")

    existing_users = session.exec(select(User)).all()
    pairs: list[tuple[str, str]] = []
    if u1 and p1:
        pairs.append((u1, p1))
    if u2 and p2:
        pairs.append((u2, p2))

    # Bootstrap de secours pour éviter un verrouillage complet après déploiement initial.
    if not existing_users and not pairs:
        pairs.append(("admin", "admin123"))

    for username, password in pairs:
        existing = session.exec(select(User).where(User.username == username)).first()
        new_hash = hash_password(password)
        if not existing:
            session.add(User(username=username, password_hash=new_hash))
        else:
            existing.password_hash = new_hash
            session.add(existing)

    session.commit()



@app.on_event("startup")
def on_startup():
    init_db()
    ensure_invoice_columns()
    ensure_invoice_document_columns()
    ensure_payment_tables_columns()
    ensure_company_bank_columns()
    
    ensure_supplier_columns()
    ensure_supplier_bank_columns()

    with Session(engine) as session:
        ensure_bootstrap_users(session)


@app.get("/setup", response_class=HTMLResponse)
def setup_page(request: Request, session: Session = Depends(get_session)):
    users = session.exec(select(User).order_by(User.username.asc())).all()
    can_bootstrap = len(users) == 0
    return templates.TemplateResponse(
        "setup.html",
        {
            "request": request,
            "user": None,
            "users": users,
            "can_bootstrap": can_bootstrap,
        },
    )


@app.post("/setup/create_user")
def setup_create_user(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    session: Session = Depends(get_session),
):
    users = session.exec(select(User)).all()
    if users:
        return RedirectResponse("/login?setup=locked", status_code=303)

    uname = (username or "").strip()
    if not uname or not password:
        return RedirectResponse("/setup?err=1", status_code=303)

    session.add(User(username=uname[:80], password_hash=hash_password(password)))
    session.commit()
    return RedirectResponse("/login?setup=ok", status_code=303)


PUBLIC_PATHS = {"/login", "/health", "/version", "/docs", "/openapi.json", "/redoc", "/setup", "/setup/create_user"}
PUBLIC_PREFIXES = ("/static", "/uploads")


@app.middleware("http")
async def auth_guard(request: Request, call_next):
    path = request.url.path

    if path in PUBLIC_PATHS or path.startswith(PUBLIC_PREFIXES):
        return await call_next(request)

    if request.method == "HEAD" and path == "/":
        return Response(status_code=200)

    with Session(engine) as session:
        user = get_current_user(request, session)

    if not user:
        return RedirectResponse("/login?err=1", status_code=303)

    return await call_next(request)


@app.get("/version", response_class=PlainTextResponse)
def version():
    return APP_VERSION


@app.get("/system/status")
def system_status(request: Request, session: Session = Depends(get_session)):
    _ = get_current_user(request, session)
    users_count = len(session.exec(select(User)).all())
    invoices_count = len(session.exec(select(Invoice)).all())
    return JSONResponse(
        content={
            "version": APP_VERSION,
            "openai_enabled": _openai_enabled(),
            "database": engine.dialect.name,
            "users_count": users_count,
            "invoices_count": invoices_count,
            "upload_dir": UPLOAD_DIR,
        }
    )


@app.get("/health", response_class=PlainTextResponse)
@app.head("/health")
def health():
    return "OK"


@app.head("/")
def head_root():
    return Response(status_code=200)


# ---------------- LOGIN / LOGOUT ----------------

@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request):
    return templates.TemplateResponse("login.html", {"request": request, "user": None})


@app.post("/login")
def login(username: str = Form(...), password: str = Form(...), session: Session = Depends(get_session)):
    u = session.exec(select(User).where(User.username == username.strip())).first()
    if not u or not verify_password(password, u.password_hash):
        return RedirectResponse("/login", status_code=303)

    resp = RedirectResponse("/", status_code=303)
    resp.set_cookie(
        "uid",
        str(u.id),
        httponly=True,
        samesite="lax",
        secure=COOKIE_SECURE,
        path="/",
        max_age=60 * 60 * 24 * 30,
    )
    return resp


@app.get("/logout")
def logout():
    resp = RedirectResponse("/login", status_code=303)
    resp.delete_cookie("uid", path="/")
    return resp


# ---------------- OPENAI VISION HELPERS ----------------

def _bytes_to_data_url(file_bytes: bytes, mime: str) -> str:
    b64 = base64.b64encode(file_bytes).decode("utf-8")
    return f"data:{mime};base64,{b64}"


def _pdf_first_page_to_png_bytes(pdf_path: str) -> bytes:
    pdf = pdfium.PdfDocument(pdf_path)
    page = pdf.get_page(0)
    pil_image = page.render(scale=2).to_pil()
    page.close()
    pdf.close()
    buf = BytesIO()
    pil_image.save(buf, format="PNG")
    return buf.getvalue()


def _parse_date(s: Optional[str]) -> Optional[datetime]:
    if s is None or s == "":
        return None
    if isinstance(s, datetime):
        return s
    try:
        import datetime as _dt
        if isinstance(s, _dt.date):
            return datetime(s.year, s.month, s.day)
    except Exception:
        pass
    s = str(s).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            pass
    try:
        return datetime.fromisoformat(s.replace("Z", ""))
    except Exception:
        return None


def _normalize_currency(cur: Optional[str]) -> Optional[str]:
    if not cur:
        return None
    c = cur.strip().upper()
    if c in ("DH", "DHS", "DIRHAM", "DIRHAMS", "MAD", "M.A.D"):
        return "MAD"
    if c in ("€", "EUR"):
        return "EUR"
    if c in ("$", "USD", "US$"):
        return "USD"
    return c[:10]


def _to_int(x: Optional[str]) -> Optional[int]:
    if x is None:
        return None
    s = str(x).strip()
    if not s:
        return None
    try:
        return int(float(s))
    except Exception:
        return None


def _to_float(x: Optional[str]) -> Optional[float]:
    if x is None:
        return None
    s = str(x).strip()
    if not s:
        return None
    s = s.replace(" ", "")
    s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None

# ---------------- AFFECTATIONS helpers ----------------

def _invoice_amount_for_alloc(inv: Invoice) -> float:
    """Montant de référence pour les affectations (TTC si dispo, sinon amount)."""
    try:
        if getattr(inv, "amount_ttc", None) not in (None, 0, 0.0):
            return float(inv.amount_ttc)  # type: ignore
    except Exception:
        pass
    try:
        return float(inv.amount or 0.0)  # type: ignore
    except Exception:
        return 0.0


def _invoice_is_allocatable(session: Session, inv: Invoice) -> bool:
    """Affectable uniquement si la facture a un BR ou une FACTURE (pas BC only)."""
    try:
        doc = session.exec(
            select(InvoiceDocument).where(
                (InvoiceDocument.invoice_id == inv.id)
                & (InvoiceDocument.doc_type.in_(["BR", "FACTURE"]))
            )
        ).first()
        if doc:
            return True
    except Exception:
        pass

    # fallback: si BR renseigné
    if getattr(inv, "br_no", None):
        return True

    return False


def _alloc_sum(session: Session, invoice_id: int, kind: str) -> float:
    rows = session.exec(
        select(AllocationLine).where(
            (AllocationLine.invoice_id == invoice_id) & (AllocationLine.kind == kind)
        )
    ).all()
    return float(sum((r.amount or 0.0) for r in rows))



# ---------------- CIRCULAIRE CALC ----------------

def _end_of_month(d: date) -> date:
    if d.month == 12:
        return date(d.year, 12, 31)
    first_next = date(d.year, d.month + 1, 1)
    return first_next - timedelta(days=1)


def _next_business_day(d: date) -> date:
    while d.weekday() >= 5:
        d = d + timedelta(days=1)
    return d


def compute_pay_delay_fields(inv: Invoice):
    """Calcule l'échéance légale selon la logique DGI.

    - délai convenu: payment_terms_days (max 120) ou dérogation (max 180)
    - à défaut: 60 jours
    - point de départ principal: date d'émission de la facture
    - à défaut: fin du mois de la livraison / exécution / prestation
    - si l'échéance tombe un week-end: report au prochain jour ouvrable
    """
    if inv.derogation_sector and inv.derogation_days:
        terms = min(int(inv.derogation_days), 180)
    elif inv.payment_terms_days:
        terms = min(int(inv.payment_terms_days), 120)
    else:
        terms = 60
    inv.applied_terms_days = terms

    start_d: Optional[date] = None

    # La circulaire prend d'abord la date de facture; à défaut on retient la fin du mois de la prestation/livraison.
    if inv.invoice_date:
        inv.calc_start_date = inv.invoice_date
        inv.calc_start_rule = "DATE_FACTURE"
        start_d = inv.invoice_date.date()
    else:
        anchor_dt = getattr(inv, "service_date", None) or getattr(inv, "br_date", None) or getattr(inv, "bc_date", None)
        if anchor_dt:
            eom = _end_of_month(anchor_dt.date())
            inv.calc_start_date = datetime.combine(eom, datetime.min.time())
            inv.calc_start_rule = "FIN_MOIS_PRESTATION"
            start_d = eom

    if not start_d:
        inv.calc_start_date = None
        inv.calc_start_rule = None
        inv.legal_due_date = None
        return

    due = start_d + timedelta(days=terms)
    due = _next_business_day(due)
    inv.legal_due_date = datetime.combine(due, datetime.min.time())


# ---------------- PDF MERGE (DOSSIER EN 1 PDF) ----------------

def _image_path_to_pdf_bytes(img_path: str) -> bytes:
    im = Image.open(img_path)
    if im.mode in ("RGBA", "P"):
        im = im.convert("RGB")
    buf = BytesIO()
    im.save(buf, format="PDF")
    return buf.getvalue()


def _merge_paths_to_one_pdf_bytes(paths: list[str]) -> bytes:
    if not PDF_MERGER_AVAILABLE or PdfMerger is None:
        raise RuntimeError(
            "Fusion PDF indisponible: installe `pypdf>=4.2.0` (recommandé) "
            "ou `PyPDF2>=3.0.0` dans requirements.txt."
        )

    merger = PdfMerger()
    try:
        for p in paths:
            ext = os.path.splitext(p)[1].lower()
            if ext == ".pdf":
                merger.append(p)
            else:
                pdf_bytes = _image_path_to_pdf_bytes(p)
                merger.append(BytesIO(pdf_bytes))

        out = BytesIO()
        merger.write(out)
        return out.getvalue()
    finally:
        try:
            merger.close()
        except Exception:
            pass


# ---------------- OPENAI EXTRACT ----------------


def extract_invoice_with_openai(file_path: str, original_mime: str) -> dict:
    if not _openai_enabled():
        return _fallback_extract_doc_data(file_path, original_mime, detailed_invoice=True)

    if original_mime == "application/pdf" or file_path.lower().endswith(".pdf"):
        img_bytes = _pdf_first_page_to_png_bytes(file_path)
        img_mime = "image/png"
    else:
        with open(file_path, "rb") as f:
            img_bytes = f.read()
        img_mime = original_mime if (original_mime or "").startswith("image/") else "image/jpeg"

    data_url = _bytes_to_data_url(img_bytes, img_mime)

    schema = {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "supplier_name": {"type": ["string", "null"]},
            "supplier_if": {"type": ["string", "null"]},
            "supplier_ice": {"type": ["string", "null"]},
            "supplier_rc": {"type": ["string", "null"]},
            "supplier_rc_city": {"type": ["string", "null"]},
            "supplier_address": {"type": ["string", "null"]},
            "invoice_no": {"type": ["string", "null"]},
            "invoice_date": {"type": ["string", "null"], "description": "YYYY-MM-DD"},
            "service_date": {"type": ["string", "null"], "description": "YYYY-MM-DD"},
            "nature_operation": {"type": ["string", "null"], "description": "BIENS/SERVICES/TRAVAUX"},
            "payment_terms_days": {"type": ["number", "null"]},
            "amount_ht": {"type": ["number", "null"]},
            "vat_rate": {"type": ["number", "null"], "description": "ex 20 for 20%"},
            "amount_vat": {"type": ["number", "null"]},
            "amount_ttc": {"type": ["number", "null"], "description": "TOTAL TTC / NET A PAYER"},
            "currency": {"type": ["string", "null"]},
            "department": {"type": ["string", "null"]},
            "analytic": {"type": ["string", "null"]},
            "bc_no": {"type": ["string", "null"]},
            "due_date": {"type": ["string", "null"], "description": "YYYY-MM-DD"},
            "due_date_planned": {"type": ["string", "null"], "description": "YYYY-MM-DD"},
            "due_date_agreed": {"type": ["string", "null"], "description": "YYYY-MM-DD"},
        },
        "required": [
            "supplier_name", "supplier_if", "supplier_ice", "supplier_rc", "supplier_rc_city", "supplier_address",
            "invoice_no", "invoice_date",
            "service_date", "nature_operation",
            "payment_terms_days",
            "amount_ht", "vat_rate", "amount_vat", "amount_ttc", "currency",
            "department", "analytic", "bc_no",
            "due_date", "due_date_planned", "due_date_agreed",
        ],
    }

    try:
        return _call_openai_json(
            model="gpt-4o",
            input_payload=[
                {
                    "role": "system",
                    "content": (
                        "Tu es un expert en lecture de factures (Maroc/FR). "
                        "Retourne STRICTEMENT un JSON conforme au schéma (pas de texte). "
                        "Priorité montants: TOTAL TTC / NET A PAYER / GRAND TOTAL. "
                        "Dates au format YYYY-MM-DD. Si absent: null. "
                        "Identifiants fournisseur: ICE, IF, RC si visibles."
                    ),
                },
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "input_text",
                            "text": (
                                "Extrait les champs demandés.\n"
                                "- amount_ttc = TOTAL TTC / NET A PAYER / GRAND TOTAL.\n"
                                "- amount_ht, amount_vat et vat_rate si visibles.\n"
                                "- service_date = date livraison/service fait si visible.\n"
                                "- payment_terms_days si une mention de délai existe (30j, 60j...).\n"
                                "Si non visible: null."
                            ),
                        },
                        {"type": "input_image", "image_url": data_url},
                    ],
                },
            ],
            schema_name="invoice_extract_v2",
            schema=schema,
            max_output_tokens=1200,
        )
    except Exception:
        return _fallback_extract_doc_data(file_path, original_mime, detailed_invoice=True)
def _safe_store_upload(content: bytes, original_name: str) -> str:
    ts = int(datetime.now().timestamp())
    original = original_name or "upload"
    safe_name = f"{ts}_{original}".replace("/", "_").replace("\\", "_")
    disk_path = os.path.join(UPLOAD_DIR, safe_name)
    with open(disk_path, "wb") as f:
        f.write(content)
    return safe_name



def extract_doc_bundle_with_openai(file_path: str, original_mime: str) -> dict:
    if not _openai_enabled():
        return _fallback_extract_doc_data(file_path, original_mime, detailed_invoice=False)

    if original_mime == "application/pdf" or file_path.lower().endswith(".pdf"):
        img_bytes = _pdf_first_page_to_png_bytes(file_path)
        img_mime = "image/png"
    else:
        with open(file_path, "rb") as f:
            img_bytes = f.read()
        img_mime = original_mime if (original_mime or "").startswith("image/") else "image/jpeg"

    data_url = _bytes_to_data_url(img_bytes, img_mime)

    schema = {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "doc_type": {"type": "string", "enum": ["FACTURE", "BC", "BR", "AUTRE"]},
            "supplier_name": {"type": ["string", "null"]},
            "supplier_ice": {"type": ["string", "null"]},
            "supplier_if": {"type": ["string", "null"]},
            "supplier_rc": {"type": ["string", "null"]},
            "supplier_rc_city": {"type": ["string", "null"]},
            "supplier_address": {"type": ["string", "null"]},
            "invoice_no": {"type": ["string", "null"]},
            "invoice_date": {"type": ["string", "null"], "description": "YYYY-MM-DD"},
            "amount_ttc": {"type": ["number", "null"]},
            "currency": {"type": ["string", "null"]},
            "payment_terms_days": {"type": ["number", "null"]},
            "bc_no": {"type": ["string", "null"]},
            "bc_date": {"type": ["string", "null"], "description": "YYYY-MM-DD"},
            "br_no": {"type": ["string", "null"]},
            "br_date": {"type": ["string", "null"], "description": "YYYY-MM-DD"},
            "service_date": {"type": ["string", "null"], "description": "YYYY-MM-DD"},
        },
        "required": [
            "doc_type",
            "supplier_name", "supplier_ice", "supplier_if", "supplier_rc", "supplier_rc_city", "supplier_address",
            "invoice_no", "invoice_date", "amount_ttc", "currency", "payment_terms_days",
            "bc_no", "bc_date", "br_no", "br_date", "service_date",
        ],
    }

    try:
        return _call_openai_json(
            model="gpt-4o",
            input_payload=[
                {
                    "role": "system",
                    "content": (
                        "Tu es un assistant ERP. "
                        "1) Identifie le type du document: FACTURE, BC, BR, AUTRE. "
                        "2) Extrait les champs visibles. "
                        "Retourne STRICTEMENT un JSON conforme au schéma."
                    ),
                },
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "input_text",
                            "text": (
                                "Classe le document et extrait les champs.\n"
                                "- FACTURE: invoice_no, invoice_date, amount_ttc, devise, fournisseur.\n"
                                "- BC: bc_no, bc_date, fournisseur.\n"
                                "- BR: br_no, br_date ou service_date.\n"
                                "Si absent => null."
                            ),
                        },
                        {"type": "input_image", "image_url": data_url},
                    ],
                },
            ],
            schema_name="doc_bundle_extract_v1",
            schema=schema,
            max_output_tokens=800,
        )
    except Exception:
        return _fallback_extract_doc_data(file_path, original_mime, detailed_invoice=False)

def extract_rib_with_openai(file_path: str, original_mime: str) -> dict:
    if not _openai_enabled():
        return _fallback_extract_doc_data(file_path, original_mime, rib_mode=True)

    if original_mime == "application/pdf" or file_path.lower().endswith(".pdf"):
        img_bytes = _pdf_first_page_to_png_bytes(file_path)
        img_mime = "image/png"
    else:
        with open(file_path, "rb") as f:
            img_bytes = f.read()
        img_mime = original_mime if (original_mime or "").startswith("image/") else "image/jpeg"

    data_url = _bytes_to_data_url(img_bytes, img_mime)

    schema = {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "beneficiary_name": {"type": ["string", "null"]},
            "bank_name": {"type": ["string", "null"]},
            "rib_or_iban": {"type": ["string", "null"]},
            "swift": {"type": ["string", "null"]},
        },
        "required": ["beneficiary_name", "bank_name", "rib_or_iban", "swift"],
    }

    try:
        return _call_openai_json(
            model="gpt-4o",
            input_payload=[
                {"role": "system", "content": "Tu lis une attestation RIB/IBAN. Retourne STRICTEMENT un JSON conforme au schéma."},
                {
                    "role": "user",
                    "content": [
                        {"type": "input_text", "text": "Extrait: bénéficiaire, banque, RIB/IBAN, SWIFT si visible."},
                        {"type": "input_image", "image_url": data_url},
                    ],
                },
            ],
            schema_name="rib_extract_v1",
            schema=schema,
            max_output_tokens=400,
        )
    except Exception:
        return _fallback_extract_doc_data(file_path, original_mime, rib_mode=True)
def _merge_extraction_into_invoice(inv: Invoice, data: dict):
    dt = (data.get("doc_type") or "AUTRE").strip().upper()

    if not inv.supplier_name and data.get("supplier_name"):
        inv.supplier_name = str(data.get("supplier_name")).strip()[:200] or inv.supplier_name
    if not inv.supplier_ice and data.get("supplier_ice"):
        inv.supplier_ice = str(data.get("supplier_ice")).strip()[:30] or inv.supplier_ice
    if not inv.supplier_if and data.get("supplier_if"):
        inv.supplier_if = str(data.get("supplier_if")).strip()[:30] or inv.supplier_if
    if not inv.supplier_rc and data.get("supplier_rc"):
        inv.supplier_rc = str(data.get("supplier_rc")).strip()[:30] or inv.supplier_rc
    if not inv.supplier_rc_city and data.get("supplier_rc_city"):
        inv.supplier_rc_city = str(data.get("supplier_rc_city")).strip()[:60] or inv.supplier_rc_city
    if not inv.supplier_address and data.get("supplier_address"):
        inv.supplier_address = str(data.get("supplier_address")).strip()[:250] or inv.supplier_address

    if dt == "FACTURE":
        if not inv.invoice_no and data.get("invoice_no"):
            inv.invoice_no = str(data.get("invoice_no")).strip()[:80] or inv.invoice_no
        if not inv.invoice_date and data.get("invoice_date"):
            inv.invoice_date = _parse_date(data.get("invoice_date"))

        if (inv.amount_ttc is None) and (data.get("amount_ttc") is not None):
            inv.amount_ttc = _to_float(data.get("amount_ttc"))

        if not inv.currency and data.get("currency"):
            inv.currency = _normalize_currency(data.get("currency")) or inv.currency

        if (inv.payment_terms_days is None) and (data.get("payment_terms_days") is not None):
            inv.payment_terms_days = _to_int(data.get("payment_terms_days"))

        if inv.amount_ttc is not None:
            inv.amount = float(inv.amount_ttc)

    elif dt == "BC":
        if not inv.bc_no and data.get("bc_no"):
            inv.bc_no = str(data.get("bc_no")).strip()[:80] or inv.bc_no
        if not getattr(inv, "bc_date", None) and data.get("bc_date"):
            inv.bc_date = _parse_date(data.get("bc_date"))

    elif dt == "BR":
        if not getattr(inv, "br_no", None) and data.get("br_no"):
            inv.br_no = str(data.get("br_no")).strip()[:80] or inv.br_no
        if not getattr(inv, "br_date", None) and data.get("br_date"):
            inv.br_date = _parse_date(data.get("br_date"))

        sd = _parse_date(data.get("service_date")) or _parse_date(data.get("br_date"))
        if not inv.service_date and sd:
            inv.service_date = sd

    compute_pay_delay_fields(inv)


def _invoice_amount(inv: Invoice) -> float:
    for cand in (getattr(inv, "amount_ttc", None), getattr(inv, "amount", None)):
        try:
            if cand is not None:
                return float(cand or 0.0)
        except Exception:
            pass
    return 0.0


def _invoice_due_date(inv: Invoice) -> Optional[date]:
    for cand in (
        getattr(inv, "legal_due_date", None),
        getattr(inv, "due_date_agreed", None),
        getattr(inv, "due_date_planned", None),
        getattr(inv, "due_date", None),
        getattr(inv, "invoice_date", None),
    ):
        try:
            if cand:
                return cand.date()
        except Exception:
            pass
    return None


def _invoice_quality_flags(inv: Invoice) -> list[str]:
    flags: list[str] = []
    if not (getattr(inv, "supplier_name", None) or getattr(inv, "supplier_id", None)):
        flags.append("fournisseur manquant")
    if not getattr(inv, "invoice_no", None):
        flags.append("n° facture manquant")
    if not _invoice_amount(inv):
        flags.append("montant manquant")
    if not _invoice_due_date(inv):
        flags.append("échéance manquante")
    if getattr(inv, "is_disputed", False):
        flags.append("litige")
    return flags


def _invoice_priority_snapshot(inv: Invoice, today: date) -> dict:
    due = _invoice_due_date(inv)
    amount = _invoice_amount(inv)
    cur = ((getattr(inv, "currency", None) or "MAD").strip().upper() or "MAD")
    flags = _invoice_quality_flags(inv)
    days_left = None if not due else (due - today).days

    score = 20
    reasons: list[str] = []

    if days_left is None:
        score += 12
        reasons.append("échéance à fiabiliser")
    elif days_left < 0:
        score += 60 + min(abs(days_left), 30)
        reasons.append("retard")
    elif days_left <= 3:
        score += 42
        reasons.append("urgence 72h")
    elif days_left <= 7:
        score += 28
        reasons.append("à payer sous 7 jours")
    elif days_left <= 15:
        score += 14
        reasons.append("à préparer")

    if amount >= 100000:
        score += 16
        reasons.append("montant élevé")
    elif amount >= 50000:
        score += 10
        reasons.append("montant significatif")
    elif amount >= 10000:
        score += 4

    if flags:
        score += min(18, 6 * len(flags))
        reasons.append("données incomplètes")

    if (getattr(inv, "status", "") or "").upper() in ("ANNULEE", "PAYEE"):
        score = 0
        reasons = ["déjà traitée"]

    if score >= 95:
        label = "CRITIQUE"
    elif score >= 70:
        label = "HAUTE"
    elif score >= 45:
        label = "MOYENNE"
    else:
        label = "NORMALE"

    return {
        "invoice": inv,
        "invoice_id": getattr(inv, "id", None),
        "supplier_name": getattr(inv, "supplier_name", None) or "—",
        "invoice_no": getattr(inv, "invoice_no", None) or "—",
        "currency": cur,
        "amount": amount,
        "due": due,
        "days_left": days_left,
        "flags": flags,
        "score": int(score),
        "label": label,
        "reasons": reasons,
    }


def _latest_bank_balances(session: Session) -> dict[str, dict]:
    rows = session.exec(
        select(BankTxn).order_by(BankTxn.value_date.desc().nullslast(), BankTxn.date.desc(), BankTxn.id.desc())
    ).all()
    latest: dict[str, dict] = {}
    for t in rows:
        key = f"{(t.bank_name or '').strip()}|{(t.account_no or '').strip()}"
        if key in latest:
            continue
        bal = None
        try:
            bal = float(t.balance) if t.balance is not None else None
        except Exception:
            bal = None
        latest[key] = {
            "bank_name": t.bank_name or "",
            "account_no": t.account_no or "",
            "currency": (t.currency or "MAD").upper(),
            "balance": bal,
            "as_of": (t.value_date or t.date),
        }
    return latest


def _recommend_company_account(accounts: list[CompanyBankAccount], bank_balances: dict[str, dict], currency: str, amount: float) -> Optional[dict]:
    if not accounts:
        return None

    currency = (currency or "MAD").upper()
    candidates: list[dict] = []
    for acc in accounts:
        key = f"{(acc.bank_name or '').strip()}|{(acc.account_no or '').strip()}"
        bal_info = bank_balances.get(key, {})
        bal_cur = (bal_info.get("currency") or currency).upper()
        bal = bal_info.get("balance")
        fit = 0
        if bal_cur == currency:
            fit += 40
        if bal is not None:
            if bal >= amount > 0:
                fit += 30
            elif bal > 0:
                fit += 12
        if getattr(acc, "is_active", False):
            fit += 10
        candidates.append(
            {
                "account": acc,
                "bank_name": acc.bank_name,
                "account_no": acc.account_no,
                "balance": bal,
                "balance_currency": bal_cur,
                "score": fit,
            }
        )

    candidates.sort(key=lambda x: (x["score"], x["balance"] if x["balance"] is not None else -10**12), reverse=True)
    return candidates[0] if candidates else None


def _candidate_invoices_for_txn(txn: BankTxn, open_invoices: list[Invoice]) -> list[Invoice]:
    txn_amt = round(_txn_amount(txn), 2)
    candidates: list[Invoice] = []
    for inv in open_invoices:
        inv_amt = round(_invoice_amount(inv), 2)
        if inv_amt <= 0:
            continue
        if abs(inv_amt - txn_amt) <= 0.01 or abs(inv_amt - txn_amt) <= 5.0:
            candidates.append(inv)
    return candidates


def _build_automation_snapshot(session: Session) -> dict:
    today = date.today()
    accounts = session.exec(
        select(CompanyBankAccount)
        .where(CompanyBankAccount.is_active == True)
        .order_by(CompanyBankAccount.bank_name.asc())
    ).all()
    bank_balances = _latest_bank_balances(session)

    open_invoices = session.exec(select(Invoice).where(Invoice.status == "A_PAYER")).all()
    priority_rows = [_invoice_priority_snapshot(inv, today) for inv in open_invoices]
    priority_rows.sort(
        key=lambda r: (
            -r["score"],
            r["days_left"] if r["days_left"] is not None else 10**6,
            -r["amount"],
        )
    )

    for row in priority_rows:
        row["recommended_account"] = _recommend_company_account(accounts, bank_balances, row["currency"], row["amount"])
        row["completion"] = _invoice_completion_snapshot(row["invoice"]) if row.get("invoice") else {"score": 0, "missing": []}

    matched_txn_ids = {m.banktxn_id for m in session.exec(select(InvoicePaymentMatch.banktxn_id)).all()}
    txns = session.exec(
        select(BankTxn)
        .where(BankTxn.debit > 0)
        .order_by(BankTxn.value_date.desc().nullslast(), BankTxn.date.desc(), BankTxn.id.desc())
        .limit(160)
    ).all()
    unmatched_txns = [t for t in txns if t.id not in matched_txn_ids]

    high_conf_matches: list[dict] = []
    for txn in unmatched_txns[:80]:
        candidates = _candidate_invoices_for_txn(txn, open_invoices)
        suggestions = _suggest_for_txn(txn, candidates, limit=3)
        if not suggestions:
            continue
        first = suggestions[0]
        second_score = suggestions[1]["score"] if len(suggestions) > 1 else 0
        if first["score"] >= 90 and first["score"] - second_score >= 8:
            high_conf_matches.append(
                {
                    "txn": txn,
                    "top": first,
                    "alternatives": suggestions[1:],
                }
            )

    action_buckets = {
        "critical": [r for r in priority_rows if r["label"] == "CRITIQUE"][:12],
        "high": [r for r in priority_rows if r["label"] == "HAUTE"][:12],
        "missing_data": [r for r in priority_rows if r["flags"]][:12],
        "recommended_now": [r for r in priority_rows if r["days_left"] is not None and r["days_left"] <= 7][:12],
    }

    return {
        "today": today,
        "priority_rows": priority_rows,
        "action_buckets": action_buckets,
        "high_conf_matches": high_conf_matches,
        "bank_balances": bank_balances,
        "accounts": accounts,
    }



# ---------------- HOME ----------------

@app.get("/", response_class=HTMLResponse)
def home(request: Request, session: Session = Depends(get_session)):
    user = get_current_user(request, session)

    # ---------------- Dashboard analytics (style C) ----------------
    today = datetime.utcnow().date()
    horizon_90 = today + timedelta(days=90)

    # Load invoices + suppliers (for "étrangers" exclusion in délais)
    invs = session.exec(select(Invoice)).all()
    suppliers = {s.id: s for s in session.exec(select(Supplier)).all()}

    def _cur(x: Optional[str]) -> str:
        c = (x or "MAD").strip().upper()
        return c or "MAD"

    def _is_foreign(inv: Invoice) -> bool:
        sid = getattr(inv, "supplier_id", None)
        if not sid:
            return False
        s = suppliers.get(sid)
        return bool(getattr(s, "is_foreign", False))

    # KPIs (amounts by currency)
    kpi = {
        "pay_7": {},
        "pay_30": {},
        "pay_90": {},
        "overdue": {},
        "total_open": 0,
        "foreign_excluded": 0,
        "missing_due": 0,
        "missing_supplier": 0,
    }

    # For bank allocation: use PaymentLine -> PaymentBatch.company_account_id when available
    # Fallback: if exactly one active account exists, allocate all to it.
    accounts = session.exec(select(CompanyBankAccount)).all()
    active_accounts = [a for a in accounts if getattr(a, "is_active", True)]
    default_account_id = active_accounts[0].id if len(active_accounts) == 1 else None

    # Build invoice -> account_id mapping via payment lines
    inv_to_account: dict[int, Optional[int]] = {}
    lines = session.exec(select(PaymentLine)).all()
    if lines:
        batch_ids = list({ln.batch_id for ln in lines if ln.batch_id})
        batches = {}
        if batch_ids:
            for b in session.exec(select(PaymentBatch).where(PaymentBatch.id.in_(batch_ids))).all():
                batches[b.id] = b
        for ln in lines:
            if not ln.invoice_id:
                continue
            b = batches.get(ln.batch_id)
            aid = getattr(b, "company_account_id", None) if b else None
            if aid:
                inv_to_account[ln.invoice_id] = aid

    def _add_amount(bucket: dict, cur: str, amt: float):
        bucket[cur] = float(bucket.get(cur, 0.0)) + float(amt or 0.0)

    upcoming = []
    top_open = []

    bank_proj: dict[int, dict] = {}
    for a in active_accounts:
        bank_proj[a.id] = {"account": a, "cur_bal": None, "out_7": 0.0, "out_30": 0.0, "out_90": 0.0, "currencies": set()}

    # Last balance per account
    txns = session.exec(select(BankTxn).order_by(BankTxn.date.desc()).limit(5000)).all()
    last_bal_by_acct: dict[str, float] = {}
    for t in txns:
        if t.balance is None:
            continue
        key = (t.account_no or "").strip()
        if key and key not in last_bal_by_acct:
            last_bal_by_acct[key] = float(t.balance)

    for a in active_accounts:
        acct_no = (getattr(a, "account_no", None) or "").strip()
        if acct_no and acct_no in last_bal_by_acct:
            bank_proj[a.id]["cur_bal"] = last_bal_by_acct[acct_no]

    for inv in invs:
        status = (getattr(inv, "status", None) or "").upper()
        if status != "A_PAYER":
            continue

        kpi["total_open"] += 1

        cur = _cur(getattr(inv, "currency", None))
        amt = float(getattr(inv, "amount_ttc", None) or getattr(inv, "amount", 0.0) or 0.0)

        if _is_foreign(inv):
            kpi["foreign_excluded"] += 1

        due_dt = getattr(inv, "legal_due_date", None)
        if not due_dt:
            kpi["missing_due"] += 1
            continue
        try:
            due = due_dt.date()
        except Exception:
            continue

        if not getattr(inv, "supplier_id", None):
            kpi["missing_supplier"] += 1

        if due < today:
            _add_amount(kpi["overdue"], cur, amt)
        if due <= today + timedelta(days=7):
            _add_amount(kpi["pay_7"], cur, amt)
        if due <= today + timedelta(days=30):
            _add_amount(kpi["pay_30"], cur, amt)
        if due <= horizon_90:
            _add_amount(kpi["pay_90"], cur, amt)

        if not _is_foreign(inv):
            upcoming.append({
                "due": due.isoformat(),
                "supplier": getattr(inv, "supplier_name", None) or "—",
                "ref": getattr(inv, "invoice_no", None) or f"ACHAT #{inv.id}",
                "amount": amt,
                "currency": cur,
                "id": inv.id,
            })
            top_open.append({
                "supplier": getattr(inv, "supplier_name", None) or "—",
                "ref": getattr(inv, "invoice_no", None) or f"ACHAT #{inv.id}",
                "amount": amt,
                "currency": cur,
                "due": due.isoformat(),
                "id": inv.id,
            })

        aid = inv_to_account.get(inv.id) or default_account_id
        if aid and aid in bank_proj:
            bank_proj[aid]["currencies"].add(cur)
            if cur == "MAD":
                if due <= today + timedelta(days=7):
                    bank_proj[aid]["out_7"] += amt
                if due <= today + timedelta(days=30):
                    bank_proj[aid]["out_30"] += amt
                if due <= horizon_90:
                    bank_proj[aid]["out_90"] += amt

    upcoming = sorted(upcoming, key=lambda x: x["due"])[:8]
    top_open = sorted(top_open, key=lambda x: x["amount"], reverse=True)[:8]

    bank_rows: list[dict] = []
    for a in active_accounts:
        d = bank_proj[a.id]
        cur_bal = d["cur_bal"]
        out_7 = d["out_7"]
        out_30 = d["out_30"]
        out_90 = d["out_90"]
        proj_30 = (cur_bal - out_30) if cur_bal is not None else None
        proj_90 = (cur_bal - out_90) if cur_bal is not None else None
        alert = None
        if proj_30 is not None and proj_30 < 0:
            alert = "DEFICIT -30j"
        elif proj_90 is not None and proj_90 < 0:
            alert = "DEFICIT -90j"
        bank_rows.append({
            "bank": getattr(a, "bank_name", None) or "Banque",
            "account": getattr(a, "account_no", None) or "—",
            "cur_bal": cur_bal,
            "out_7": out_7,
            "out_30": out_30,
            "out_90": out_90,
            "proj_30": proj_30,
            "proj_90": proj_90,
            "alert": alert,
            "multi_currency": (len(d["currencies"]) > 1) or (len(d["currencies"]) == 1 and (list(d["currencies"])[0] != "MAD")),
        })

    total_local = 0
    conforme_local = 0
    for inv in invs:
        if (getattr(inv, "status", "") or "").upper() != "A_PAYER":
            continue
        if _is_foreign(inv):
            continue
        total_local += 1
        if getattr(inv, "legal_due_date", None):
            conforme_local += 1

    compliance_pct = (round((conforme_local / total_local) * 100) if total_local else 100)

    def _sum_cur(bucket: dict, cur: str) -> float:
        return float(bucket.get(cur, 0.0) or 0.0)

    # ---------------- Chart data (by currency) ----------------
    # Dashboard chart is switchable (MAD/EUR/USD). If a currency has no data,
    # the frontend will hide it.

    def _inv_amount(inv: Invoice) -> float:
        return float(getattr(inv, "amount_ttc", None) or getattr(inv, "amount", 0.0) or 0.0)

    # Precompute pay <= 60 days by currency
    pay_60_by_cur: dict[str, float] = {}
    for inv in invs:
        if (getattr(inv, "status", "") or "").upper() != "A_PAYER":
            continue
        due_dt = getattr(inv, "legal_due_date", None)
        if not due_dt:
            continue
        try:
            due = due_dt.date()
        except Exception:
            continue
        if due <= today + timedelta(days=60):
            c = _cur(getattr(inv, "currency", None))
            pay_60_by_cur[c] = float(pay_60_by_cur.get(c, 0.0)) + _inv_amount(inv)

    chart_by_cur: dict[str, dict] = {}
    cur_set = {*(kpi["pay_7"].keys()), *(kpi["pay_30"].keys()), *(kpi["pay_90"].keys()), *(kpi["overdue"].keys())}
    if not cur_set:
        cur_set = {"MAD"}

    for c in sorted(cur_set):
        chart_by_cur[c] = {
            "labels": ["7 jours", "30 jours", "60 jours", "90 jours"],
            "a_payer": [
                _sum_cur(kpi["pay_7"], c),
                _sum_cur(kpi["pay_30"], c),
                float(pay_60_by_cur.get(c, 0.0) or 0.0),
                _sum_cur(kpi["pay_90"], c),
            ],
            # Retard = stock des factures en retard (répété pour lecture).
            "retard": [
                _sum_cur(kpi["overdue"], c),
                _sum_cur(kpi["overdue"], c),
                _sum_cur(kpi["overdue"], c),
                _sum_cur(kpi["overdue"], c),
            ],
        }

    default_chart_cur = "MAD" if "MAD" in chart_by_cur else next(iter(chart_by_cur.keys()))

    alerts = []
    if _sum_cur(kpi["overdue"], "MAD") > 0:
        alerts.append("Factures en retard : régulariser ou justifier (litige / dérogation).")
    if kpi["missing_due"] > 0:
        alerts.append(f"{kpi['missing_due']} facture(s) sans échéance légale : compléter pour le pilotage.")
    if kpi["missing_supplier"] > 0:
        alerts.append(f"{kpi['missing_supplier']} facture(s) sans fournisseur : lier pour conformité délais.")
    if not active_accounts:
        alerts.append("Aucun compte société actif : ajouter au moins un compte (Paramètres → Comptes société).")

    automation = _build_automation_snapshot(session)

    payload = {
        "request": request,
        "user": user,
        "kpi": kpi,
        "upcoming": upcoming,
        "top_open": top_open,
        "bank_rows": bank_rows,
        "compliance_pct": compliance_pct,
        "chart_by_cur": chart_by_cur,
        "default_chart_cur": default_chart_cur,
        "alerts": alerts,
        "automation": automation,
    }

    return templates.TemplateResponse("home.html", payload)


@app.get("/automation", response_class=HTMLResponse)
def automation_hub(request: Request, session: Session = Depends(get_session)):
    user = get_current_user(request, session)
    automation = _build_automation_snapshot(session)
    return templates.TemplateResponse(
        "automation.html",
        {
            "request": request,
            "user": user,
            "automation": automation,
        },
    )


@app.get("/todo", response_class=HTMLResponse)
def todo_page(request: Request, session: Session = Depends(get_session)):
    user = get_current_user(request, session)
    automation = _build_automation_snapshot(session)
    quick_actions = [
        {
            "title": "Importer une nouvelle facture",
            "desc": "Commencez par déposer un PDF ou une photo. L'application tente de remplir automatiquement les champs.",
            "href": "/factures",
            "cta": "Ajouter une facture",
        },
        {
            "title": "Compléter les dossiers incomplets",
            "desc": f"{len(automation['action_buckets']['missing_data'])} dossier(s) demandent une vérification ou un complément.",
            "href": "/automation",
            "cta": "Voir les éléments à corriger",
        },
        {
            "title": "Préparer les paiements urgents",
            "desc": f"{len(automation['action_buckets']['critical'])} priorité(s) critique(s) et {len(automation['action_buckets']['high'])} haute(s).",
            "href": "/planning",
            "cta": "Ouvrir le planning",
        },
    ]
    return templates.TemplateResponse(
        "todo.html",
        {
            "request": request,
            "user": user,
            "automation": automation,
            "quick_actions": quick_actions,
        },
    )



# ---------------- ASSISTANT (Q/R DB) ----------------



def _assistant_db_answer(session: Session, q: str) -> dict:
    q = (q or "").strip()
    if not q:
        return {"answer_lines": [], "tables": {}, "q": ""}

    qn = q.lower().strip()
    today = datetime.utcnow().date()

    def fmt_amt(x: float) -> str:
        try:
            return "{:,.2f}".format(float(x)).replace(",", " ")
        except Exception:
            return str(x)

    answer_lines: list[str] = []
    tables: dict[str, list[dict]] = {}

    if any(k in qn for k in ["priorit", "priorité", "urgent", "urgente", "quoi payer", "paiement", "a payer aujourd", "à payer aujourd"]):
        automation = _build_automation_snapshot(session)
        critical = automation["action_buckets"]["critical"]
        high = automation["action_buckets"]["high"]
        answer_lines.append(f"Priorités détectées au {today.isoformat()} :")
        answer_lines.append(f"- Critiques : {len(critical)}")
        answer_lines.append(f"- Hautes : {len(high)}")
        if automation["high_conf_matches"]:
            answer_lines.append(f"- Matchings auto très fiables disponibles : {len(automation['high_conf_matches'])}")
        if critical:
            tables["priorities"] = [
                {
                    "id": r["invoice_id"],
                    "supplier": r["supplier_name"],
                    "invoice_no": r["invoice_no"],
                    "due": r["due"].isoformat() if r["due"] else "—",
                    "days_left": r["days_left"] if r["days_left"] is not None else "—",
                    "amount": f"{fmt_amt(r['amount'])} {r['currency']}",
                    "priority": r["label"],
                    "reasons": ", ".join(r["reasons"]) or "—",
                }
                for r in critical[:10]
            ]
            answer_lines.append("Top priorités prêtes à être traitées affichées dans le tableau.")
        else:
            answer_lines.append("Aucune priorité critique détectée.")

    elif any(k in qn for k in ["match", "matching", "rapprochement", "releve", "relevé automatique"]):
        automation = _build_automation_snapshot(session)
        rows = automation["high_conf_matches"]
        if rows:
            answer_lines.append(f"J'ai trouvé {len(rows)} rapprochements auto à forte confiance.")
            tables["matches"] = [
                {
                    "date": (r["txn"].value_date or r["txn"].date).date().isoformat(),
                    "label": (r["txn"].label or "")[:80],
                    "amount": fmt_amt(_txn_amount(r["txn"])),
                    "invoice": f'#{r["top"]["inv"].id} · {r["top"]["inv"].supplier_name or "—"} · {r["top"]["inv"].invoice_no or "—"}',
                    "score": r["top"]["score"],
                }
                for r in rows[:10]
            ]
        else:
            answer_lines.append("Aucun matching automatique à forte confiance n'a été détecté pour le moment.")

    elif any(k in qn for k in ["risque", "retard", "compliance", "conform", "bloqu", "bloqué"]):
        automation = _build_automation_snapshot(session)
        missing = automation["action_buckets"]["missing_data"]
        overdue = [r for r in automation["priority_rows"] if r["days_left"] is not None and r["days_left"] < 0][:10]
        answer_lines.append(f"Risques principaux au {today.isoformat()} :")
        answer_lines.append(f"- Factures en retard : {len([r for r in automation['priority_rows'] if r['days_left'] is not None and r['days_left'] < 0])}")
        answer_lines.append(f"- Factures avec données incomplètes : {len(missing)} (aperçu)")
        if overdue:
            tables["priorities"] = [
                {
                    "id": r["invoice_id"],
                    "supplier": r["supplier_name"],
                    "invoice_no": r["invoice_no"],
                    "due": r["due"].isoformat() if r["due"] else "—",
                    "days_left": r["days_left"],
                    "amount": f"{fmt_amt(r['amount'])} {r['currency']}",
                    "priority": r["label"],
                    "reasons": ", ".join(r["reasons"]) or "—",
                }
                for r in overdue
            ]

    elif any(k in qn for k in ["cette semaine", "7 jours", "semaine", "7j"]):
        invs = session.exec(select(Invoice).where((Invoice.status == "A_PAYER"))).all()
        sums: dict[str, float] = {}
        for inv in invs:
            due = _invoice_due_date(inv)
            if due and due <= today + timedelta(days=7):
                cur = (getattr(inv, "currency", None) or "MAD").upper().strip() or "MAD"
                sums[cur] = sums.get(cur, 0.0) + _invoice_amount(inv)
        if sums:
            answer_lines.append("Montant total à payer sur 7 jours (par devise) :")
            for cur, amt in sorted(sums.items()):
                answer_lines.append(f"- {fmt_amt(amt)} {cur}")
        else:
            answer_lines.append("Aucune échéance trouvée sur 7 jours.")

    elif any(k in qn for k in ["30 jours", "30j", "mois"]):
        invs = session.exec(select(Invoice).where((Invoice.status == "A_PAYER"))).all()
        sums: dict[str, float] = {}
        for inv in invs:
            due = _invoice_due_date(inv)
            if due and due <= today + timedelta(days=30):
                cur = (getattr(inv, "currency", None) or "MAD").upper().strip() or "MAD"
                sums[cur] = sums.get(cur, 0.0) + _invoice_amount(inv)
        if sums:
            answer_lines.append("Montant total à payer sur 30 jours (par devise) :")
            for cur, amt in sorted(sums.items()):
                answer_lines.append(f"- {fmt_amt(amt)} {cur}")
        else:
            answer_lines.append("Aucune échéance trouvée sur 30 jours.")

    elif any(k in qn for k in ["solde", "balance", "banque", "compte"]):
        balances = _latest_bank_balances(session)
        rows = []
        for key, info in balances.items():
            rows.append({
                "bank_name": info["bank_name"] or "Banque",
                "account_no": info["account_no"] or "—",
                "balance": "—" if info["balance"] is None else f'{fmt_amt(info["balance"])} {info["currency"]}',
                "as_of": info["as_of"].date().isoformat() if info["as_of"] else "—",
            })
        rows = rows[:20]
        tables["balances"] = rows
        answer_lines.append(f"J'ai trouvé {len(rows)} soldes bancaires récents.")

    elif any(k in qn for k in ["transfert", "virement", "transaction"]):
        txns = session.exec(select(BankTxn).order_by(BankTxn.date.desc()).limit(10)).all()
        tables["transactions"] = [
            {
                "date": t.date.date().isoformat(),
                "label": t.label,
                "debit": fmt_amt(t.debit or 0.0),
                "credit": fmt_amt(t.credit or 0.0),
                "currency": (t.currency or "MAD").upper(),
            }
            for t in txns
        ]
        answer_lines.append("Voici les 10 dernières transactions bancaires :")

    else:
        token = q.strip()
        inv = None
        try:
            if token.isdigit():
                inv = session.get(Invoice, int(token))
        except Exception:
            inv = None
        if not inv:
            invs = session.exec(select(Invoice).order_by(Invoice.id.desc()).limit(300)).all()
            token_low = token.lower()
            for i in invs:
                searchable = " ".join([
                    str(getattr(i, "invoice_no", "") or ""),
                    str(getattr(i, "supplier_name", "") or ""),
                    str(getattr(i, "bc_no", "") or ""),
                    str(getattr(i, "br_no", "") or ""),
                ]).lower()
                if token_low in searchable:
                    inv = i
                    break
        if inv:
            answer_lines.append(f"Facture/Achat #{inv.id} — {getattr(inv, 'invoice_no', None) or '—'}")
            answer_lines.append(f"- Fournisseur: {getattr(inv, 'supplier_name', None) or '—'}")
            answer_lines.append(f"- Montant: {fmt_amt(_invoice_amount(inv))} {(getattr(inv, 'currency', None) or 'MAD').upper()}")
            due = _invoice_due_date(inv)
            answer_lines.append(f"- Échéance: {due.isoformat() if due else '—'}")
            answer_lines.append(f"- Statut: {(getattr(inv, 'status', None) or '—')}")
            flags = _invoice_quality_flags(inv)
            if flags:
                answer_lines.append(f"- Points à corriger: {', '.join(flags)}")
        else:
            answer_lines.append("Questions reconnues : priorités, matching, risques, à payer 7 jours, à payer 30 jours, soldes, facture XYZ.")

    return {"answer_lines": answer_lines, "tables": tables, "q": q}


@app.get("/assistant", response_class=HTMLResponse)
def assistant_page(request: Request, q: Optional[str] = None, session: Session = Depends(get_session)):
    user = get_current_user(request, session)
    result = _assistant_db_answer(session, q or "")
    return templates.TemplateResponse(
        "assistant.html",
        {
            "request": request,
            "user": user,
            "q": result["q"],
            "answer_lines": result["answer_lines"],
            "tables": result["tables"],
        },
    )


def _assistant_answer(session: Session, q: str) -> dict:
    result = _assistant_db_answer(session, q)
    return {"answer": "\n".join(result["answer_lines"]).strip(), "tables": result["tables"]}


@app.post("/assistant/api")
def assistant_api(request: Request, payload: dict, session: Session = Depends(get_session)):
    """JSON endpoint for the floating assistant widget."""
    _ = get_current_user(request, session)
    q = (payload or {}).get("q") or ""
    return JSONResponse(content=_assistant_answer(session, q))


# ---------------- TRESORERIE ----------------

@app.get("/tresorerie", response_class=HTMLResponse)
def tresorerie_page(request: Request, session: Session = Depends(get_session)):
    user = get_current_user(request, session)

    today = datetime.utcnow().date()
    horizon_90 = today + timedelta(days=90)

    # --- accounts / balances ---
    accounts = session.exec(select(CompanyBankAccount)).all()
    active_accounts = [a for a in accounts if getattr(a, "is_active", True)]

    txns = session.exec(select(BankTxn).order_by(BankTxn.date.desc()).limit(8000)).all()
    last_bal_by_acct: dict[str, float] = {}
    for t in txns:
        if t.balance is None:
            continue
        key = (t.account_no or "").strip()
        if key and key not in last_bal_by_acct:
            last_bal_by_acct[key] = float(t.balance)

    # --- invoices horizon ---
    invs = session.exec(select(Invoice)).all()

    def _cur(x: Optional[str]) -> str:
        c = (x or "MAD").strip().upper()
        return c or "MAD"

    def _inv_amount(inv: Invoice) -> float:
        return float(getattr(inv, "amount_ttc", None) or getattr(inv, "amount", 0.0) or 0.0)

    # map invoice -> account via payment lines if possible
    inv_to_account: dict[int, Optional[int]] = {}
    lines = session.exec(select(PaymentLine)).all()
    batches = {b.id: b for b in session.exec(select(PaymentBatch)).all()}
    for ln in lines:
        if not ln.invoice_id:
            continue
        b = batches.get(ln.batch_id)
        aid = getattr(b, "company_account_id", None) if b else None
        if aid:
            inv_to_account[ln.invoice_id] = aid

    default_account_id = active_accounts[0].id if len(active_accounts) == 1 else None

    # KPIs by currency
    kpi = {"pay_7": {}, "pay_30": {}, "pay_60": {}, "pay_90": {}, "overdue": {}}

    def _add(bucket: dict, cur: str, amt: float):
        bucket[cur] = float(bucket.get(cur, 0.0) or 0.0) + float(amt or 0.0)

    # Bank projections (MAD-based)
    bank_proj: dict[int, dict] = {}
    for a in active_accounts:
        bank_proj[a.id] = {
            "bank": getattr(a, "bank_name", None) or "Banque",
            "account_no": (getattr(a, "account_no", None) or "").strip() or "—",
            "cur_bal": last_bal_by_acct.get((getattr(a, "account_no", None) or "").strip(), None),
            "out_7": 0.0,
            "out_30": 0.0,
            "out_60": 0.0,
            "out_90": 0.0,
            "currencies": set(),
        }

    upcoming: list[dict] = []

    for inv in invs:
        if (getattr(inv, "status", "") or "").upper() != "A_PAYER":
            continue
        due_dt = getattr(inv, "legal_due_date", None)
        if not due_dt:
            continue
        try:
            due = due_dt.date()
        except Exception:
            continue

        cur = _cur(getattr(inv, "currency", None))
        amt = _inv_amount(inv)

        if due < today:
            _add(kpi["overdue"], cur, amt)
        if due <= today + timedelta(days=7):
            _add(kpi["pay_7"], cur, amt)
        if due <= today + timedelta(days=30):
            _add(kpi["pay_30"], cur, amt)
        if due <= today + timedelta(days=60):
            _add(kpi["pay_60"], cur, amt)
        if due <= horizon_90:
            _add(kpi["pay_90"], cur, amt)

        if due <= horizon_90:
            upcoming.append({
                "due": due.isoformat(),
                "supplier": getattr(inv, "supplier_name", None) or "—",
                "ref": getattr(inv, "invoice_no", None) or f"ACHAT #{inv.id}",
                "amount": amt,
                "currency": cur,
                "id": inv.id,
            })

        aid = inv_to_account.get(inv.id) or default_account_id
        if aid and aid in bank_proj:
            bank_proj[aid]["currencies"].add(cur)
            if cur == "MAD":
                if due <= today + timedelta(days=7):
                    bank_proj[aid]["out_7"] += amt
                if due <= today + timedelta(days=30):
                    bank_proj[aid]["out_30"] += amt
                if due <= today + timedelta(days=60):
                    bank_proj[aid]["out_60"] += amt
                if due <= horizon_90:
                    bank_proj[aid]["out_90"] += amt

    upcoming = sorted(upcoming, key=lambda x: x["due"])[:25]

    # Chart data per currency (7/30/60/90)
    def _sum(bucket: dict, cur: str) -> float:
        return float(bucket.get(cur, 0.0) or 0.0)

    cur_set = {*(kpi["pay_7"].keys()), *(kpi["pay_30"].keys()), *(kpi["pay_60"].keys()), *(kpi["pay_90"].keys()), *(kpi["overdue"].keys())}
    if not cur_set:
        cur_set = {"MAD"}

    chart_by_cur: dict[str, dict] = {}
    for c in sorted(cur_set):
        chart_by_cur[c] = {
            "labels": ["7 jours", "30 jours", "60 jours", "90 jours"],
            "a_payer": [_sum(kpi["pay_7"], c), _sum(kpi["pay_30"], c), _sum(kpi["pay_60"], c), _sum(kpi["pay_90"], c)],
            "retard": [_sum(kpi["overdue"], c)] * 4,
        }
    default_chart_cur = "MAD" if "MAD" in chart_by_cur else next(iter(chart_by_cur.keys()))

    bank_rows: list[dict] = []
    for a in active_accounts:
        d = bank_proj[a.id]
        cur_bal = d["cur_bal"]
        out_30 = d["out_30"]
        out_90 = d["out_90"]
        proj_30 = (cur_bal - out_30) if cur_bal is not None else None
        proj_90 = (cur_bal - out_90) if cur_bal is not None else None
        alert = None
        if proj_30 is not None and proj_30 < 0:
            alert = "DEFICIT -30j"
        elif proj_90 is not None and proj_90 < 0:
            alert = "DEFICIT -90j"
        bank_rows.append({
            "bank": d["bank"],
            "account": d["account_no"],
            "cur_bal": cur_bal,
            "out_7": d["out_7"],
            "out_30": d["out_30"],
            "out_60": d["out_60"],
            "out_90": d["out_90"],
            "proj_30": proj_30,
            "proj_90": proj_90,
            "alert": alert,
            "multi_currency": (len(d["currencies"]) > 1) or (len(d["currencies"]) == 1 and (list(d["currencies"])[0] != "MAD")),
        })

    # Flux by date (last 30 days) for the small table
    by_date: dict[str, dict] = {}
    for t in txns:
        d = t.date.date().isoformat()
        by_date.setdefault(d, {"date": d, "debit": 0.0, "credit": 0.0})
        by_date[d]["debit"] += t.debit or 0.0
        by_date[d]["credit"] += t.credit or 0.0
    flux_rows = [by_date[d] for d in sorted(by_date.keys(), reverse=True)[:30]]

    return templates.TemplateResponse(
        "tresorerie.html",
        {
            "request": request,
            "user": user,
            "bank_rows": bank_rows,
            "chart_by_cur": chart_by_cur,
            "default_chart_cur": default_chart_cur,
            "upcoming": upcoming,
            "flux_rows": flux_rows,
        },
    )


# ---------------- DELAIS ----------------


# ---------------- RELEVES BANCAIRES (import + matching) ----------------

def _guess_csv_delimiter(sample: str) -> str:
    # Heuristique simple: ; est très fréquent au Maroc/France
    if sample.count(";") >= sample.count(","):
        return ";"
    return ","



def _parse_bank_rows_from_file(content: bytes, filename: str):
    """Parse CSV/XLSX de relevé en lignes standardisées.
    Gère aussi les relevés Excel sans ligne d'entête (cas BMCE simple :
    col A=date, col B=débit, col C=crédit, col D=libellé).
    """
    fn = (filename or "").lower()
    rows = []

    def _is_nonempty(v):
        return v is not None and str(v).strip() != ""

    def _normalize_debit_credit(dv, cv):
        dv = float(dv or 0.0)
        cv = float(cv or 0.0)
        # certains relevés stockent le débit en négatif dans une seule colonne
        if dv < 0 and cv == 0:
            return abs(dv), 0.0
        if cv < 0 and dv == 0:
            return abs(cv), 0.0
        return max(dv, 0.0), max(cv, 0.0)

    if fn.endswith(".xlsx") or fn.endswith(".xlsm") or fn.endswith(".xls"):
        wb = openpyxl.load_workbook(filename=BytesIO(content), data_only=True)

        def _sheet_score(ws):
            sample = list(ws.iter_rows(values_only=True, max_row=min(ws.max_row, 25)))
            score = 0
            for r in sample:
                vals = list(r)
                if not vals:
                    continue
                if len(vals) >= 4:
                    d = _parse_date(vals[0]) if vals[0] is not None else None
                    b = _to_float(vals[1]) if len(vals) > 1 else None
                    c = _to_float(vals[2]) if len(vals) > 2 else None
                    lbl = str(vals[3]).strip() if len(vals) > 3 and vals[3] is not None else ""
                    if d:
                        score += 2
                    if b is not None or c is not None:
                        score += 1
                    if lbl:
                        score += 1
            return score

        visible_sheets = [ws for ws in wb.worksheets if ws.sheet_state == "visible"]
        ws = max(visible_sheets or [wb.active], key=_sheet_score)

        raw_data = [list(r) for r in ws.iter_rows(values_only=True)]
        data = [[("" if v is None else str(v)) for v in r] for r in raw_data]

        # chercher header standard
        header_idx = None
        for i, r in enumerate(data[:20]):
            joined = " ".join(r).lower()
            if ("date" in joined and ("libell" in joined or "label" in joined or "motif" in joined)) and ("debit" in joined or "débit" in joined or "crédit" in joined or "credit" in joined or "montant" in joined):
                header_idx = i
                break

        if header_idx is not None:
            headers = [h.strip().lower() for h in data[header_idx]]
            for r in data[header_idx + 1:]:
                if not any(str(x).strip() for x in r):
                    continue
                row = {headers[i]: (r[i] if i < len(r) else "") for i in range(len(headers))}
                rows.append(row)
        else:
            # fallback format simple sans entête : A=date, B=débit, C=crédit, D=libellé, E=solde éventuel
            for r in raw_data:
                if not r or len(r) < 4:
                    continue
                d = _parse_date(r[0]) if len(r) > 0 else None
                b = _to_float(r[1]) if len(r) > 1 else None
                c = _to_float(r[2]) if len(r) > 2 else None
                lbl = str(r[3]).strip() if len(r) > 3 and r[3] is not None else ""
                bal = _to_float(r[4]) if len(r) > 4 else None
                if not d:
                    continue
                if (b is None and c is None) or not lbl:
                    continue
                rows.append({
                    "date": d.isoformat() if hasattr(d, "isoformat") else str(d),
                    "date valeur": d.isoformat() if hasattr(d, "isoformat") else str(d),
                    "debit": "" if b is None else str(b),
                    "credit": "" if c is None else str(c),
                    "libellé": lbl,
                    "solde": "" if bal is None else str(bal),
                })

    else:
        txt = content.decode("utf-8", errors="ignore")
        sample = txt[:2000]
        delim = _guess_csv_delimiter(sample)
        reader = csv.DictReader(txt.splitlines(), delimiter=delim)
        for row in reader:
            if not row:
                continue
            rows.append({(k or "").strip().lower(): (v or "").strip() for k, v in row.items()})

    def pick(row: dict, keys: list[str]) -> str:
        for k in keys:
            for kk, vv in row.items():
                if k in str(kk).lower():
                    return ("" if vv is None else str(vv)).strip()
        return ""

    parsed = []
    for r in rows:
        d0 = pick(r, ["date valeur", "value date", "valeur"])
        d1 = pick(r, ["date", "operation", "opération"])
        label = pick(r, ["libell", "label", "motif", "designation", "désignation", "objet"])
        debit = pick(r, ["debit", "débit", "sortie"])
        credit = pick(r, ["credit", "crédit", "entrée", "entree"])
        balance = pick(r, ["solde", "balance"])

        dt = _parse_date(d1) or _parse_date(d0) or None
        vd = _parse_date(d0) or dt

        if not vd and not dt:
            continue

        dv = _to_float(debit) or 0.0
        cv = _to_float(credit) or 0.0
        dv, cv = _normalize_debit_credit(dv, cv)
        bal = _to_float(balance)

        if not label and dv == 0.0 and cv == 0.0:
            continue

        parsed.append(
            {
                "date": dt or vd,
                "value_date": vd or dt,
                "label": label or "",
                "debit": float(dv),
                "credit": float(cv),
                "balance": bal,
            }
        )
    return parsed


def _ensure_cashflow_actual_from_match(session: Session, match: InvoicePaymentMatch):
    """Crée/maj le flux réel à partir d'un match facture<->banque."""
    inv = session.get(Invoice, match.invoice_id)
    txn = session.get(BankTxn, match.banktxn_id)
    if not inv or not txn:
        return

    # date valeur prioritaire
    dt = txn.value_date or txn.date
    month = f"{dt.year:04d}-{dt.month:02d}"

    # rubrique: on mappe invoice.category -> CashflowRubrique.rubrique
    rubrique_id = None
    cat = (inv.category or "").strip()
    if cat:
        rub = session.exec(select(CashflowRubrique).where(CashflowRubrique.rubrique == cat)).first()
        if rub:
            rubrique_id = rub.id

    # montant: si match.matched_amount renseigné, sinon debit/credit
    amt = float(match.matched_amount or 0.0)
    if amt <= 0:
        amt = float(txn.debit or 0.0) if float(txn.debit or 0.0) > 0 else float(txn.credit or 0.0)

    # upsert (un match -> un flux réel)
    existing = session.exec(select(CashflowActual).where(CashflowActual.banktxn_id == txn.id, CashflowActual.invoice_id == inv.id)).first()
    if existing:
        existing.rubrique_id = rubrique_id
        existing.actual_debit_date = dt
        existing.actual_month = month
        existing.amount = amt
        session.add(existing)
        session.commit()
        return

    cf = CashflowActual(
        rubrique_id=rubrique_id,
        invoice_id=inv.id,
        banktxn_id=txn.id,
        actual_debit_date=dt,
        actual_month=month,
        amount=amt,
    )
    session.add(cf)
    session.commit()


@app.get("/releves", response_class=HTMLResponse)
def releves_page(request: Request, session: Session = Depends(get_session)):
    user = get_current_user(request, session)

    only_unmatched = request.query_params.get("unmatched") == "1"

    # txns + statut match
    txns = session.exec(select(BankTxn).order_by(BankTxn.value_date.desc().nullslast(), BankTxn.date.desc()).limit(800)).all()
    matched_ids = {m.banktxn_id for m in session.exec(select(InvoicePaymentMatch.banktxn_id)).all()}

    if only_unmatched:
        txns = [t for t in txns if t.id not in matched_ids]

    # factures candidates (A_PAYER / PAYEE récente)
    invs = session.exec(select(Invoice).order_by(Invoice.due_date.desc().nullslast(), Invoice.invoice_date.desc().nullslast()).limit(500)).all()
    invs = [i for i in invs if (i.status or "").upper() in ("A_PAYER", "PAYEE")]

    rubs = session.exec(select(CashflowRubrique).order_by(CashflowRubrique.rubrique.asc())).all()

    accounts = session.exec(
        select(CompanyBankAccount)
        .where(CompanyBankAccount.is_active == True)
        .order_by(CompanyBankAccount.bank_name.asc(), CompanyBankAccount.account_no.asc())
    ).all()

    msg = request.query_params.get("msg")
    n = request.query_params.get("n")
    return templates.TemplateResponse(
        "releves.html",
        {
            "request": request,
            "user": user,
            "txns": txns,
            "matched_ids": matched_ids,
            "invs": invs,
            "rubs": rubs,
            "accounts": accounts,
            "msg": msg,
            "n": n,
            "only_unmatched": only_unmatched,
        },
    )


@app.post("/releves/import")
async def releves_import(
    request: Request,
    session: Session = Depends(get_session),
    bank_name: str = Form(""),
    account_no: str = Form(""),
    currency: str = Form("MAD"),
    company_bank_account_id: Optional[int] = Form(None),
    file: UploadFile = File(...),
):
    selected_acc = None
    if company_bank_account_id:
        selected_acc = session.get(CompanyBankAccount, company_bank_account_id)
        if not selected_acc:
            return RedirectResponse("/releves?msg=bad_account", status_code=303)
        bank_name = selected_acc.bank_name or bank_name
        account_no = selected_acc.account_no or account_no

    if not (bank_name or "").strip() or not (account_no or "").strip():
        return RedirectResponse("/releves?msg=missing_account", status_code=303)

    filename = (file.filename or "releve").lower()
    allowed_ext = (".csv", ".xlsx", ".xls")
    if not filename.endswith(allowed_ext):
        return RedirectResponse("/releves?msg=bad_format", status_code=303)

    content = await file.read()
    if not content:
        return RedirectResponse("/releves?msg=empty", status_code=303)

    try:
        parsed = _parse_bank_rows_from_file(content, file.filename or "releve")
    except Exception:
        return RedirectResponse("/releves?msg=parse_error", status_code=303)

    inserted = 0
    for r in parsed:
        dt = r["date"]
        vd = r["value_date"] or dt
        label = (r["label"] or "").strip()[:400]
        label_norm = _norm_label(label)[:400]

        amt = float(r.get("debit") or 0.0) if float(r.get("debit") or 0.0) > 0 else float(r.get("credit") or 0.0)

        dedup = _make_banktxn_dedup_key(bank_name, account_no, vd, amt, label_norm)

        exists = session.exec(select(BankTxn).where(BankTxn.dedup_key == dedup)).first()
        if exists:
            continue

        txn = BankTxn(
            bank_name=(bank_name.strip()[:80] if bank_name else None),
            account_no=(account_no.strip()[:80] if account_no else None),
            currency=(currency.strip()[:10] if currency else None),
            date=dt,
            value_date=vd,
            label=label,
            label_norm=label_norm,
            debit=float(r.get("debit") or 0.0),
            credit=float(r.get("credit") or 0.0),
            balance=r.get("balance"),
            dedup_key=dedup,
        )
        session.add(txn)
        inserted += 1

    session.commit()
    return RedirectResponse(f"/releves?msg=import_ok&n={inserted}", status_code=303)


@app.post("/matching/manual")
def matching_manual(
    request: Request,
    session: Session = Depends(get_session),
    banktxn_id: int = Form(...),
    invoice_id: int = Form(...),
    matched_amount: Optional[str] = Form(None),
):
    def _back(msg_code: str) -> RedirectResponse:
        ref = (request.headers.get("referer") or "")
        if "/matching" in ref:
            # on reste sur l'écran matching (plus pratique)
            return RedirectResponse(f"/matching?msg={msg_code}", status_code=303)
        return RedirectResponse(f"/releves?msg={msg_code}", status_code=303)

    txn = session.get(BankTxn, banktxn_id)
    inv = session.get(Invoice, invoice_id)
    if not txn or not inv:
        return _back("bad_ids")

    # ---------------- Matching 1:1 (choix ERP) ----------------
    # Décision: 1 virement = 1 facture (pas de paiements groupés).
    # Donc: un mouvement bancaire ne peut matcher qu'une seule facture, et inversement.
    other_for_txn = session.exec(
        select(InvoicePaymentMatch).where(
            InvoicePaymentMatch.banktxn_id == banktxn_id,
            InvoicePaymentMatch.invoice_id != invoice_id,
        )
    ).first()
    if other_for_txn:
        return _back("txn_already_matched")

    other_for_inv = session.exec(
        select(InvoicePaymentMatch).where(
            InvoicePaymentMatch.invoice_id == invoice_id,
            InvoicePaymentMatch.banktxn_id != banktxn_id,
        )
    ).first()
    if other_for_inv:
        return _back("inv_already_matched")

    amt = _to_float(matched_amount) if matched_amount else None
    if amt is None:
        # défaut = debit si dispo sinon credit
        amt = float(txn.debit or 0.0) if float(txn.debit or 0.0) > 0 else float(txn.credit or 0.0)

    existing = session.exec(select(InvoicePaymentMatch).where(InvoicePaymentMatch.banktxn_id == banktxn_id, InvoicePaymentMatch.invoice_id == invoice_id)).first()
    if existing:
        existing.matched_amount = float(amt)
        existing.method = "MANUAL"
        session.add(existing)
        session.commit()
        _ensure_cashflow_actual_from_match(session, existing)
        return _back("match_updated")

    m = InvoicePaymentMatch(
        invoice_id=invoice_id,
        banktxn_id=banktxn_id,
        matched_amount=float(amt),
        method="MANUAL",
    )
    session.add(m)
    session.commit()
    session.refresh(m)

    # statut facture (si sortie et montant >= TTC)
    if float(txn.debit or 0.0) > 0 and float(amt) > 0:
        inv.status = "PAYEE"
        inv.payment_date = (txn.value_date or txn.date)
        session.add(inv)
        session.commit()

    _ensure_cashflow_actual_from_match(session, m)
    return _back("match_ok")


# ---------------- MATCHING ASSISTE (suggestions) ----------------

def _txn_amount(txn: BankTxn) -> float:
    d = float(txn.debit or 0.0)
    c = float(txn.credit or 0.0)
    if d > 0:
        return d
    if c > 0:
        return c
    return 0.0


def _score_match(txn: BankTxn, inv: Invoice) -> int:
    """Score simple (0..100) pour proposer un match 1:1."""
    txn_amt = _txn_amount(txn)
    inv_amt = float(inv.amount or 0.0)
    if txn_amt <= 0 or inv_amt <= 0:
        return 0

    # 1) Montant (strict, car pas de groupés)
    if abs(txn_amt - inv_amt) > 0.01:
        return 0
    score = 70

    # 2) Proximité date (date valeur)
    dt = (txn.value_date or txn.date).date()
    tgt = None
    for cand in (inv.due_date_planned, inv.due_date, inv.invoice_date):
        if cand:
            tgt = cand.date()
            break
    if tgt:
        delta = abs((dt - tgt).days)
        if delta <= 3:
            score += 20
        elif delta <= 7:
            score += 10

    # 3) Libellé vs fournisseur / numéro facture
    lab = (txn.label_norm or _norm_label(txn.label or ""))
    if inv.invoice_no:
        invno = _norm_label(inv.invoice_no)
        if invno and invno in (lab or ""):
            score += 10
    if inv.supplier_name:
        sup = _norm_label(inv.supplier_name)
        tok = ""
        for t in sup.split():
            if len(t) >= 5:
                tok = t
                break
        if tok and tok in (lab or ""):
            score += 10

    return int(score)


def _suggest_for_txn(txn: BankTxn, invs: list[Invoice], limit: int = 5) -> list[dict]:
    scored: list[dict] = []
    for inv in invs:
        if (inv.status or "").upper() in ("PAYEE", "ANNULEE"):
            continue
        s = _score_match(txn, inv)
        if s > 0:
            scored.append({"inv": inv, "score": s})
    scored.sort(key=lambda x: x["score"], reverse=True)
    return scored[:limit]


@app.get("/matching", response_class=HTMLResponse)
def matching_page(request: Request, session: Session = Depends(get_session)):
    user = get_current_user(request, session)

    matched_txn_ids = {m.banktxn_id for m in session.exec(select(InvoicePaymentMatch.banktxn_id)).all()}
    txns_all = session.exec(
        select(BankTxn)
        .where(BankTxn.debit > 0)
        .order_by(BankTxn.value_date.desc().nullslast(), BankTxn.date.desc(), BankTxn.id.desc())
        .limit(500)
    ).all()
    txns = [t for t in txns_all if t.id not in matched_txn_ids]

    invs = session.exec(
        select(Invoice)
        .where(Invoice.status == "A_PAYER")
        .order_by(Invoice.legal_due_date.asc().nullslast(), Invoice.invoice_date.asc().nullslast())
        .limit(1200)
    ).all()

    auto = request.query_params.get("auto") == "1"
    focused_id = request.query_params.get("txn")
    focus_txn = session.get(BankTxn, int(focused_id)) if (focused_id and focused_id.isdigit()) else None

    suggestions_map: dict[int, list[dict]] = {}
    txns_for_scan = txns[:80] if auto else ([focus_txn] if focus_txn else [])
    for t in [x for x in txns_for_scan if x is not None]:
        candidates = _candidate_invoices_for_txn(t, invs)
        suggestions_map[t.id] = _suggest_for_txn(t, candidates, limit=8 if focus_txn else 5)

    automation = _build_automation_snapshot(session)
    high_conf_matches = automation["high_conf_matches"][:20]

    msg = request.query_params.get("msg")
    return templates.TemplateResponse(
        "matching.html",
        {
            "request": request,
            "user": user,
            "txns": txns,
            "focus_txn": focus_txn,
            "suggestions_map": suggestions_map,
            "msg": msg,
            "auto": auto,
            "high_conf_matches": high_conf_matches,
        },
    )


@app.post("/matching/auto_apply")
def matching_auto_apply(request: Request, session: Session = Depends(get_session), limit: int = Form(10)):
    _ = get_current_user(request, session)
    snapshot = _build_automation_snapshot(session)
    applied = 0

    for row in snapshot["high_conf_matches"][: max(1, min(int(limit or 10), 50))]:
        txn = row["txn"]
        inv = row["top"]["inv"]

        other_for_txn = session.exec(
            select(InvoicePaymentMatch).where(InvoicePaymentMatch.banktxn_id == txn.id)
        ).first()
        other_for_inv = session.exec(
            select(InvoicePaymentMatch).where(InvoicePaymentMatch.invoice_id == inv.id)
        ).first()
        if other_for_txn or other_for_inv:
            continue

        amt = _txn_amount(txn)
        if amt <= 0:
            continue

        match = InvoicePaymentMatch(
            invoice_id=inv.id,
            banktxn_id=txn.id,
            matched_amount=float(amt),
            method="AUTO_HIGH_CONF",
        )
        session.add(match)
        session.commit()
        session.refresh(match)

        inv.status = "PAYEE"
        inv.payment_date = (txn.value_date or txn.date)
        session.add(inv)
        session.commit()

        _ensure_cashflow_actual_from_match(session, match)
        applied += 1

    return RedirectResponse(f"/matching?auto=1&msg=auto_applied_{applied}", status_code=303)




@app.post("/releves/delete-selected")
def releves_delete_selected(
    request: Request,
    session: Session = Depends(get_session),
    txn_ids: Optional[list[int]] = Form(None),
):
    _ = get_current_user(request, session)
    ids = [int(x) for x in (txn_ids or []) if str(x).isdigit() or isinstance(x, int)]
    if not ids:
        return RedirectResponse("/releves?msg=no_selection", status_code=303)

    deleted = 0
    blocked = 0
    for txn_id in ids:
        txn = session.get(BankTxn, txn_id)
        if not txn:
            continue
        has_match = session.exec(select(InvoicePaymentMatch).where(InvoicePaymentMatch.banktxn_id == txn_id)).first()
        if has_match:
            blocked += 1
            continue
        # supprimer éventuel classement hors facture lié à ce mouvement
        cf_rows = session.exec(select(CashflowActual).where(CashflowActual.banktxn_id == txn_id)).all()
        for cf in cf_rows:
            session.delete(cf)
        session.delete(txn)
        deleted += 1
    session.commit()
    return RedirectResponse(f"/releves?msg=delete_ok&n={deleted}&blocked={blocked}", status_code=303)


@app.post("/releves/{txn_id}/delete")
def releves_delete_txn(
    txn_id: int,
    request: Request,
    session: Session = Depends(get_session),
):
    _ = get_current_user(request, session)
    txn = session.get(BankTxn, txn_id)
    if not txn:
        return RedirectResponse("/releves?msg=bad_txn", status_code=303)

    has_match = session.exec(select(InvoicePaymentMatch).where(InvoicePaymentMatch.banktxn_id == txn_id)).first()
    if has_match:
        return RedirectResponse("/releves?msg=delete_blocked_matched", status_code=303)

    cf_rows = session.exec(select(CashflowActual).where(CashflowActual.banktxn_id == txn_id)).all()
    for cf in cf_rows:
        session.delete(cf)
    session.delete(txn)
    session.commit()
    return RedirectResponse("/releves?msg=delete_one_ok", status_code=303)


@app.post("/releves/{txn_id}/classify")
def releves_classify_txn(
    txn_id: int,
    request: Request,
    session: Session = Depends(get_session),
    rubrique_id: int = Form(...),
    amount: Optional[str] = Form(None),
):
    txn = session.get(BankTxn, txn_id)
    if not txn:
        return RedirectResponse("/releves?msg=bad_txn", status_code=303)

    dt = txn.value_date or txn.date
    month = f"{dt.year:04d}-{dt.month:02d}"

    amt = _to_float(amount) if amount else None
    if amt is None:
        amt = float(txn.debit or 0.0) if float(txn.debit or 0.0) > 0 else float(txn.credit or 0.0)

    existing = session.exec(select(CashflowActual).where(CashflowActual.banktxn_id == txn_id, CashflowActual.invoice_id == None)).first()
    if existing:
        existing.rubrique_id = rubrique_id
        existing.actual_debit_date = dt
        existing.actual_month = month
        existing.amount = float(amt)
        session.add(existing)
        session.commit()
        return RedirectResponse("/releves?msg=class_ok", status_code=303)

    cf = CashflowActual(
        rubrique_id=rubrique_id,
        invoice_id=None,
        banktxn_id=txn_id,
        actual_debit_date=dt,
        actual_month=month,
        amount=float(amt),
    )
    session.add(cf)
    session.commit()
    return RedirectResponse("/releves?msg=class_ok", status_code=303)


# ---------------- CASHFLOW REEL (par mois, depuis relevé) ----------------

@app.get("/cashflow", response_class=HTMLResponse)
def cashflow_reel(request: Request, session: Session = Depends(get_session)):
    user = get_current_user(request, session)

    month = request.query_params.get("month")  # YYYY-MM ou None
    q = select(CashflowActual)
    if month:
        q = q.where(CashflowActual.actual_month == month)

    rows = session.exec(q.order_by(CashflowActual.actual_debit_date.desc())).all()

    # agrégation
    totals = {}
    for r in rows:
        key = r.actual_month
        totals.setdefault(key, 0.0)
        totals[key] += float(r.amount or 0.0)

    rub_map = {rb.id: rb for rb in session.exec(select(CashflowRubrique)).all()}

    return templates.TemplateResponse(
        "cashflow_reel.html",
        {
            "request": request,
            "user": user,
            "rows": rows,
            "totals": totals,
            "month": month,
            "rub_map": rub_map,
        },
    )


# ---------------- PLANNING PAIEMENT (séparé) ----------------

@app.get("/planning", response_class=HTMLResponse)
def planning_paiement(request: Request, session: Session = Depends(get_session)):
    user = get_current_user(request, session)

    today = date.today()
    horizon_end = today + timedelta(days=90)

    invs = session.exec(
        select(Invoice).order_by(
            Invoice.legal_due_date.asc().nullslast(),
            Invoice.invoice_date.desc().nullslast(),
        )
    ).all()
    invs = [i for i in invs if (i.status or "").upper() == "A_PAYER"]

    accs = session.exec(
        select(CompanyBankAccount)
        .where(CompanyBankAccount.is_active == True)
        .order_by(CompanyBankAccount.bank_name.asc())
    ).all()
    bank_balances = _latest_bank_balances(session)

    rows = []
    totals_by_cur: dict[str, float] = {}
    for inv in invs:
        snap = _invoice_priority_snapshot(inv, today)
        due = snap["due"]
        if due and not (today <= due <= horizon_end) and snap["days_left"] is not None and snap["days_left"] > 90:
            continue
        snap["recommended_account"] = _recommend_company_account(accs, bank_balances, snap["currency"], snap["amount"])
        rows.append(snap)
        totals_by_cur[snap["currency"]] = totals_by_cur.get(snap["currency"], 0.0) + snap["amount"]

    rows.sort(
        key=lambda r: (
            -r["score"],
            r["days_left"] if r["days_left"] is not None else 10**6,
            -r["amount"],
        )
    )

    summary = {
        "critical_count": len([r for r in rows if r["label"] == "CRITIQUE"]),
        "high_count": len([r for r in rows if r["label"] == "HAUTE"]),
        "missing_count": len([r for r in rows if r["flags"]]),
        "totals_by_cur": totals_by_cur,
    }

    return templates.TemplateResponse(
        "planning.html",
        {
            "request": request,
            "user": user,
            "rows": rows,
            "today": today,
            "horizon_days": 90,
            "accs": accs,
            "summary": summary,
        },
    )


# ---------------- ANALYTIQUE (pilotage) ----------------

@app.get("/analytics", response_class=HTMLResponse)
def analytics_page(request: Request, session: Session = Depends(get_session)):
    user = get_current_user(request, session)

    # filtres simples
    year = request.query_params.get("year")
    scope = (request.query_params.get("scope") or "open").lower()  # open / all
    try:
        year_i = int(year) if year else date.today().year
    except Exception:
        year_i = date.today().year

    invs = session.exec(select(Invoice)).all()
    # filtre année sur invoice_date si dispo, sinon reception_date
    def _inv_year(i: Invoice) -> int | None:
        d = i.invoice_date or i.reception_date
        return d.year if d else None

    invs = [i for i in invs if (_inv_year(i) == year_i)]
    if scope == "open":
        invs = [i for i in invs if (i.status or "").upper() == "A_PAYER"]

    def amt(i: Invoice) -> float:
        v = i.amount_ttc if i.amount_ttc is not None else i.amount
        try:
            return float(v or 0.0)
        except Exception:
            return 0.0

    currency = "MAD"
    for i in invs:
        if i.currency:
            currency = i.currency
            break

    def group_by(key_fn):
        m: dict[str, dict] = {}
        for i in invs:
            k = (key_fn(i) or "Non renseigné").strip() or "Non renseigné"
            if k not in m:
                m[k] = {"key": k, "total": 0.0, "count": 0}
            m[k]["total"] += amt(i)
            m[k]["count"] += 1
        rows = list(m.values())
        rows.sort(key=lambda r: r["total"], reverse=True)
        return rows

    by_analytic = group_by(lambda i: i.analytic)
    by_project = group_by(lambda i: i.project)
    by_site = group_by(lambda i: i.site)
    by_rubrique = group_by(lambda i: i.cashflow_rubrique)

    return templates.TemplateResponse(
        "analytics.html",
        {
            "request": request,
            "user": user,
            "year": year_i,
            "scope": scope,
            "currency": currency,
            "by_analytic": by_analytic,
            "by_project": by_project,
            "by_site": by_site,
            "by_rubrique": by_rubrique,
        },
    )




# ---------------- ANALYTIQUE : AFFECTATIONS ----------------

@app.get("/analytics/allocations", response_class=HTMLResponse)
def analytics_allocations(request: Request, session: Session = Depends(get_session)):
    user = get_current_user(request, session)

    q = (request.query_params.get("q") or "").strip().lower()

    invs = session.exec(select(Invoice).order_by(Invoice.id.desc())).all()
    # uniquement affectables (BR/FACTURE)
    invs = [i for i in invs if _invoice_is_allocatable(session, i)]

    if q:
        def _hit(i: Invoice) -> bool:
            s = " ".join([
                str(i.id or ""),
                (i.supplier_name or ""),
                (i.invoice_no or ""),
                (i.bc_no or ""),
                (i.br_no or ""),
            ]).lower()
            return q in s
        invs = [i for i in invs if _hit(i)]

    posts = session.exec(select(Post).where((Post.kind == "ANALYTIC") & (Post.is_active == True)).order_by(Post.code, Post.name)).all()  # type: ignore

    rows = []
    for i in invs:
        total = _invoice_amount_for_alloc(i)
        allocated = _alloc_sum(session, int(i.id), "ANALYTIC") if i.id else 0.0
        remaining = total - allocated
        rows.append({"inv": i, "total": total, "allocated": allocated, "remaining": remaining})

    
    # charger les lignes d'affectation (pour affichage)
    inv_ids = [int(i.id) for i in invs if i.id]
    if inv_ids:
        allocs = session.exec(
            select(AllocationLine).where(
                (AllocationLine.kind == "ANALYTIC") & (AllocationLine.invoice_id.in_(inv_ids))
            ).order_by(AllocationLine.id.desc())
        ).all()
        by_inv = {}
        for a in allocs:
            by_inv.setdefault(int(a.invoice_id), []).append(a)
        for i in invs:
            if i.id:
                setattr(i, "alloc_analytic_lines", by_inv.get(int(i.id), []))
    return templates.TemplateResponse(
        "analytics_allocations.html",
        {"request": request, "user": user, "rows": rows, "posts": posts, "q": q},
    )


@app.post("/analytics/allocations/add")
def analytics_allocations_add(
    request: Request,
    session: Session = Depends(get_session),
    invoice_id: int = Form(...),
    post_id: int = Form(...),
    amount: str = Form(""),
    note: Optional[str] = Form(None),
):
    user = get_current_user(request, session)

    inv = session.get(Invoice, invoice_id)
    if not inv:
        return RedirectResponse("/analytics/allocations", status_code=303)

    # sécurité : pas d'affectation si BC only
    if not _invoice_is_allocatable(session, inv):
        return RedirectResponse("/analytics/allocations", status_code=303)

    post = session.get(Post, post_id)
    if not post or (post.kind != "ANALYTIC"):
        return RedirectResponse("/analytics/allocations", status_code=303)

    amt = _to_float(amount) or 0.0
    line = AllocationLine(
        invoice_id=invoice_id,
        kind="ANALYTIC",
        post_id=post_id,
        post_code=post.code,
        post_name=post.name,
        amount=float(amt),
        note=(note or None),
    )
    session.add(line)
    session.commit()
    return RedirectResponse("/analytics/allocations", status_code=303)


@app.post("/analytics/allocations/{line_id}/delete")
def analytics_allocations_delete(line_id: int, request: Request, session: Session = Depends(get_session)):
    user = get_current_user(request, session)
    line = session.get(AllocationLine, line_id)
    if line and line.kind == "ANALYTIC":
        session.delete(line)
        session.commit()
    return RedirectResponse("/analytics/allocations", status_code=303)


# ---------------- BUDGET (pilotage) ----------------

@app.get("/budget", response_class=HTMLResponse)
def budget_page(request: Request, session: Session = Depends(get_session)):
    user = get_current_user(request, session)

    year = request.query_params.get("year")
    try:
        year_i = int(year) if year else date.today().year
    except Exception:
        year_i = date.today().year

    lines = session.exec(
        select(BudgetLine).where(BudgetLine.year == year_i).order_by(BudgetLine.month.asc(), BudgetLine.cashflow_rubrique.asc().nullslast())
    ).all()

    return templates.TemplateResponse(
        "budget.html",
        {
            "request": request,
            "user": user,
            "year": year_i,
            "now_month": date.today().month,
            "lines": lines,
        },
    )


@app.post("/budget/upsert")
def budget_upsert(
    request: Request,
    session: Session = Depends(get_session),
    year: int = Form(...),
    month: int = Form(...),
    amount_planned: float = Form(...),
    currency: str = Form("MAD"),
    cashflow_rubrique: Optional[str] = Form(None),
    cashflow_category: Optional[str] = Form(None),
    analytic: Optional[str] = Form(None),
    project: Optional[str] = Form(None),
    site: Optional[str] = Form(None),
    cost_center: Optional[str] = Form(None),
    notes: Optional[str] = Form(None),
):
    _ = get_current_user(request, session)

    # normalisation
    month = max(1, min(12, int(month)))
    currency = (currency or "MAD").upper().strip()[:10]

    def norm(s: Optional[str], n: int):
        if not s:
            return None
        s2 = str(s).strip()
        return s2[:n] if s2 else None

    cashflow_rubrique = norm(cashflow_rubrique, 80)
    cashflow_category = norm(cashflow_category, 60)
    analytic = norm(analytic, 80)
    project = norm(project, 80)
    site = norm(site, 30)
    cost_center = norm(cost_center, 60)
    notes = norm(notes, 200)

    # clé métier : (year, month, currency, rubrique, analytic, project, site, cost_center)
    existing = session.exec(
        select(BudgetLine)
        .where(BudgetLine.year == year)
        .where(BudgetLine.month == month)
        .where(BudgetLine.currency == currency)
        .where(BudgetLine.cashflow_rubrique == cashflow_rubrique)
        .where(BudgetLine.analytic == analytic)
        .where(BudgetLine.project == project)
        .where(BudgetLine.site == site)
        .where(BudgetLine.cost_center == cost_center)
    ).first()

    if existing:
        existing.amount_planned = float(amount_planned)
        existing.cashflow_category = cashflow_category
        existing.notes = notes
        session.add(existing)
        session.commit()
        return RedirectResponse(f"/budget?year={year}", status_code=303)

    bl = BudgetLine(
        year=int(year),
        month=int(month),
        currency=currency,
        amount_planned=float(amount_planned),
        cashflow_category=cashflow_category,
        cashflow_rubrique=cashflow_rubrique,
        analytic=analytic,
        project=project,
        site=site,
        cost_center=cost_center,
        notes=notes,
    )
    session.add(bl)
    session.commit()
    return RedirectResponse(f"/budget?year={year}", status_code=303)


@app.post("/budget/{budget_id}/delete")
def budget_delete(budget_id: int, request: Request, session: Session = Depends(get_session)):
    _ = get_current_user(request, session)
    bl = session.get(BudgetLine, budget_id)
    if bl:
        y = bl.year
        session.delete(bl)
        session.commit()
        return RedirectResponse(f"/budget?year={y}", status_code=303)
    return RedirectResponse("/budget", status_code=303)



# ---------------- BUDGET : AFFECTATIONS ----------------

@app.get("/budget/allocations", response_class=HTMLResponse)
def budget_allocations(request: Request, session: Session = Depends(get_session)):
    user = get_current_user(request, session)

    q = (request.query_params.get("q") or "").strip().lower()

    invs = session.exec(select(Invoice).order_by(Invoice.id.desc())).all()
    invs = [i for i in invs if _invoice_is_allocatable(session, i)]

    if q:
        def _hit(i: Invoice) -> bool:
            s = " ".join([
                str(i.id or ""),
                (i.supplier_name or ""),
                (i.invoice_no or ""),
                (i.bc_no or ""),
                (i.br_no or ""),
            ]).lower()
            return q in s
        invs = [i for i in invs if _hit(i)]

    posts = session.exec(select(Post).where((Post.kind == "BUDGET") & (Post.is_active == True)).order_by(Post.code, Post.name)).all()  # type: ignore

    rows = []
    for i in invs:
        total = _invoice_amount_for_alloc(i)
        allocated = _alloc_sum(session, int(i.id), "BUDGET") if i.id else 0.0
        remaining = total - allocated
        rows.append({"inv": i, "total": total, "allocated": allocated, "remaining": remaining})

    
    # charger les lignes d'affectation (pour affichage)
    inv_ids = [int(i.id) for i in invs if i.id]
    if inv_ids:
        allocs = session.exec(
            select(AllocationLine).where(
                (AllocationLine.kind == "BUDGET") & (AllocationLine.invoice_id.in_(inv_ids))
            ).order_by(AllocationLine.id.desc())
        ).all()
        by_inv = {}
        for a in allocs:
            by_inv.setdefault(int(a.invoice_id), []).append(a)
        for i in invs:
            if i.id:
                setattr(i, "alloc_budget_lines", by_inv.get(int(i.id), []))
    return templates.TemplateResponse(
        "budget_allocations.html",
        {"request": request, "user": user, "rows": rows, "posts": posts, "q": q},
    )


@app.post("/budget/allocations/add")
def budget_allocations_add(
    request: Request,
    session: Session = Depends(get_session),
    invoice_id: int = Form(...),
    post_id: int = Form(...),
    amount: str = Form(""),
    note: Optional[str] = Form(None),
):
    user = get_current_user(request, session)

    inv = session.get(Invoice, invoice_id)
    if not inv:
        return RedirectResponse("/budget/allocations", status_code=303)

    if not _invoice_is_allocatable(session, inv):
        return RedirectResponse("/budget/allocations", status_code=303)

    post = session.get(Post, post_id)
    if not post or (post.kind != "BUDGET"):
        return RedirectResponse("/budget/allocations", status_code=303)

    amt = _to_float(amount) or 0.0
    line = AllocationLine(
        invoice_id=invoice_id,
        kind="BUDGET",
        post_id=post_id,
        post_code=post.code,
        post_name=post.name,
        amount=float(amt),
        note=(note or None),
    )
    session.add(line)
    session.commit()
    return RedirectResponse("/budget/allocations", status_code=303)


@app.post("/budget/allocations/{line_id}/delete")
def budget_allocations_delete(line_id: int, request: Request, session: Session = Depends(get_session)):
    user = get_current_user(request, session)
    line = session.get(AllocationLine, line_id)
    if line and line.kind == "BUDGET":
        session.delete(line)
        session.commit()
    return RedirectResponse("/budget/allocations", status_code=303)


@app.get("/delais", response_class=HTMLResponse)
def delais_page(request: Request, session: Session = Depends(get_session)):
    user = get_current_user(request, session)
    # ✅ Déclaration: exclure les fournisseurs étrangers
    rows = session.exec(
        select(Invoice)
        .outerjoin(Supplier, Supplier.id == Invoice.supplier_id)
        .where((Supplier.is_foreign == False) | (Invoice.supplier_id == None))
        .order_by(Invoice.invoice_date.desc().nullslast(), Invoice.id.desc())
        .limit(300)
    ).all()

    today = datetime.utcnow().date()
    kpi_total = 0.0
    kpi_late_amt = 0.0
    kpi_late_cnt = 0
    kpi_missing_due = 0
    kpi_missing_supplier = 0

    for inv in rows:
        amt = float((inv.amount_ttc if inv.amount_ttc is not None else inv.amount) or 0.0)
        kpi_total += amt
        if not (inv.supplier_name or "").strip():
            kpi_missing_supplier += 1
        if not inv.legal_due_date:
            kpi_missing_due += 1
        else:
            try:
                if (inv.payment_date is None) and (inv.legal_due_date.date() < today):
                    kpi_late_cnt += 1
                    kpi_late_amt += amt
            except Exception:
                pass

    return templates.TemplateResponse(
        "delais.html",
        {
            "request": request,
            "user": user,
            "rows": rows,
            "kpi_total": kpi_total,
            "kpi_late_amt": kpi_late_amt,
            "kpi_late_cnt": kpi_late_cnt,
            "kpi_missing_due": kpi_missing_due,
            "kpi_missing_supplier": kpi_missing_supplier,
        },
    )


@app.get("/delais/export.xlsx")
def delais_export_xlsx(request: Request, session: Session = Depends(get_session)):
    """Export simple (Excel) des délais de paiement — fournisseurs marocains uniquement."""
    _ = get_current_user(request, session)

    invs = session.exec(
        select(Invoice)
        .outerjoin(Supplier, Supplier.id == Invoice.supplier_id)
        .where((Supplier.is_foreign == False) | (Invoice.supplier_id == None))
        .order_by(Invoice.invoice_date.asc().nullslast(), Invoice.id.asc())
    ).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DelaisPaiement"

    headers = [
        "ID",
        "IF",
        "ICE",
        "Fournisseur",
        "N° facture",
        "Date facture",
        "Nature",
        "Date début calcul",
        "Échéance légale",
        "Date paiement",
        "Montant",
        "Devise",
        "BC",
        "BR",
        "Statut",
    ]
    ws.append(headers)

    for inv in invs:
        amt = inv.amount_ttc if inv.amount_ttc is not None else inv.amount
        ws.append(
            [
                inv.id,
                inv.supplier_if or "",
                inv.supplier_ice or "",
                inv.supplier_name or "",
                inv.invoice_no or "",
                inv.invoice_date.date().isoformat() if inv.invoice_date else "",
                (inv.nature_operation or ""),
                inv.calc_start_date.date().isoformat() if inv.calc_start_date else "",
                inv.legal_due_date.date().isoformat() if inv.legal_due_date else "",
                inv.payment_date.date().isoformat() if inv.payment_date else "",
                float(amt or 0.0),
                (inv.currency or "MAD").upper(),
                inv.bc_no or "",
                inv.br_no or "",
                (inv.status or ""),
            ]
        )

    # format simple
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
    ws.freeze_panes = "A2"

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return Response(
        content=out.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=delais_paiement.xlsx"},
    )


# ---------------- DOCS LIST ----------------

@app.get("/docs_list", response_class=HTMLResponse)
def docs_page(request: Request, session: Session = Depends(get_session)):
    user = get_current_user(request, session)
    docs = session.exec(select(Document).order_by(Document.created_at.desc()).limit(200)).all()
    return templates.TemplateResponse("docs.html", {"request": request, "user": user, "docs": docs})


# ---------------- FACTURES ----------------

@app.get("/factures", response_class=HTMLResponse)
def factures_page(request: Request, session: Session = Depends(get_session)):
    user = get_current_user(request, session)
    invs = session.exec(select(Invoice).order_by(Invoice.id.desc()).limit(300)).all()
    return templates.TemplateResponse(
        "factures.html",
        {"request": request, "user": user, "invs": invs, "today": datetime.utcnow().date()},
    )


@app.get("/factures/{invoice_id}", response_class=HTMLResponse)
def facture_detail(invoice_id: int, request: Request, session: Session = Depends(get_session)):
    user = get_current_user(request, session)
    inv = session.get(Invoice, invoice_id)
    if not inv:
        raise HTTPException(status_code=404, detail="Facture introuvable")

    docs = session.exec(
        select(InvoiceDocument)
        .where(InvoiceDocument.invoice_id == invoice_id)
        .order_by(InvoiceDocument.created_at.desc())
    ).all()

    doc_id = request.query_params.get("doc_id")
    active_doc = None
    if doc_id:
        try:
            active_doc = session.get(InvoiceDocument, int(doc_id))
            if active_doc and active_doc.invoice_id != invoice_id:
                active_doc = None
        except Exception:
            active_doc = None

    if not active_doc and docs:
        active_doc = docs[0]

    try:
        rubriques = session.exec(select(CashflowRubrique).order_by(CashflowRubrique.rubrique.asc())).all()
    except Exception:
        rubriques = []

    return templates.TemplateResponse(
        "facture_detail.html",
        {"request": request, "user": user, "inv": inv, "docs": docs, "active_doc": active_doc, "rubriques": rubriques, "today": datetime.utcnow().date()},
    )


def _best_invoice_file_path(session: Session, invoice_id: int, inv: Invoice) -> Optional[str]:
    if inv and inv.file_path:
        return inv.file_path

    d = session.exec(
        select(InvoiceDocument)
        .where(InvoiceDocument.invoice_id == invoice_id)
        .where(InvoiceDocument.doc_type == "FACTURE")
        .order_by(InvoiceDocument.created_at.desc())
    ).first()
    if d:
        return d.stored_path
    return None



@app.get("/factures/{invoice_id}/open")
def facture_open(invoice_id: int, doc_id: Optional[int] = None, session: Session = Depends(get_session)):
    inv = session.get(Invoice, invoice_id)
    if not inv:
        raise HTTPException(status_code=404, detail="Facture introuvable")

    # Si un doc_id est fourni, ouvrir ce document précis (BC/BR/FACTURE/AUTRE)
    if doc_id:
        d = session.get(InvoiceDocument, doc_id)
        if not d or d.invoice_id != invoice_id:
            raise HTTPException(status_code=404, detail="Document introuvable")
        file_path = d.stored_path
    else:
        file_path = _best_invoice_file_path(session, invoice_id, inv)

    if not file_path:
        raise HTTPException(status_code=404, detail="Aucun fichier trouvé")

    disk_path = os.path.join(UPLOAD_DIR, file_path)
    if not os.path.exists(disk_path):
        raise HTTPException(status_code=404, detail="Fichier introuvable sur le serveur")

    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".pdf":
        media = "application/pdf"
    elif ext == ".png":
        media = "image/png"
    elif ext in (".jpg", ".jpeg"):
        media = "image/jpeg"
    elif ext == ".webp":
        media = "image/webp"
    else:
        media = "application/octet-stream"

    return Response(content=open(disk_path, "rb").read(), media_type=media, headers={"Content-Disposition": "inline"})


@app.get("/factures/{invoice_id}/download_pdf")

def facture_download_pdf(invoice_id: int, request: Request, session: Session = Depends(get_session)):
    inv = session.get(Invoice, invoice_id)
    if not inv:
        raise HTTPException(status_code=404, detail="Facture introuvable")

    docs = session.exec(
        select(InvoiceDocument)
        .where(InvoiceDocument.invoice_id == invoice_id)
        .order_by(InvoiceDocument.created_at.asc())
    ).all()

    order_rank = {"BC": 1, "BR": 2, "FACTURE": 3, "AUTRE": 4}
    docs_sorted = sorted(docs, key=lambda d: (order_rank.get((d.doc_type or "AUTRE").upper(), 9), d.created_at))

    paths: list[str] = []
    if inv.file_path:
        main_path = os.path.join(UPLOAD_DIR, inv.file_path)
        if os.path.exists(main_path):
            paths.append(main_path)

    for d in docs_sorted:
        p = os.path.join(UPLOAD_DIR, d.stored_path)
        if os.path.exists(p):
            if inv.file_path and d.stored_path == inv.file_path:
                continue
            paths.append(p)

    if not paths:
        raise HTTPException(status_code=404, detail="Aucun document à fusionner")
    if not PDF_MERGER_AVAILABLE:
        raise HTTPException(status_code=500, detail="Fusion en 1 seul PDF non disponible. Ajoute pypdf>=4.2.0 puis redeploy.")

    pdf_bytes = _merge_paths_to_one_pdf_bytes(paths)
    filename = f"dossier_facture_{invoice_id}.pdf"
    return Response(content=pdf_bytes, media_type="application/pdf", headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@app.post("/factures/{invoice_id}/delete")
def facture_delete(invoice_id: int, request: Request, session: Session = Depends(get_session)):
    # Optionnel: permet de revenir à la page d'origine après suppression.
    # Ex: /factures/12/delete?next=/delais
    next_url = (request.query_params.get("next") or "").strip() or "/factures"
    inv = session.get(Invoice, invoice_id)
    if not inv:
        return RedirectResponse(next_url, status_code=303)

    # --- Sécuriser la suppression: nettoyer toutes les tables dépendantes ---
    # Évite les erreurs 500 (contraintes FK) si la facture est liée à:
    # - matching bancaire (InvoicePaymentMatch)
    # - cashflow réel (CashflowActual)
    # - batch de règlements (PaymentLine)
    # - paiements manuels (Payment)
    try:
        session.exec(text("DELETE FROM invoicepaymentmatch WHERE invoice_id = :id"), {"id": invoice_id})
    except Exception:
        pass
    try:
        session.exec(text("DELETE FROM cashflowactual WHERE invoice_id = :id"), {"id": invoice_id})
    except Exception:
        pass
    try:
        session.exec(text("DELETE FROM paymentline WHERE invoice_id = :id"), {"id": invoice_id})
    except Exception:
        pass
    try:
        session.exec(text("DELETE FROM payment WHERE invoice_id = :id"), {"id": invoice_id})
    except Exception:
        pass
    try:
        session.exec(text("DELETE FROM allocationline WHERE invoice_id = :id"), {"id": invoice_id})
    except Exception:
        pass

    # ⚠️ Certains environnements DB ont une contrainte FK sur InvoiceDocument.
    # Comme le modèle peut évoluer, on sécurise aussi par SQL direct pour éviter
    # l'erreur 500 sur quelques factures (docs attachés).
    try:
        session.exec(text("DELETE FROM invoicedocument WHERE invoice_id = :id"), {"id": invoice_id})
    except Exception:
        pass

    # Nettoyage fichiers (best effort)
    docs = session.exec(select(InvoiceDocument).where(InvoiceDocument.invoice_id == invoice_id)).all()
    for d in docs:
        try:
            fp = os.path.join(UPLOAD_DIR, d.stored_path)
            if os.path.exists(fp):
                os.remove(fp)
        except Exception:
            pass
        try:
            session.delete(d)
        except Exception:
            pass

    try:
        if inv.file_path:
            fp = os.path.join(UPLOAD_DIR, inv.file_path)
            if os.path.exists(fp):
                os.remove(fp)
    except Exception:
        pass

    session.delete(inv)
    try:
        session.commit()
    except Exception:
        # Si une contrainte FK résiduelle existe, on rollback pour éviter
        # de laisser la session dans un état cassé.
        session.rollback()

        msg = (
            "Impossible de supprimer cette facture car elle est liée à d'autres données "
            "(règlements, matching bancaire, cashflow, affectations, documents...). "
            "Supprime d'abord les éléments liés ou annule la facture."
        )

        # Si appel en AJAX (fetch), on renvoie du JSON (la page reste ouverte).
        xrw = (request.headers.get("x-requested-with") or "").lower()
        accept = (request.headers.get("accept") or "").lower()
        if xrw == "fetch" or "application/json" in accept:
            raise HTTPException(status_code=400, detail=msg)

        # Sinon (submit normal), on redirige vers /factures avec un message.
        from urllib.parse import quote
        sep = "&" if ("?" in next_url) else "?"
        return RedirectResponse(f"{next_url}{sep}toast={quote(msg)}", status_code=303)
        # OK
    if ((request.headers.get("x-requested-with") or "").lower() == "fetch") or ("application/json" in (request.headers.get("accept") or "").lower()):
        return {"ok": True}
    return RedirectResponse(next_url, status_code=303)


# Empêche le "Internal Server Error" si quelqu'un ouvre /factures/{id}/delete au navigateur (GET)
@app.get("/factures/{invoice_id}/delete")
def facture_delete_get(invoice_id: int):
    # La suppression doit rester en POST (form). En GET, on redirige vers la fiche.
    return RedirectResponse(f"/factures/{invoice_id}", status_code=303)


def _clean_str(v: Optional[str], maxlen: Optional[int] = None) -> Optional[str]:
    if v is None:
        return None
    s = v.strip()
    if not s:
        return None
    if maxlen:
        s = s[:maxlen]
    return s


def _apply_str(current: Optional[str], v: Optional[str], maxlen: Optional[int] = None) -> Optional[str]:
    if v is None:
        return current
    return _clean_str(v, maxlen=maxlen)


def _apply_date(current: Optional[datetime], v: Optional[str]) -> Optional[datetime]:
    if v is None:
        return current
    s = v.strip()
    return _parse_date(s) if s else None


def _apply_int(current: Optional[int], v: Optional[str]) -> Optional[int]:
    if v is None:
        return current
    s = v.strip()
    return _to_int(s) if s else None


def _apply_float(current: Optional[float], v: Optional[str]) -> Optional[float]:
    if v is None:
        return current
    s = str(v).strip()
    if not s:
        return None
    s = s.replace(" ", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None



# ---------------- FOURNISSEUR = SOURCE DE VERITE (ERP) ----------------

def _norm_spaces(s: str) -> str:
    """Normalise un texte (trim + espaces multiples)."""
    s = (s or "").strip()
    if not s:
        return ""
    while "  " in s:
        s = s.replace("  ", " ")
    return s




def _norm_label(s: str) -> str:
    """Normalisation simple pour matching (uppercase, sans accents, sans ponctuation)."""
    s = (s or "").strip().upper()
    if not s:
        return ""
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    # ponctuation -> espaces
    for ch in ["-", "_", "/", "\\", ".", ",", ";", ":", "(", ")", "[", "]", "{", "}", "|", "’", "'", '"']:
        s = s.replace(ch, " ")
    # espaces multiples
    while "  " in s:
        s = s.replace("  ", " ")
    return s.strip()


def _txn_amount(txn: BankTxn) -> float:
    """Sortie (debit) positive, entrée (credit) positive."""
    d = float(txn.debit or 0.0)
    c = float(txn.credit or 0.0)
    # convention: montant cashflow réel = débit (sortie) ou crédit (entrée)
    return d if d > 0 else c


def _make_banktxn_dedup_key(bank_name: str, account_no: str, value_date: datetime, amount: float, label_norm: str) -> str:
    b = (bank_name or "").strip().upper()[:40]
    a = (account_no or "").strip().upper()[:40]
    vd = value_date.date().isoformat() if value_date else ""
    amt = f"{float(amount):.2f}"
    lab = (label_norm or "")[:60]
    return f"{b}|{a}|{vd}|{amt}|{lab}"
def upsert_supplier_from_invoice(session: Session, inv: Invoice) -> Optional[Supplier]:
    """
    ERP pro: Supplier = source de vérité.
    Crée / lie / complète la fiche fournisseur à partir des champs facture
    sans écraser des données déjà renseignées côté Supplier.
    """

    name = _norm_spaces(inv.supplier_name or "")[:200]
    ice = _clean_str(getattr(inv, "supplier_ice", None), 30)
    if_code = _clean_str(getattr(inv, "supplier_if", None), 30)
    rc = _clean_str(getattr(inv, "supplier_rc", None), 30)
    rc_city = _clean_str(getattr(inv, "supplier_rc_city", None), 60)
    address = _clean_str(getattr(inv, "supplier_address", None), 250)

    if not name and not ice:
        return None

    sup: Optional[Supplier] = None

    # 1) matching stable: ICE
    if ice:
        sup = session.exec(select(Supplier).where(Supplier.ice == ice)).first()

    # 2) fallback: nom exact
    if (not sup) and name:
        sup = session.exec(select(Supplier).where(Supplier.name == name)).first()

    if not sup:
        sup = Supplier(
            name=name or "A_COMPLETER",
            ice=ice,
            if_code=if_code,
            rc=rc,
            rc_city=rc_city,
            address=address,
            # par défaut: marocain (modifiable depuis la fiche fournisseur)
            is_foreign=False,
            country_code="MA",
        )
        session.add(sup)
        session.commit()
        session.refresh(sup)
    else:
        # ✅ on complète uniquement les champs vides
        if name and (not sup.name or sup.name == "A_COMPLETER"):
            sup.name = name
        if ice and not sup.ice:
            sup.ice = ice
        if if_code and not sup.if_code:
            sup.if_code = if_code
        if rc and not sup.rc:
            sup.rc = rc
        if rc_city and not sup.rc_city:
            sup.rc_city = rc_city
        if address and not sup.address:
            sup.address = address

        session.add(sup)
        session.commit()

    # lien facture -> fournisseur
    inv.supplier_id = sup.id

    # snapshot facture (uniquement si facture vide)
    if not inv.supplier_name:
        inv.supplier_name = sup.name
    if not getattr(inv, "supplier_ice", None):
        inv.supplier_ice = sup.ice
    if not getattr(inv, "supplier_if", None):
        inv.supplier_if = sup.if_code
    if not getattr(inv, "supplier_rc", None):
        inv.supplier_rc = sup.rc
    if not getattr(inv, "supplier_rc_city", None):
        inv.supplier_rc_city = sup.rc_city
    if not getattr(inv, "supplier_address", None):
        inv.supplier_address = sup.address

    return sup

@app.post("/factures/{invoice_id}/update")
def facture_update(
    invoice_id: int,
    request: Request,
    session: Session = Depends(get_session),
    supplier_name: Optional[str] = Form(None),
    invoice_no: Optional[str] = Form(None),
    invoice_date: Optional[str] = Form(None),
    amount: Optional[str] = Form(None),
    currency: Optional[str] = Form(None),
    department: Optional[str] = Form(None),
    analytic: Optional[str] = Form(None),
    bc_no: Optional[str] = Form(None),
    bc_date: Optional[str] = Form(None),
    br_no: Optional[str] = Form(None),
    br_date: Optional[str] = Form(None),
    due_date: Optional[str] = Form(None),
    due_date_planned: Optional[str] = Form(None),
    due_date_agreed: Optional[str] = Form(None),
    status: Optional[str] = Form(None),
    payment_date: Optional[str] = Form(None),
    payment_mode: Optional[str] = Form(None),
    supplier_if: Optional[str] = Form(None),
    supplier_ice: Optional[str] = Form(None),
    supplier_rc: Optional[str] = Form(None),
    supplier_rc_city: Optional[str] = Form(None),
    supplier_address: Optional[str] = Form(None),
    service_date: Optional[str] = Form(None),
    nature_operation: Optional[str] = Form(None),
    payment_terms_days: Optional[str] = Form(None),
    derogation_sector: Optional[str] = Form(None),
    derogation_days: Optional[str] = Form(None),
    derogation_ref: Optional[str] = Form(None),
    amount_ht: Optional[str] = Form(None),
    vat_rate: Optional[str] = Form(None),
    amount_vat: Optional[str] = Form(None),
    amount_ttc: Optional[str] = Form(None),
    cashflow_category: Optional[str] = Form(None),
    cashflow_rubrique: Optional[str] = Form(None),
    reporting_groupe: Optional[str] = Form(None),
    impact_budget: Optional[str] = Form(None),
    site: Optional[str] = Form(None),
    cost_center: Optional[str] = Form(None),
    project: Optional[str] = Form(None),
    gl_account: Optional[str] = Form(None),
    expense_nature: Optional[str] = Form(None),
    is_disputed: Optional[str] = Form(None),
    dispute_reason: Optional[str] = Form(None),
    disputed_amount: Optional[str] = Form(None),
):
    inv = session.get(Invoice, invoice_id)
    if not inv:
        raise HTTPException(status_code=404, detail="Facture introuvable")

    inv.supplier_name = _apply_str(inv.supplier_name, supplier_name, 200)
    inv.invoice_no = _apply_str(inv.invoice_no, invoice_no, 80) or inv.invoice_no or f"MANUAL-{invoice_id}"
    inv.invoice_date = _apply_date(inv.invoice_date, invoice_date)

    inv.cashflow_rubrique = _apply_str(getattr(inv, "cashflow_rubrique", None), cashflow_rubrique, 80)
    inv.reporting_groupe = _apply_str(getattr(inv, "reporting_groupe", None), reporting_groupe, 80)
    if impact_budget is not None:
        inv.impact_budget = (impact_budget == "on")

    if amount_ttc is not None:
        inv.amount_ttc = _to_float(amount_ttc)
        if inv.amount_ttc is not None:
            inv.amount = float(inv.amount_ttc)

    if amount is not None and amount_ttc is None:
        inv.amount = _to_float(amount) or (inv.amount or 0.0)

    if currency is not None:
        inv.currency = _normalize_currency(currency) if _clean_str(currency) else inv.currency

    inv.department = _apply_str(inv.department, department, 80)
    inv.analytic = _apply_str(inv.analytic, analytic, 80)

    inv.bc_no = _apply_str(inv.bc_no, bc_no, 80)
    inv.bc_date = _apply_date(getattr(inv, "bc_date", None), bc_date)

    inv.br_no = _apply_str(getattr(inv, "br_no", None), br_no, 80)
    inv.br_date = _apply_date(getattr(inv, "br_date", None), br_date)

    inv.due_date = _apply_date(inv.due_date, due_date)
    inv.due_date_planned = _apply_date(inv.due_date_planned, due_date_planned)
    inv.due_date_agreed = _apply_date(inv.due_date_agreed, due_date_agreed)

    if status is not None:
        inv.status = (_clean_str(status, 20) or inv.status or "A_PAYER").upper()

    inv.payment_date = _apply_date(inv.payment_date, payment_date)
    inv.payment_mode = _apply_str(inv.payment_mode, payment_mode, 60)

    inv.supplier_if = _apply_str(inv.supplier_if, supplier_if, 30)
    inv.supplier_ice = _apply_str(inv.supplier_ice, supplier_ice, 30)
    inv.supplier_rc = _apply_str(inv.supplier_rc, supplier_rc, 30)
    inv.supplier_rc_city = _apply_str(inv.supplier_rc_city, supplier_rc_city, 60)
    inv.supplier_address = _apply_str(inv.supplier_address, supplier_address, 250)

    # ✅ ERP pro: synchronise fiche fournisseur (sans écraser)
    upsert_supplier_from_invoice(session, inv)

    inv.service_date = _apply_date(inv.service_date, service_date)
    if nature_operation is not None:
        inv.nature_operation = (_clean_str(nature_operation, 20) or "").upper() or None

    inv.payment_terms_days = _apply_int(inv.payment_terms_days, payment_terms_days)

    if derogation_sector is not None:
        inv.derogation_sector = (derogation_sector == "on")
    inv.derogation_days = _apply_int(inv.derogation_days, derogation_days)
    inv.derogation_ref = _apply_str(inv.derogation_ref, derogation_ref, 120)

    inv.amount_ht = _apply_float(inv.amount_ht, amount_ht)
    inv.vat_rate = _apply_float(inv.vat_rate, vat_rate)
    inv.amount_vat = _apply_float(inv.amount_vat, amount_vat)

    inv.cashflow_category = _apply_str(inv.cashflow_category, cashflow_category, 60)
    inv.site = _apply_str(inv.site, site, 30)
    inv.cost_center = _apply_str(inv.cost_center, cost_center, 60)
    inv.project = _apply_str(inv.project, project, 80)
    inv.gl_account = _apply_str(inv.gl_account, gl_account, 20)
    inv.expense_nature = _apply_str(inv.expense_nature, expense_nature, 60)

    if is_disputed is not None:
        inv.is_disputed = (is_disputed == "on")
    inv.dispute_reason = _apply_str(inv.dispute_reason, dispute_reason, 200)
    inv.disputed_amount = _apply_float(inv.disputed_amount, disputed_amount)

    compute_pay_delay_fields(inv)

    session.add(inv)
    session.commit()
    return RedirectResponse(f"/factures/{invoice_id}", status_code=303)



# ---------------- DUPLICATES (factures / BC / BR) ----------------

def _norm_key(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"[^a-z0-9 ]+", "", s)
    return s

def _find_duplicate_invoice(
    session: Session,
    invoice_no: str,
    supplier_name: str | None,
    supplier_ice: str | None,
) -> Optional["Invoice"]:
    """Détecte un doublon probable (même invoice_no + même fournisseur/ICE)."""
    inv_no = (invoice_no or "").strip()
    if not inv_no:
        return None
    if inv_no.upper().startswith("SCAN-"):
        return None

    candidates = session.exec(select(Invoice).where(Invoice.invoice_no == inv_no)).all()
    if not candidates:
        return None

    ice = (supplier_ice or "").strip()
    sname = _norm_key(supplier_name or "")
    for c in candidates:
        if ice and (c.supplier_ice or "").strip() == ice:
            return c
        if sname and _norm_key(getattr(c, "supplier_name", "") or "") == sname:
            return c
    return candidates[0]

def _find_duplicate_doc(
    session: Session,
    doc_type: str,
    ref_no: str | None,
) -> Optional["InvoiceDocument"]:
    r = (ref_no or "").strip()
    if not r:
        return None
    return session.exec(
        select(InvoiceDocument).where(
            InvoiceDocument.doc_type == doc_type,
            InvoiceDocument.ref_no == r,
        )
    ).first()


def _ensure_invoice_document_facture(
    session: Session,
    inv: "Invoice",
    stored_path: str | None,
    filename: str | None = None,
) -> None:
    """Garantit qu'une facture issue d'un scan possède un InvoiceDocument FACTURE.

    Les écrans Analytique/Budget filtrent les factures "affectables" via la présence
    d'un InvoiceDocument de type BR/FACTURE (ou via br_no). Si une facture est créée
    via le scan simple, sans InvoiceDocument, elle n'apparaît pas.

    On évite les doublons (même invoice_id + doc_type + stored_path).
    """
    try:
        if not inv or not getattr(inv, "id", None):
            return
        if not stored_path:
            return

        existing = session.exec(
            select(InvoiceDocument).where(
                (InvoiceDocument.invoice_id == inv.id)
                & (InvoiceDocument.doc_type == "FACTURE")
                & (InvoiceDocument.stored_path == stored_path)
            )
        ).first()
        if existing:
            return

        fname = (filename or stored_path or "scan.pdf")[:255]
        session.add(
            InvoiceDocument(
                invoice_id=int(inv.id),
                doc_type="FACTURE",
                filename=fname,
                stored_path=(stored_path or "")[:255],
                ref_no=(getattr(inv, "invoice_no", None) or None),
                ref_date=(getattr(inv, "invoice_date", None) or None),
            )
        )
    except Exception:
        # ne jamais bloquer le flux facture pour un souci de pièces
        return

def _finalize_invoice_fields(inv: "Invoice") -> None:
    """Complète des champs calculables si l'extraction n'a pas tout rempli."""
    try:
        ttc = float(inv.amount_ttc) if inv.amount_ttc is not None else None
    except Exception:
        ttc = None

    try:
        ht = float(inv.amount_ht) if inv.amount_ht is not None else None
    except Exception:
        ht = None

    try:
        vat = float(inv.amount_vat) if inv.amount_vat is not None else None
    except Exception:
        vat = None

    try:
        rate = float(inv.vat_rate) if inv.vat_rate is not None else None
    except Exception:
        rate = None

    if rate is not None and rate <= 1.0:
        rate = rate * 100.0
        inv.vat_rate = round(rate, 2)

    r = (rate / 100.0) if rate is not None else None

    if ttc is None and ht is not None and vat is not None:
        ttc = round(ht + vat, 2)
        inv.amount_ttc = ttc
    if vat is None and ht is not None and ttc is not None:
        vat = round(ttc - ht, 2)
        inv.amount_vat = vat
    if ht is None and ttc is not None and r is not None:
        ht = round(ttc / (1 + r), 2)
        inv.amount_ht = ht
    if vat is None and ht is not None and r is not None:
        vat = round(ht * r, 2)
        inv.amount_vat = vat
    if ttc is None and ht is not None and r is not None:
        ttc = round(ht * (1 + r), 2)
        inv.amount_ttc = ttc
    if rate is None and ht not in (None, 0.0) and vat is not None:
        inv.vat_rate = round((vat / ht) * 100.0, 2)

    if (inv.amount is None or float(inv.amount or 0.0) == 0.0):
        base_amt = inv.amount_ttc if inv.amount_ttc is not None else (inv.amount_ht if inv.amount_ht is not None else None)
        if base_amt is not None:
            try:
                inv.amount = float(base_amt)
            except Exception:
                pass

    if inv.due_date is None and inv.invoice_date is not None and inv.payment_terms_days:
        try:
            inv.due_date = inv.invoice_date + timedelta(days=int(inv.payment_terms_days))
        except Exception:
            pass

    if not getattr(inv, "currency", None):
        inv.currency = "MAD"


# ✅ IMPORT UNIQUE : 1 fichier => facture simple | plusieurs => dossier
@app.post("/factures/scan")
async def factures_scan(
    request: Request,
    session: Session = Depends(get_session),
    files: Optional[list[UploadFile]] = File(None),
    file: Optional[UploadFile] = File(None),
):
    """Import Achats.

    - 1 seul fichier : on classe d'abord (BC/BR/FACTURE). On ne remplit les champs facture
      que si c'est une FACTURE.
    - Plusieurs fichiers : on crée un dossier (Invoice INCOMPLET) et on attache les docs.
    """
    if not files:
        files = []
    if file is not None:
        files = [file]

    if not files:
        return RedirectResponse("/factures", status_code=303)

    # ---------------- 1 fichier ----------------
    if len(files) == 1:
        up = files[0]
        ts = int(datetime.now().timestamp())
        original = up.filename or "piece"
        safe_name = f"{ts}_{original}".replace("/", "_").replace("\\", "_")
        disk_path = os.path.join(UPLOAD_DIR, safe_name)

        content = await up.read()
        sha1 = hashlib.sha1(content).hexdigest()
        with open(disk_path, "wb") as f:
            f.write(content)

        # 1) Classifier BC/BR/FACTURE (et extraire le minimum)
        try:
            bundle = extract_doc_bundle_with_openai(disk_path, up.content_type or "")
        except Exception as e:
            print("BUNDLE OPENAI ERROR =", repr(e))
            bundle = {"doc_type": "AUTRE"}

        doc_type = (bundle.get("doc_type") or "AUTRE").strip().upper()
        if doc_type not in ("FACTURE", "BC", "BR"):
            doc_type = "AUTRE"

        # 2) Si FACTURE => extraction détaillée
        data = bundle
        if doc_type == "FACTURE":
            try:
                data = extract_invoice_with_openai(disk_path, up.content_type or "")
                data["doc_type"] = "FACTURE"
            except Exception as e:
                print("OPENAI INVOICE ERROR =", repr(e))
                data = bundle

        supplier = (data.get("supplier_name") or "A_COMPLETER").strip()[:200]

        # Champs facture (uniquement si FACTURE)
        inv_no = (data.get("invoice_no") or "").strip()[:80] or None
        inv_date = _parse_date(data.get("invoice_date")) if doc_type == "FACTURE" else None
        currency = _normalize_currency(data.get("currency")) or "MAD"
        amount_ttc = _to_float(data.get("amount_ttc")) if doc_type == "FACTURE" else None
        amount_value = float(amount_ttc) if amount_ttc is not None else 0.0

        # Champs BC/BR
        bc_no = (data.get("bc_no") or "").strip()[:80] or None
        br_no = (data.get("br_no") or "").strip()[:80] or None
        service_date = _parse_date(data.get("service_date")) if doc_type == "BR" else None
        bc_date = _parse_date(data.get("bc_date")) if doc_type == "BC" else None
        br_date = _parse_date(data.get("br_date")) if doc_type == "BR" else None

        department = (data.get("department") or "").strip()[:80] or None
        analytic = (data.get("analytic") or "").strip()[:80] or None

        # --- doublons docs BC/BR ---
        duplicate_docs_scan: list[dict] = []
        try:
            if doc_type == "BC" and bc_no:
                d = _find_duplicate_doc(session, "BC", bc_no)
                if d:
                    duplicate_docs_scan.append({"kind": "BC", "ref_no": bc_no, "existing_invoice_id": d.invoice_id})
        except Exception:
            pass
        try:
            if doc_type == "BR" and br_no:
                d = _find_duplicate_doc(session, "BR", br_no)
                if d:
                    duplicate_docs_scan.append({"kind": "BR", "ref_no": br_no, "existing_invoice_id": d.invoice_id})
        except Exception:
            pass

        # Anti-doublon dur (même upload)
        dedup_key = f"scan_sha1|{sha1}"
        exists = session.exec(select(Invoice).where(Invoice.dedup_key == dedup_key)).first()
        if exists:
            return RedirectResponse("/factures", status_code=303)

        # Doublon probable facture => confirmation
        if doc_type == "FACTURE" and inv_no:
            dup = _find_duplicate_invoice(
                session=session,
                invoice_no=inv_no,
                supplier_name=supplier,
                supplier_ice=(data.get("supplier_ice") or None),
            )
            if dup:
                payload_obj = {
                    "safe_name": safe_name,
                    "data": data,
                    "content_type": (up.content_type or ""),
                    "doc_type": doc_type,
                }
                payload_b64 = base64.urlsafe_b64encode(json.dumps(payload_obj, default=str).encode("utf-8")).decode("utf-8")
                return templates.TemplateResponse(
                    "confirm_duplicate.html",
                    {
                        "request": request,
                        "mode": "invoice_scan",
                        "payload": payload_b64,
                        "safe_name": safe_name,
                        "duplicate": dup,
                        "duplicate_docs_scan": duplicate_docs_scan,
                        "candidate": {
                            "supplier_name": supplier,
                            "supplier_ice": (data.get("supplier_ice") or None),
                            "invoice_no": inv_no,
                            "invoice_date": inv_date,
                            "amount": amount_value,
                            "currency": currency,
                        },
                    },
                )

        # Doublon BC/BR => confirmation
        if duplicate_docs_scan:
            payload_obj = {
                "safe_name": safe_name,
                "data": data,
                "content_type": (up.content_type or ""),
                "doc_type": doc_type,
            }
            payload_b64 = base64.urlsafe_b64encode(json.dumps(payload_obj, default=str).encode("utf-8")).decode("utf-8")
            return templates.TemplateResponse(
                "confirm_duplicate.html",
                {
                    "request": request,
                    "mode": "invoice_scan",
                    "payload": payload_b64,
                    "safe_name": safe_name,
                    "duplicate": None,
                    "duplicate_docs_scan": duplicate_docs_scan,
                    "candidate": {
                        "supplier_name": supplier,
                        "supplier_ice": (data.get("supplier_ice") or None),
                        "invoice_no": inv_no,
                        "invoice_date": inv_date,
                        "amount": amount_value,
                        "currency": currency,
                    },
                },
            )

        # --- Créer la ligne Achats ---
        status = "A_PAYER" if doc_type == "FACTURE" else ("BC_RECU" if doc_type == "BC" else ("BR_RECU" if doc_type == "BR" else "INCOMPLET"))

        inv = Invoice(
            supplier_name=supplier,
            supplier_ice=(data.get("supplier_ice") or None),
            supplier_if=(data.get("supplier_if") or None),
            supplier_rc=(data.get("supplier_rc") or None),
            supplier_rc_city=(data.get("supplier_rc_city") or None),
            supplier_address=(data.get("supplier_address") or None),

            invoice_no=(inv_no if doc_type == "FACTURE" else None),
            invoice_date=inv_date,
            reception_date=_parse_date(data.get("reception_date")) if doc_type == "FACTURE" else None,

            service_date=service_date,

            amount=(amount_value if doc_type == "FACTURE" else 0.0),
            amount_ht=_to_float(data.get("amount_ht")) if doc_type == "FACTURE" else None,
            amount_vat=_to_float(data.get("amount_vat")) if doc_type == "FACTURE" else None,
            amount_ttc=_to_float(data.get("amount_ttc")) if doc_type == "FACTURE" else None,
            vat_rate=_to_float(data.get("vat_rate")) if doc_type == "FACTURE" else None,

            currency=currency,
            status=status,
            dedup_key=dedup_key,
            department=department,
            analytic=analytic,

            bc_no=bc_no,
            bc_date=bc_date,
            br_no=br_no,
            br_date=br_date,

            payment_terms_days=(int(data.get("payment_terms_days")) if str(data.get("payment_terms_days") or "").isdigit() else None),
            nature_operation=(data.get("nature_operation") or None),

            file_path=(safe_name if doc_type == "FACTURE" else None),
        )

        upsert_supplier_from_invoice(session, inv)
        _finalize_invoice_fields(inv)

        session.add(inv)
        session.commit()
        session.refresh(inv)

        # Enregistrer le document
        ref_no = None
        ref_date = None
        if doc_type == "BC":
            ref_no = bc_no
            ref_date = bc_date
        elif doc_type == "BR":
            ref_no = br_no
            ref_date = br_date or service_date
        elif doc_type == "FACTURE":
            ref_no = inv_no
            ref_date = inv_date

        session.add(
            InvoiceDocument(
                invoice_id=inv.id,
                doc_type=(doc_type if doc_type in ("FACTURE", "BC", "BR") else "AUTRE"),
                filename=(up.filename or safe_name)[:255],
                stored_path=safe_name,
                ref_no=(ref_no[:80] if isinstance(ref_no, str) else ref_no),
                ref_date=ref_date,
            )
        )

        # Calcul échéance uniquement si point de départ fiable
        if doc_type in ("FACTURE", "BR") and (inv.invoice_date or inv.service_date):
            compute_pay_delay_fields(inv)
            session.add(inv)

        # FACTURE => document principal
        if doc_type == "FACTURE":
            _ensure_invoice_document_facture(session, inv, safe_name, filename=(up.filename or safe_name))

        session.commit()
        return RedirectResponse("/factures", status_code=303)

    # ---------------- plusieurs fichiers : dossier ----------------
    ts = int(datetime.now().timestamp())
    inv = Invoice(
        supplier_name="A_COMPLETER",
        invoice_no=None,
        status="INCOMPLET",
        dedup_key=f"bundle|{ts}",
        currency="MAD",
        amount=0.0,
    )
    session.add(inv)
    session.commit()
    session.refresh(inv)

    items: list[dict] = []
    dups: list[dict] = []

    for up in files:
        content = await up.read()
        stored = _safe_store_upload(content, up.filename or "piece")
        disk_path = os.path.join(UPLOAD_DIR, stored)

        try:
            data = extract_doc_bundle_with_openai(disk_path, up.content_type or "")
        except Exception as e:
            print("BUNDLE OPENAI ERROR =", repr(e))
            data = {"doc_type": "AUTRE"}

        doc_type = (data.get("doc_type") or "AUTRE").strip().upper()
        if doc_type not in ("FACTURE", "BC", "BR"):
            doc_type = "AUTRE"

        ref_no = None
        ref_date = None
        if doc_type == "BC":
            ref_no = data.get("bc_no")
            ref_date = _parse_date(data.get("bc_date"))
        elif doc_type == "BR":
            ref_no = data.get("br_no")
            ref_date = _parse_date(data.get("br_date")) or _parse_date(data.get("service_date"))
        elif doc_type == "FACTURE":
            ref_no = data.get("invoice_no")
            ref_date = _parse_date(data.get("invoice_date"))

        # --- détection doublons ---
        if doc_type in ("BC", "BR") and ref_no:
            doc_dup = _find_duplicate_doc(session, doc_type, str(ref_no).strip()[:80])
            if doc_dup:
                dups.append(
                    {"kind": doc_type, "ref_no": str(ref_no).strip()[:80], "existing_invoice_id": doc_dup.invoice_id, "existing_doc_id": doc_dup.id}
                )

        if doc_type == "FACTURE" and ref_no:
            supplier_guess = (data.get("supplier_name") or inv.supplier_name or "").strip()[:200]
            ice_guess = (data.get("supplier_ice") or inv.supplier_ice or None)
            inv_dup = _find_duplicate_invoice(session, str(ref_no).strip()[:80], supplier_guess, ice_guess)
            if inv_dup:
                dups.append(
                    {"kind": "FACTURE", "ref_no": str(ref_no).strip()[:80], "existing_invoice_id": inv_dup.id, "existing_doc_id": None}
                )

        items.append(
            {
                "stored": stored,
                "filename": (up.filename or stored)[:255],
                "doc_type": doc_type,
                "ref_no": (str(ref_no).strip()[:80] if ref_no else None),
                "ref_date": (ref_date.isoformat() if ref_date else None),
                "data": data,
            }
        )

    if dups:
        payload_obj = {"invoice_id": inv.id, "items": items}
        payload_b64 = base64.urlsafe_b64encode(json.dumps(payload_obj, default=str).encode("utf-8")).decode("utf-8")
        return templates.TemplateResponse(
            "confirm_duplicate.html",
            {"request": request, "mode": "scan_bundle", "payload": payload_b64, "duplicate_docs": dups},
        )

    for it in items:
        doc_type = it["doc_type"]
        ref_no = it.get("ref_no")
        ref_date = _parse_date(it.get("ref_date")) if it.get("ref_date") else None

        session.add(
            InvoiceDocument(
                invoice_id=inv.id,
                doc_type=doc_type,
                filename=it["filename"],
                stored_path=it["stored"],
                ref_no=ref_no,
                ref_date=ref_date,
            )
        )

        if doc_type == "FACTURE" and not inv.file_path:
            inv.file_path = it["stored"]

        _merge_extraction_into_invoice(inv, it["data"] or {})

    upsert_supplier_from_invoice(session, inv)
    _finalize_invoice_fields(inv)
    if inv.invoice_date or inv.service_date:
        compute_pay_delay_fields(inv)

    session.add(inv)
    session.commit()
    return RedirectResponse(f"/factures/{inv.id}", status_code=303)


@app.post("/factures/scan_commit")
def factures_scan_commit(
    request: Request,
    session: Session = Depends(get_session),
    payload: str = Form(...),
    force: str = Form("0"),
):
    """Commit du scan après confirmation doublon (facture ou BC/BR)."""
    try:
        raw = base64.urlsafe_b64decode(payload.encode("utf-8"))
        obj = json.loads(raw.decode("utf-8"))
    except Exception:
        return RedirectResponse("/factures", status_code=303)

    safe_name = obj.get("safe_name")
    data = obj.get("data") or {}
    doc_type = (obj.get("doc_type") or data.get("doc_type") or "FACTURE").strip().upper()
    if doc_type not in ("FACTURE", "BC", "BR"):
        doc_type = "AUTRE"

    # anti-doublon dur (même upload)
    sha1 = None
    try:
        if safe_name:
            fp = os.path.join(UPLOAD_DIR, safe_name)
            if os.path.exists(fp):
                with open(fp, "rb") as f:
                    sha1 = hashlib.sha1(f.read()).hexdigest()
    except Exception:
        sha1 = None

    dedup_key = f"scan_sha1|{sha1 or safe_name or int(datetime.now().timestamp())}"
    exists = session.exec(select(Invoice).where(Invoice.dedup_key == dedup_key)).first()
    if exists:
        return RedirectResponse("/factures", status_code=303)

    supplier = (data.get("supplier_name") or "A_COMPLETER").strip()[:200]
    currency = _normalize_currency(data.get("currency")) or "MAD"

    inv_no = (data.get("invoice_no") or "").strip()[:80] or None
    inv_date = _parse_date(data.get("invoice_date")) if doc_type == "FACTURE" else None
    amount_ttc = _to_float(data.get("amount_ttc")) if doc_type == "FACTURE" else None
    amount_value = float(amount_ttc) if amount_ttc is not None else 0.0

    bc_no = (data.get("bc_no") or "").strip()[:80] or None
    br_no = (data.get("br_no") or "").strip()[:80] or None
    bc_date = _parse_date(data.get("bc_date")) if doc_type == "BC" else None
    service_date = _parse_date(data.get("service_date")) if doc_type == "BR" else None
    br_date = _parse_date(data.get("br_date")) if doc_type == "BR" else None

    department = (data.get("department") or "").strip()[:80] or None
    analytic = (data.get("analytic") or "").strip()[:80] or None

    # si pas "force", on refuse quand même le doublon probable sur facture (si n° existe)
    if force != "1" and doc_type == "FACTURE" and inv_no:
        dup = _find_duplicate_invoice(
            session=session,
            invoice_no=inv_no,
            supplier_name=supplier,
            supplier_ice=(data.get("supplier_ice") or None),
        )
        if dup:
            return RedirectResponse("/factures", status_code=303)

    status = "A_PAYER" if doc_type == "FACTURE" else ("BC_RECU" if doc_type == "BC" else ("BR_RECU" if doc_type == "BR" else "INCOMPLET"))

    inv = Invoice(
        supplier_name=supplier,
        supplier_if=(data.get("supplier_if") or None),
        supplier_ice=(data.get("supplier_ice") or None),
        supplier_rc=(data.get("supplier_rc") or None),
        supplier_rc_city=(data.get("supplier_rc_city") or None),
        supplier_address=(data.get("supplier_address") or None),

        invoice_no=(inv_no if doc_type == "FACTURE" else None),
        invoice_date=inv_date,
        reception_date=_parse_date(data.get("reception_date")) if doc_type == "FACTURE" else None,

        service_date=service_date,
        bc_no=bc_no,
        bc_date=bc_date,
        br_no=br_no,
        br_date=br_date,

        payment_terms_days=_to_int(data.get("payment_terms_days")),
        nature_operation=((data.get("nature_operation") or "").strip().upper() or None),

        amount=(amount_value if doc_type == "FACTURE" else 0.0),
        amount_ht=_to_float(data.get("amount_ht")) if doc_type == "FACTURE" else None,
        vat_rate=_to_float(data.get("vat_rate")) if doc_type == "FACTURE" else None,
        amount_vat=_to_float(data.get("amount_vat")) if doc_type == "FACTURE" else None,
        amount_ttc=amount_ttc if doc_type == "FACTURE" else None,

        currency=currency,
        status=status,
        dedup_key=dedup_key,
        department=department,
        analytic=analytic,

        file_path=(safe_name if doc_type == "FACTURE" else None),
    )

    upsert_supplier_from_invoice(session, inv)
    _finalize_invoice_fields(inv)

    session.add(inv)
    session.commit()
    session.refresh(inv)

    # Document lié
    ref_no = None
    ref_date = None
    if doc_type == "BC":
        ref_no = bc_no
        ref_date = bc_date
    elif doc_type == "BR":
        ref_no = br_no
        ref_date = br_date or service_date
    elif doc_type == "FACTURE":
        ref_no = inv_no
        ref_date = inv_date

    session.add(
        InvoiceDocument(
            invoice_id=inv.id,
            doc_type=(doc_type if doc_type in ("FACTURE", "BC", "BR") else "AUTRE"),
            filename=(safe_name or "")[:255],
            stored_path=(safe_name or ""),
            ref_no=(ref_no[:80] if isinstance(ref_no, str) else ref_no),
            ref_date=ref_date,
        )
    )

    if doc_type in ("FACTURE", "BR") and (inv.invoice_date or inv.service_date):
        compute_pay_delay_fields(inv)
        session.add(inv)

    # Affectable uniquement si FACTURE
    if doc_type == "FACTURE" and safe_name:
        _ensure_invoice_document_facture(session, inv, safe_name, filename=safe_name)

    session.commit()
    return RedirectResponse("/factures", status_code=303)
@app.get("/factures/scan_cancel")
def factures_scan_cancel(safe_name: str, request: Request):
    """Annule un scan en supprimant le fichier uploadé."""
    try:
        fp = os.path.join(UPLOAD_DIR, safe_name)
        if os.path.exists(fp):
            os.remove(fp)
    except Exception:
        pass
    return RedirectResponse("/factures", status_code=303)

@app.post("/factures/scan_bundle_commit")
def factures_scan_bundle_commit(
    request: Request,
    session: Session = Depends(get_session),
    payload: str = Form(...),
    force: str = Form("0"),
):
    """Commit d'un import dossier après confirmation doublon."""
    try:
        raw = base64.urlsafe_b64decode(payload.encode("utf-8"))
        obj = json.loads(raw.decode("utf-8"))
        invoice_id = int(obj.get("invoice_id"))
        items = obj.get("items") or []
    except Exception:
        return RedirectResponse("/factures", status_code=303)

    inv = session.get(Invoice, invoice_id)
    if not inv:
        # si la facture n'existe plus, nettoyer les fichiers si possible
        for it in items:
            try:
                fp = os.path.join(UPLOAD_DIR, it.get("stored") or "")
                if fp and os.path.exists(fp):
                    os.remove(fp)
            except Exception:
                pass
        return RedirectResponse("/factures", status_code=303)

    # si pas force => annuler (et nettoyer)
    if force != "1":
        for it in items:
            try:
                fp = os.path.join(UPLOAD_DIR, it.get("stored") or "")
                if fp and os.path.exists(fp):
                    os.remove(fp)
            except Exception:
                pass
        try:
            session.delete(inv)
            session.commit()
        except Exception:
            pass
        return RedirectResponse("/factures", status_code=303)

    # commit documents + merge
    for it in items:
        doc_type = it.get("doc_type") or "AUTRE"
        ref_no = it.get("ref_no")
        ref_date = _parse_date(it.get("ref_date")) if it.get("ref_date") else None

        session.add(
            InvoiceDocument(
                invoice_id=inv.id,
                doc_type=doc_type,
                filename=(it.get("filename") or it.get("stored") or "")[:255],
                stored_path=it.get("stored") or "",
                ref_no=ref_no,
                ref_date=ref_date,
            )
        )

        if doc_type == "FACTURE" and not inv.file_path:
            inv.file_path = it.get("stored") or inv.file_path

        _merge_extraction_into_invoice(inv, it.get("data") or {})

    upsert_supplier_from_invoice(session, inv)
    _finalize_invoice_fields(inv)
    compute_pay_delay_fields(inv)

    session.add(inv)
    session.commit()
    return RedirectResponse(f"/factures/{inv.id}", status_code=303)


@app.get("/factures/scan_bundle_cancel")
def factures_scan_bundle_cancel(payload: str, request: Request, session: Session = Depends(get_session)):
    """Annule un import dossier: supprime fichiers + facture vide."""
    try:
        raw = base64.urlsafe_b64decode(payload.encode("utf-8"))
        obj = json.loads(raw.decode("utf-8"))
        invoice_id = int(obj.get("invoice_id"))
        items = obj.get("items") or []
    except Exception:
        return RedirectResponse("/factures", status_code=303)

    for it in items:
        try:
            fp = os.path.join(UPLOAD_DIR, it.get("stored") or "")
            if fp and os.path.exists(fp):
                os.remove(fp)
        except Exception:
            pass

    inv = session.get(Invoice, invoice_id)
    if inv:
        try:
            session.delete(inv)
            session.commit()
        except Exception:
            pass

    return RedirectResponse("/factures", status_code=303)




@app.post("/factures/{invoice_id}/add_docs")
async def factures_add_docs(
    invoice_id: int,
    request: Request,
    files: list[UploadFile] = File(...),
    session: Session = Depends(get_session),
):
    inv = session.get(Invoice, invoice_id)
    if not inv:
        raise HTTPException(status_code=404, detail="Facture introuvable")

    if not files:
        return RedirectResponse(f"/factures/{invoice_id}", status_code=303)

    items: list[dict] = []
    dups: list[dict] = []

    for up in files:
        content = await up.read()
        stored = _safe_store_upload(content, up.filename or "piece")
        disk_path = os.path.join(UPLOAD_DIR, stored)

        try:
            data = extract_doc_bundle_with_openai(disk_path, up.content_type or "")
        except Exception as e:
            print("ADD_DOCS OPENAI ERROR =", repr(e))
            data = {"doc_type": "AUTRE"}

        doc_type = (data.get("doc_type") or "AUTRE").strip().upper()
        if doc_type not in ("FACTURE", "BC", "BR"):
            doc_type = "AUTRE"

        ref_no = None
        ref_date = None
        if doc_type == "BC":
            ref_no = data.get("bc_no")
            ref_date = _parse_date(data.get("bc_date"))
        elif doc_type == "BR":
            ref_no = data.get("br_no")
            ref_date = _parse_date(data.get("br_date")) or _parse_date(data.get("service_date"))
        elif doc_type == "FACTURE":
            ref_no = data.get("invoice_no")
            ref_date = _parse_date(data.get("invoice_date"))

        # --- détection doublons ---
        if doc_type in ("BC", "BR") and ref_no:
            doc_dup = _find_duplicate_doc(session, doc_type, str(ref_no).strip()[:80])
            if doc_dup:
                dups.append(
                    {
                        "kind": doc_type,
                        "ref_no": str(ref_no).strip()[:80],
                        "existing_invoice_id": doc_dup.invoice_id,
                        "existing_doc_id": doc_dup.id,
                    }
                )

        if doc_type == "FACTURE" and ref_no:
            supplier_guess = (data.get("supplier_name") or inv.supplier_name or "").strip()[:200]
            ice_guess = (data.get("supplier_ice") or inv.supplier_ice or None)
            inv_dup = _find_duplicate_invoice(session, str(ref_no).strip()[:80], supplier_guess, ice_guess)
            if inv_dup:
                dups.append(
                    {
                        "kind": "FACTURE",
                        "ref_no": str(ref_no).strip()[:80],
                        "existing_invoice_id": inv_dup.id,
                        "existing_doc_id": None,
                    }
                )

        items.append(
            {
                "stored": stored,
                "filename": (up.filename or stored)[:255],
                "doc_type": doc_type,
                "ref_no": (str(ref_no).strip()[:80] if ref_no else None),
                "ref_date": (ref_date.isoformat() if ref_date else None),
                "data": data,
            }
        )

    if dups:
        payload_obj = {"invoice_id": invoice_id, "items": items}
        payload_b64 = base64.urlsafe_b64encode(json.dumps(payload_obj, default=str).encode("utf-8")).decode("utf-8")
        return templates.TemplateResponse(
            "confirm_duplicate.html",
            {
                "request": request,
                "mode": "add_docs",
                "payload": payload_b64,
                "invoice_id": invoice_id,
                "duplicate_docs": dups,
            },
        )

    # sinon commit direct
    for it in items:
        doc_type = it["doc_type"]
        ref_no = it.get("ref_no")
        ref_date = _parse_date(it.get("ref_date")) if it.get("ref_date") else None

        session.add(
            InvoiceDocument(
                invoice_id=inv.id,
                doc_type=doc_type,
                filename=it["filename"],
                stored_path=it["stored"],
                ref_no=ref_no,
                ref_date=ref_date,
            )
        )

        if doc_type == "FACTURE" and not inv.file_path:
            inv.file_path = it["stored"]

        _merge_extraction_into_invoice(inv, it["data"] or {})

    _finalize_invoice_fields(inv)
    compute_pay_delay_fields(inv)

    session.add(inv)
    session.commit()
    return RedirectResponse(f"/factures/{invoice_id}", status_code=303)


@app.post("/factures/{invoice_id}/add_docs_commit")
def factures_add_docs_commit(
    invoice_id: int,
    request: Request,
    session: Session = Depends(get_session),
    payload: str = Form(...),
    force: str = Form("0"),
):
    inv = session.get(Invoice, invoice_id)
    if not inv:
        return RedirectResponse("/factures", status_code=303)

    try:
        raw = base64.urlsafe_b64decode(payload.encode("utf-8"))
        obj = json.loads(raw.decode("utf-8"))
        items = obj.get("items") or []
    except Exception:
        return RedirectResponse(f"/factures/{invoice_id}", status_code=303)

    # si pas "force", on annule
    if force != "1":
        # nettoyage fichiers temporaires
        for it in items:
            try:
                fp = os.path.join(UPLOAD_DIR, it.get("stored") or "")
                if fp and os.path.exists(fp):
                    os.remove(fp)
            except Exception:
                pass
        return RedirectResponse(f"/factures/{invoice_id}", status_code=303)

    for it in items:
        doc_type = (it.get("doc_type") or "AUTRE").strip().upper()
        ref_no = it.get("ref_no")
        ref_date = _parse_date(it.get("ref_date")) if it.get("ref_date") else None

        session.add(
            InvoiceDocument(
                invoice_id=inv.id,
                doc_type=doc_type,
                filename=(it.get("filename") or it.get("stored") or "")[:255],
                stored_path=it.get("stored"),
                ref_no=ref_no,
                ref_date=ref_date,
            )
        )

        if doc_type == "FACTURE" and not inv.file_path:
            inv.file_path = it.get("stored")

        _merge_extraction_into_invoice(inv, it.get("data") or {})

    _finalize_invoice_fields(inv)
    compute_pay_delay_fields(inv)
    session.add(inv)
    session.commit()
    return RedirectResponse(f"/factures/{invoice_id}", status_code=303)


@app.get("/factures/{invoice_id}/add_docs_cancel")
def factures_add_docs_cancel(invoice_id: int, payload: str):
    # supprime fichiers temporaires
    try:
        raw = base64.urlsafe_b64decode(payload.encode("utf-8"))
        obj = json.loads(raw.decode("utf-8"))
        items = obj.get("items") or []
    except Exception:
        items = []
    for it in items:
        try:
            fp = os.path.join(UPLOAD_DIR, it.get("stored") or "")
            if fp and os.path.exists(fp):
                os.remove(fp)
        except Exception:
            pass
    return RedirectResponse(f"/factures/{invoice_id}", status_code=303)



@app.post("/factures/{invoice_id}/docs/{doc_id}/delete")
def facture_doc_delete(invoice_id: int, doc_id: int, session: Session = Depends(get_session)):
    inv = session.get(Invoice, invoice_id)
    if not inv:
        return RedirectResponse("/factures", status_code=303)

    doc = session.get(InvoiceDocument, doc_id)
    if (not doc) or (doc.invoice_id != invoice_id):
        return RedirectResponse(f"/factures/{invoice_id}", status_code=303)

    if inv.file_path and doc.stored_path == inv.file_path:
        inv.file_path = None
        session.add(inv)

    try:
        fp = os.path.join(UPLOAD_DIR, doc.stored_path)
        if os.path.exists(fp):
            os.remove(fp)
    except Exception:
        pass

    session.delete(doc)
    session.commit()
    return RedirectResponse(f"/factures/{invoice_id}", status_code=303)


# ---------------- SUPPLIERS ----------------

@app.get("/suppliers", response_class=HTMLResponse)
def suppliers_page(request: Request, session: Session = Depends(get_session)):
    user = get_current_user(request, session)

    sups = session.exec(select(Supplier).order_by(Supplier.name.asc()).limit(500)).all()
    accs = session.exec(select(SupplierBankAccount).order_by(SupplierBankAccount.created_at.desc())).all()

    by_sup: dict[int, SupplierBankAccount] = {}
    accs_by_sup: dict[int, list[SupplierBankAccount]] = {}
    for a in accs:
        accs_by_sup.setdefault(a.supplier_id, []).append(a)
        if a.supplier_id not in by_sup:
            by_sup[a.supplier_id] = a

    return templates.TemplateResponse(
        "suppliers.html",
        {"request": request, "user": user, "suppliers": sups, "by_sup": by_sup, "accs_by_sup": accs_by_sup},
    )


@app.post("/suppliers/upsert")
async def suppliers_upsert(
    request: Request,
    session: Session = Depends(get_session),

    name: str = Form(...),
    is_foreign: str = Form("0"),   # 0 = Maroc / 1 = Étranger
    country_code: Optional[str] = Form("MA"),

    ice: Optional[str] = Form(None),
    if_code: Optional[str] = Form(None),
    rc: Optional[str] = Form(None),
    rc_city: Optional[str] = Form(None),
    address: Optional[str] = Form(None),

    bank_name: Optional[str] = Form(None),
    agency_name: Optional[str] = Form(None),

    rib_or_iban: Optional[str] = Form(None),
    swift: Optional[str] = Form(None),

    rib_attestation: Optional[UploadFile] = File(None),
):
    """ERP pro: fiche fournisseur indépendante des factures + attestation RIB/IBAN."""

    # --- parsing ---
    is_foreign_bool = (is_foreign == "1")
    cc = (country_code or ("MA" if not is_foreign_bool else "")).strip().upper()[:2] or None

    name_clean = (name or "").strip()[:200] or "A_COMPLETER"
    ice_clean = (ice or "").strip()[:30] or None
    sw = (swift or "").strip()[:20] or None

    # --- retrouver / créer supplier ---
    sup = None
    if ice_clean:
        sup = session.exec(select(Supplier).where(Supplier.ice == ice_clean)).first()
    if (not sup) and name_clean:
        sup = session.exec(select(Supplier).where(Supplier.name == name_clean)).first()

    if not sup:
        sup = Supplier(
            name=name_clean,
            ice=ice_clean,
            if_code=((if_code or "").strip()[:30] or None),
            rc=((rc or "").strip()[:30] or None),
            rc_city=((rc_city or "").strip()[:60] or None),
            address=((address or "").strip()[:250] or None),
            is_foreign=is_foreign_bool,
            country_code=cc,
        )
        session.add(sup)
        session.commit()
        session.refresh(sup)
    else:
        # ✅ ici, comme on est sur la page "fournisseurs", on considère que l'utilisateur valide => on met à jour.
        sup.name = name_clean
        if ice_clean:
            sup.ice = ice_clean
        sup.if_code = ((if_code or "").strip()[:30] or None)
        sup.rc = ((rc or "").strip()[:30] or None)
        sup.rc_city = ((rc_city or "").strip()[:60] or None)
        sup.address = ((address or "").strip()[:250] or None)
        sup.is_foreign = is_foreign_bool
        sup.country_code = cc

        session.add(sup)
        session.commit()

    # --- attestation upload (optionnel) ---
    att_fn = None
    att_path = None
    if rib_attestation is not None and (rib_attestation.filename or "").strip():
        content = await rib_attestation.read()
        if content:
            stored = _safe_store_upload(content, rib_attestation.filename)
            att_fn = (rib_attestation.filename or stored)[:255]
            att_path = stored

    rib_clean = (rib_or_iban or "").strip()[:50] or None
    bank_clean = ((bank_name or "").strip()[:80] or None)
    agency_clean = ((agency_name or "").strip()[:80] or None)

    # --- créer un compte bancaire seulement si on a des infos ou une attestation ---
    has_bank_payload = any([rib_clean, bank_clean, agency_clean, sw, att_path])
    if has_bank_payload:
        acc = SupplierBankAccount(
            supplier_id=sup.id,
            bank_name=bank_clean,
            agency_name=agency_clean,
            rib_or_iban=rib_clean or "",
            swift=sw,
            attestation_filename=att_fn,
            attestation_path=att_path,
        )
        session.add(acc)
        session.commit()

    return RedirectResponse("/suppliers", status_code=303)



@app.post("/suppliers/accounts/{acc_id}/delete")
def supplier_account_delete(acc_id: int, request: Request, session: Session = Depends(get_session)):
    _ = get_current_user(request, session)

    xrw = (request.headers.get("x-requested-with") or "").lower()
    accept = (request.headers.get("accept") or "").lower()
    is_ajax = (xrw == "fetch") or ("application/json" in accept)

    acc = session.get(SupplierBankAccount, acc_id)
    if not acc:
        if is_ajax:
            raise HTTPException(status_code=404, detail="Compte introuvable.")
        return RedirectResponse("/suppliers", status_code=303)

    # Best effort: supprimer l'attestation si elle existe
    try:
        if acc.attestation_path:
            fp = os.path.join(UPLOAD_DIR, acc.attestation_path)
            if os.path.exists(fp):
                os.remove(fp)
    except Exception:
        pass

    session.delete(acc)
    try:
        session.commit()
    except Exception:
        session.rollback()
        msg = "Impossible de supprimer ce compte (il est peut-être utilisé ailleurs)."
        if is_ajax:
            raise HTTPException(status_code=400, detail=msg)
        from urllib.parse import quote
        return RedirectResponse(f"/suppliers?toast={quote(msg)}", status_code=303)

    # OK
    if is_ajax:
        return {"ok": True}
    return RedirectResponse("/suppliers", status_code=303)


@app.post("/suppliers/{supplier_id}/delete")
def supplier_delete(supplier_id: int, request: Request, session: Session = Depends(get_session)):
    _ = get_current_user(request, session)

    xrw = (request.headers.get("x-requested-with") or "").lower()
    accept = (request.headers.get("accept") or "").lower()
    is_ajax = (xrw == "fetch") or ("application/json" in accept)

    sup = session.get(Supplier, supplier_id)
    if not sup:
        if is_ajax:
            raise HTTPException(status_code=404, detail="Fournisseur introuvable.")
        return RedirectResponse("/suppliers", status_code=303)

    # 🔒 Sécurité ERP : si des factures existent, on ne supprime pas.
    linked_inv = session.exec(select(Invoice.id).where(Invoice.supplier_id == supplier_id).limit(1)).first()
    if linked_inv:
        msg = "Impossible de supprimer ce fournisseur : il est lié à des factures. Supprime/annule d'abord les factures ou détache le fournisseur."
        if is_ajax:
            raise HTTPException(status_code=400, detail=msg)
        from urllib.parse import quote
        return RedirectResponse(f"/suppliers?toast={quote(msg)}", status_code=303)

    # Supprimer les comptes associés (et leurs attestations) avant le fournisseur
    accs = session.exec(select(SupplierBankAccount).where(SupplierBankAccount.supplier_id == supplier_id)).all()
    for a in accs:
        try:
            if a.attestation_path:
                fp = os.path.join(UPLOAD_DIR, a.attestation_path)
                if os.path.exists(fp):
                    os.remove(fp)
        except Exception:
            pass
        session.delete(a)

    session.delete(sup)
    try:
        session.commit()
    except Exception:
        session.rollback()
        msg = "Impossible de supprimer ce fournisseur (il est peut-être utilisé ailleurs)."
        if is_ajax:
            raise HTTPException(status_code=400, detail=msg)
        from urllib.parse import quote
        return RedirectResponse(f"/suppliers?toast={quote(msg)}", status_code=303)

    if is_ajax:
        return {"ok": True}
    return RedirectResponse("/suppliers", status_code=303)

# ---------------- BANQUES (modèles ordre de virement) ----------------

@app.get("/banques", response_class=HTMLResponse)
def banques_list(request: Request, session: Session = Depends(get_session)):
    user = get_current_user(request, session)
    banks = session.exec(select(Bank).order_by(Bank.name.asc())).all()
    return templates.TemplateResponse("banques.html", {"request": request, "user": user, "banks": banks})


@app.post("/banques/create")
def banques_create(
    request: Request,
    session: Session = Depends(get_session),
    name: str = Form(...),
    template_code: str = Form(...),
    debit_account: str = Form(""),
    agency: str = Form(""),
    city: str = Form(""),
):
    b = Bank(
        name=name.strip(),
        template_code=template_code.strip().upper(),
        debit_account=(debit_account.strip() or None),
        agency=(agency.strip() or None),
        city=(city.strip() or None),
    )
    session.add(b)
    session.commit()
    return RedirectResponse(url="/banques", status_code=303)



@app.post("/banques/{bank_id}/delete")
def banques_delete(bank_id: int, request: Request, session: Session = Depends(get_session)):
    _ = get_current_user(request, session)

    xrw = (request.headers.get("x-requested-with") or "").lower()
    accept = (request.headers.get("accept") or "").lower()
    is_ajax = (xrw == "fetch") or ("application/json" in accept)

    b = session.get(Bank, bank_id)
    if not b:
        if is_ajax:
            raise HTTPException(status_code=404, detail="Banque introuvable.")
        return RedirectResponse(url="/banques", status_code=303)

    # Si un batch de règlement référence cette banque, la suppression peut être bloquée
    session.delete(b)
    try:
        session.commit()
    except Exception:
        session.rollback()
        msg = "Impossible de supprimer cette banque (elle est peut-être utilisée dans des règlements)."
        if is_ajax:
            raise HTTPException(status_code=400, detail=msg)
        from urllib.parse import quote
        return RedirectResponse(url=f"/banques?toast={quote(msg)}", status_code=303)

    if is_ajax:
        return {"ok": True}
    return RedirectResponse(url="/banques", status_code=303)


# ---------------- BANQUES (comptes société) ----------------

@app.get("/banks", response_class=HTMLResponse)
def banks_page(request: Request, session: Session = Depends(get_session)):
    user = get_current_user(request, session)

    accs = session.exec(
        select(CompanyBankAccount).order_by(CompanyBankAccount.is_active.desc(), CompanyBankAccount.bank_name.asc())
    ).all()

    msg = request.query_params.get("msg")
    return templates.TemplateResponse("banks.html", {"request": request, "user": user, "accs": accs, "msg": msg})


@app.post("/banks/create")
def banks_create(
    request: Request,
    session: Session = Depends(get_session),
    bank_name: str = Form(...),
    account_no: str = Form(...),
    agency_name: Optional[str] = Form(None),
    iban: Optional[str] = Form(None),
    swift: Optional[str] = Form(None),
):
    b = CompanyBankAccount(
        bank_name=bank_name.strip()[:60],
        account_no=account_no.strip()[:80],
        agency_name=(agency_name.strip()[:80] if agency_name else None),
        iban=(iban.strip()[:64] if iban else None),
        swift=(swift.strip()[:32] if swift else None),
        is_active=True,
    )
    session.add(b)
    session.commit()
    return RedirectResponse("/banks", status_code=303)


@app.post("/banks/{acc_id}/toggle")
def banks_toggle(acc_id: int, session: Session = Depends(get_session)):
    acc = session.get(CompanyBankAccount, acc_id)
    if acc:
        acc.is_active = not bool(acc.is_active)
        session.add(acc)
        session.commit()
    return RedirectResponse("/banks", status_code=303)


@app.post("/banks/{acc_id}/upload_attestation")
async def banks_upload_attestation(
    acc_id: int,
    request: Request,
    session: Session = Depends(get_session),
    file: UploadFile = File(...),
):
    acc = session.get(CompanyBankAccount, acc_id)
    if not acc:
        return RedirectResponse("/banks", status_code=303)

    content = await file.read()
    if not content:
        return RedirectResponse("/banks?msg=upload_empty", status_code=303)

    stored = _safe_store_upload(content, file.filename or "attestation_rib")
    acc.attestation_filename = (file.filename or stored)[:255]
    acc.attestation_path = stored
    session.add(acc)
    session.commit()
    return RedirectResponse("/banks?msg=rib_ok", status_code=303)


# ---------------- REGLEMENTS ----------------

def _recalc_batch_total(session: Session, batch_id: int) -> float:
    lines = session.exec(select(PaymentLine).where(PaymentLine.batch_id == batch_id)).all()
    total = 0.0
    for ln in lines:
        total += float(ln.amount or 0.0)
    b = session.get(PaymentBatch, batch_id)
    if b:
        b.total_amount = total
        session.add(b)
        session.commit()
    return total



@app.post("/banks/{acc_id}/delete")
def banks_delete(acc_id: int, request: Request, session: Session = Depends(get_session)):
    _ = get_current_user(request, session)

    xrw = (request.headers.get("x-requested-with") or "").lower()
    accept = (request.headers.get("accept") or "").lower()
    is_ajax = (xrw == "fetch") or ("application/json" in accept)

    acc = session.get(CompanyBankAccount, acc_id)
    if not acc:
        if is_ajax:
            raise HTTPException(status_code=404, detail="Compte bancaire introuvable.")
        return RedirectResponse(url="/banks", status_code=303)

    # Supprimer fichier attestation (best effort)
    try:
        if getattr(acc, "attestation_path", None):
            fp = os.path.join(UPLOAD_DIR, acc.attestation_path)
            if os.path.exists(fp):
                os.remove(fp)
    except Exception:
        pass

    session.delete(acc)
    try:
        session.commit()
    except Exception:
        session.rollback()
        msg = "Impossible de supprimer ce compte bancaire (il est peut-être utilisé)."
        if is_ajax:
            raise HTTPException(status_code=400, detail=msg)
        from urllib.parse import quote
        return RedirectResponse(url=f"/banks?toast={quote(msg)}", status_code=303)

    if is_ajax:
        return {"ok": True}
    return RedirectResponse(url="/banks", status_code=303)



@app.get("/reglements", response_class=HTMLResponse)
def reglements_page(request: Request, session: Session = Depends(get_session)):
    user = get_current_user(request, session)

    batches = session.exec(select(PaymentBatch).order_by(PaymentBatch.id.desc()).limit(200)).all()
    accs = session.exec(
        select(CompanyBankAccount).where(CompanyBankAccount.is_active == True).order_by(CompanyBankAccount.bank_name.asc())
    ).all()

    return templates.TemplateResponse(
        "reglements.html",
        {"request": request, "user": user, "batches": batches, "accs": accs, "today": datetime.utcnow().date().isoformat()},
    )



@app.post("/reglements/{batch_id}/delete")
def reglement_delete_batch(batch_id: int, request: Request, session: Session = Depends(get_session)):
    _ = get_current_user(request, session)

    xrw = (request.headers.get("x-requested-with") or "").lower()
    accept = (request.headers.get("accept") or "").lower()
    is_ajax = (xrw == "fetch") or ("application/json" in accept)

    batch = session.get(PaymentBatch, batch_id)
    if not batch:
        if is_ajax:
            raise HTTPException(status_code=404, detail="Ordre de virement introuvable.")
        return RedirectResponse(url="/reglements", status_code=303)

    # ✅ Règle ERP : si déjà exporté, demander de réinitialiser l'export avant suppression
    if (batch.status or "").upper() == "EXPORTED":
        msg = "Cet ordre est déjà exporté. Clique sur “Réinitialiser export” puis supprime-le."
        if is_ajax:
            raise HTTPException(status_code=400, detail=msg)
        from urllib.parse import quote
        return RedirectResponse(url=f"/reglements?toast={quote(msg)}", status_code=303)

    # Supprimer les lignes enfants d'abord
    try:
        session.exec(text("DELETE FROM paymentline WHERE batch_id = :id"), {"id": batch_id})
    except Exception:
        pass

    session.delete(batch)
    try:
        session.commit()
    except Exception:
        session.rollback()
        msg = "Impossible de supprimer cet ordre de virement (il est peut-être lié à des paiements/exports)."
        if is_ajax:
            raise HTTPException(status_code=400, detail=msg)
        from urllib.parse import quote
        return RedirectResponse(url=f"/reglements?toast={quote(msg)}", status_code=303)

    if is_ajax:
        return {"ok": True}
    return RedirectResponse(url="/reglements", status_code=303)


@app.post("/reglements/{batch_id}/reset_export")
def reglement_reset_export(batch_id: int, request: Request, session: Session = Depends(get_session)):
    _ = get_current_user(request, session)

    xrw = (request.headers.get("x-requested-with") or "").lower()
    accept = (request.headers.get("accept") or "").lower()
    is_ajax = (xrw == "fetch") or ("application/json" in accept)

    batch = session.get(PaymentBatch, batch_id)
    if not batch:
        if is_ajax:
            raise HTTPException(status_code=404, detail="Ordre de virement introuvable.")
        return RedirectResponse(url="/reglements", status_code=303)

    if (batch.status or "").upper() != "EXPORTED":
        msg = "Cet ordre n'est pas dans l'état EXPORTED."
        if is_ajax:
            raise HTTPException(status_code=400, detail=msg)
        from urllib.parse import quote
        return RedirectResponse(url=f"/reglements?toast={quote(msg)}", status_code=303)

    batch.status = "DRAFT"
    session.add(batch)
    session.commit()

    if is_ajax:
        return {"ok": True}
    return RedirectResponse(url=f"/reglements/{batch_id}", status_code=303)



@app.post("/reglements/create")
def reglements_create(
    request: Request,
    session: Session = Depends(get_session),
    account_id: int = Form(...),
    payment_date: Optional[str] = Form(None),
):
    acc = session.get(CompanyBankAccount, account_id)
    if not acc or not acc.is_active:
        return RedirectResponse("/reglements", status_code=303)

    dt = _parse_date(payment_date) if payment_date else None
    if not dt:
        dt = datetime.utcnow()

    b = PaymentBatch(
        company_account_id=acc.id,
        bank_name=(acc.bank_name or "").strip()[:60],
        debit_account=(acc.account_no or "").strip()[:80],
        payment_date=dt,
        total_amount=0.0,
        status="DRAFT",
    )
    session.add(b)
    session.commit()
    session.refresh(b)
    return RedirectResponse(f"/reglements/{b.id}", status_code=303)


@app.post("/reglements/create_from_invoice")
def reglements_create_from_invoice(
    request: Request,
    session: Session = Depends(get_session),
    invoice_id: int = Form(...),
    account_id: int = Form(...),
    payment_date: Optional[str] = Form(None),
    amount: Optional[str] = Form(None),
):
    """Crée un ordre de virement (PaymentBatch) depuis le planning, puis ajoute la ligne."""
    acc = session.get(CompanyBankAccount, account_id)
    if not acc or not acc.is_active:
        return RedirectResponse("/planning?msg=bank", status_code=303)

    inv = session.get(Invoice, invoice_id)
    if not inv or (inv.status or "").upper() != "A_PAYER":
        return RedirectResponse("/planning?msg=inv", status_code=303)

    dt = _parse_date(payment_date) if payment_date else None
    if not dt:
        dt = datetime.utcnow()

    b = PaymentBatch(
        company_account_id=acc.id,
        bank_name=(acc.bank_name or "").strip()[:60],
        debit_account=(acc.account_no or "").strip()[:80],
        payment_date=dt,
        total_amount=0.0,
        status="DRAFT",
    )
    session.add(b)
    session.commit()
    session.refresh(b)

    amt = _to_float(amount) if amount else None
    if amt is None:
        amt = float(inv.amount_ttc if inv.amount_ttc is not None else (inv.amount or 0.0))

    ln = PaymentLine(batch_id=b.id, invoice_id=inv.id, amount=float(amt))
    session.add(ln)
    session.commit()
    _recalc_batch_total(session, b.id)

    return RedirectResponse(f"/reglements/{b.id}", status_code=303)


@app.get("/reglements/{batch_id}", response_class=HTMLResponse)
def reglement_detail(batch_id: int, request: Request, session: Session = Depends(get_session)):
    user = get_current_user(request, session)
    batch = session.get(PaymentBatch, batch_id)
    if not batch:
        raise HTTPException(status_code=404, detail="Batch introuvable")

    lines = session.exec(select(PaymentLine).where(PaymentLine.batch_id == batch_id).order_by(PaymentLine.id.asc())).all()

    inv_ids = [ln.invoice_id for ln in lines]
    invs_in_batch = session.exec(select(Invoice).where(Invoice.id.in_(inv_ids))).all() if inv_ids else []
    inv_map = {i.id: i for i in invs_in_batch}

    invs = session.exec(select(Invoice).where(Invoice.status == "A_PAYER").order_by(Invoice.id.desc()).limit(500)).all()
    invs = [i for i in invs if (i.id not in inv_map)]

    return templates.TemplateResponse("reglement_detail.html", {"request": request, "user": user, "batch": batch, "lines": lines, "invs": invs, "inv_map": inv_map})


@app.post("/reglements/{batch_id}/add_line")
def reglement_add_line(
    batch_id: int,
    request: Request,
    session: Session = Depends(get_session),
    invoice_id: int = Form(...),
    amount: Optional[str] = Form(None),
):
    batch = session.get(PaymentBatch, batch_id)
    if not batch:
        raise HTTPException(status_code=404, detail="Batch introuvable")

    if (batch.status or "").upper() != "DRAFT":
        return RedirectResponse(f"/reglements/{batch_id}", status_code=303)

    inv = session.get(Invoice, invoice_id)
    if not inv or (inv.status or "").upper() != "A_PAYER":
        return RedirectResponse(f"/reglements/{batch_id}", status_code=303)

    already = session.exec(select(PaymentLine).where(PaymentLine.batch_id == batch_id, PaymentLine.invoice_id == invoice_id)).first()
    if already:
        return RedirectResponse(f"/reglements/{batch_id}", status_code=303)

    amt = _to_float(amount) if amount else None
    if amt is None:
        amt = float(inv.amount_ttc if inv.amount_ttc is not None else (inv.amount or 0.0))

    existing_lines = session.exec(select(PaymentLine).where(PaymentLine.batch_id == batch_id)).all()
    if existing_lines:
        any_inv = None
        for ln in existing_lines:
            any_inv = session.get(Invoice, ln.invoice_id)
            if any_inv:
                break
        if any_inv:
            cur0 = (any_inv.currency or "MAD").upper()
            cur1 = (inv.currency or "MAD").upper()
            if cur0 != cur1:
                return RedirectResponse(f"/reglements/{batch_id}", status_code=303)

    ln = PaymentLine(batch_id=batch_id, invoice_id=invoice_id, amount=float(amt))
    session.add(ln)
    session.commit()

    _recalc_batch_total(session, batch_id)
    return RedirectResponse(f"/reglements/{batch_id}", status_code=303)


@app.post("/reglements/{batch_id}/delete_line/{line_id}")
def reglement_delete_line(batch_id: int, line_id: int, session: Session = Depends(get_session)):
    batch = session.get(PaymentBatch, batch_id)
    if not batch:
        return RedirectResponse("/reglements", status_code=303)

    if (batch.status or "").upper() != "DRAFT":
        return RedirectResponse(f"/reglements/{batch_id}", status_code=303)

    ln = session.get(PaymentLine, line_id)
    if not ln or ln.batch_id != batch_id:
        return RedirectResponse(f"/reglements/{batch_id}", status_code=303)

    session.delete(ln)
    session.commit()

    _recalc_batch_total(session, batch_id)
    return RedirectResponse(f"/reglements/{batch_id}", status_code=303)


@app.get("/reglements/{batch_id}/export_csv")
def reglement_export_csv(batch_id: int, session: Session = Depends(get_session)):
    batch = session.get(PaymentBatch, batch_id)
    if not batch:
        raise HTTPException(status_code=404, detail="Batch introuvable")

    lines = session.exec(select(PaymentLine).where(PaymentLine.batch_id == batch_id)).all()

    if (batch.status or "").upper() == "DRAFT":
        batch.status = "EXPORTED"
        session.add(batch)
        session.commit()

    out = ["invoice_id;supplier;invoice_no;amount;currency"]
    for ln in lines:
        inv = session.get(Invoice, ln.invoice_id)
        if not inv:
            continue
        cur = (inv.currency or "MAD").upper()
        out.append(f"{inv.id};{(inv.supplier_name or '').replace(';',' ')};{(inv.invoice_no or '').replace(';',' ')};{ln.amount or 0.0};{cur}")

    content = "\n".join(out)
    filename = f"batch_{batch_id}_export.csv"
    return Response(content=content.encode("utf-8"), media_type="text/csv; charset=utf-8", headers={"Content-Disposition": f'attachment; filename="{filename}"'})


# ✅ ORDRE DE VIREMENT (PDF template via generate_order_pdf)

@app.get("/reglements/{batch_id}/ordre_virement.pdf")
def reglement_ordre_virement_pdf(batch_id: int, session: Session = Depends(get_session)):
    batch = session.get(PaymentBatch, batch_id)
    if not batch:
        raise HTTPException(status_code=404, detail="Batch introuvable")

    # lignes du batch
    lines_db = session.exec(select(PaymentLine).where(PaymentLine.batch_id == batch_id)).all()
    if not lines_db:
        raise HTTPException(status_code=400, detail="Aucune ligne dans le batch")

    # factures du batch
    inv_ids = [ln.invoice_id for ln in lines_db]
    invs = session.exec(select(Invoice).where(Invoice.id.in_(inv_ids))).all() if inv_ids else []
    inv_map = {i.id: i for i in invs}

    # infos banque société (depuis batch)
    bank_name = (batch.bank_name or "").strip()
    debit_account = (batch.debit_account or "").strip()
    payment_date = (batch.payment_date.date() if batch.payment_date else datetime.utcnow().date())

    # template banque (table Bank)
    template_code = "BMCE"
    agency = None
    city = None
    b = session.exec(select(Bank).where(Bank.name == bank_name)).first()
    if b:
        template_code = (b.template_code or "BMCE").strip().upper()
        agency = (b.agency or None)
        city = (b.city or None)

    # ✅ Construire orders: 1 ordre = 1 facture (1 page)
    orders: list[dict] = []

    for ln in lines_db:
        inv = inv_map.get(ln.invoice_id)
        if not inv:
            continue

        amt = float(ln.amount or 0.0)
        cur = (inv.currency or "MAD").upper()

        # bénéficiaire
        beneficiary_name = (inv.supplier_name or "").strip() or "A_COMPLETER"
        beneficiary_rib = ""
        beneficiary_swift = ""

        sup_id = getattr(inv, "supplier_id", None)
        if sup_id:
            acc = session.exec(
                select(SupplierBankAccount)
                .where(SupplierBankAccount.supplier_id == sup_id)
                .order_by(SupplierBankAccount.created_at.desc())
            ).first()
            if acc:
                beneficiary_rib = (acc.rib_or_iban or "").strip()
                beneficiary_swift = (acc.swift or "").strip()

        # 1 seule line => le PDF met le RIB bénéficiaire en haut
        one_line = {
            "beneficiary_name": beneficiary_name,
            "beneficiary_rib": beneficiary_rib,
            "beneficiary_iban": "",
            "beneficiary_swift": beneficiary_swift,

            # optionnel (si ton template l'utilise)
            "ordering_name": "SPIMACO MAROC",

            # libellé par facture
            "motif": f"PAIEMENT FACTURE {inv.invoice_no or inv.id}",

            "amount": amt,
        }

        orders.append(
            {
                "lines": [one_line],
                "total_amount": amt,
                "currency": cur,
            }
        )

    if not orders:
        raise HTTPException(status_code=400, detail="Aucune facture valide dans le batch")

    pdf_bytes = generate_orders_pdf(
        template_code=template_code,
        bank_name=bank_name,
        debit_account=debit_account,
        agency=agency,
        city=city,
        payment_date=payment_date,
        orders=orders,  # ✅ multi-pages
    )

    filename = f"ordres_virement_batch_{batch_id}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{filename}"'},
    )


# ✅ ORDRES PAR FOURNISSEUR (ZIP: 1 PDF par fournisseur)

@app.get("/reglements/{batch_id}/ordres_fournisseurs.zip")
def reglement_ordres_fournisseurs_zip(batch_id: int, session: Session = Depends(get_session)):
    batch = session.get(PaymentBatch, batch_id)
    if not batch:
        raise HTTPException(status_code=404, detail="Batch introuvable")

    lines_db = session.exec(select(PaymentLine).where(PaymentLine.batch_id == batch_id)).all()
    if not lines_db:
        raise HTTPException(status_code=400, detail="Aucune ligne dans le batch")

    # Récupérer factures
    inv_ids = [ln.invoice_id for ln in lines_db]
    invs = session.exec(select(Invoice).where(Invoice.id.in_(inv_ids))).all() if inv_ids else []
    inv_map = {i.id: i for i in invs}

    # Infos banque société (depuis batch)
    bank_name = (batch.bank_name or "").strip()
    debit_account = (batch.debit_account or "").strip()
    payment_date = (batch.payment_date.date() if batch.payment_date else datetime.utcnow().date())

    template_code = "BMCE"
    agency = None
    city = None
    b = session.exec(select(Bank).where(Bank.name == bank_name)).first()
    if b:
        template_code = (b.template_code or "BMCE").strip().upper()
        agency = (b.agency or None)
        city = (b.city or None)

    # Grouper par fournisseur (supplier_id si possible, sinon supplier_name)
    groups: dict[str, list[PaymentLine]] = {}
    for ln in lines_db:
        inv = inv_map.get(ln.invoice_id)
        if not inv:
            continue
        sup_id = getattr(inv, "supplier_id", None)
        if sup_id:
            key = f"ID:{sup_id}"
        else:
            key = f"NAME:{(inv.supplier_name or '').strip().upper()}"
        groups.setdefault(key, []).append(ln)

    # Construire le ZIP
    buf = BytesIO()
    with zipfile.ZipFile(buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for key, lns in groups.items():
            # Déterminer fournisseur + coordonnées bancaires
            any_inv = inv_map.get(lns[0].invoice_id)
            supplier_name = (any_inv.supplier_name or "A_COMPLETER").strip() if any_inv else "A_COMPLETER"
            beneficiary_rib = ""
            beneficiary_swift = ""
            currency = (any_inv.currency or "MAD").upper() if any_inv else "MAD"

            sup_id = getattr(any_inv, "supplier_id", None) if any_inv else None
            if sup_id:
                acc = session.exec(
                    select(SupplierBankAccount)
                    .where(SupplierBankAccount.supplier_id == sup_id)
                    .order_by(SupplierBankAccount.created_at.desc())
                ).first()
                if acc:
                    beneficiary_rib = (acc.rib_or_iban or "").strip()
                    beneficiary_swift = (acc.swift or "").strip()

            # Construire lines pour generate_order_pdf (1 ordre = 1 fournisseur)
            pdf_lines: list[dict] = []
            total_amount = 0.0
            for ln in lns:
                inv = inv_map.get(ln.invoice_id)
                if not inv:
                    continue
                amt = float(ln.amount or 0.0)
                total_amount += amt
                currency = (inv.currency or currency).upper()
                pdf_lines.append(
                    {
                        "beneficiary_name": supplier_name,
                        "beneficiary_rib": beneficiary_rib,
                        "beneficiary_iban": "",
                        "beneficiary_swift": beneficiary_swift,
                        "motif": f"PAIEMENT FACTURE {inv.invoice_no or inv.id}",
                        "amount": amt,
                    }
                )

            pdf_bytes = generate_order_pdf(
                template_code=template_code,
                bank_name=bank_name,
                debit_account=debit_account,
                agency=agency,
                city=city,
                payment_date=payment_date,
                lines=pdf_lines,
                total_amount=float(total_amount),
                currency=currency,
            )

            safe_sup = _slugify_filename(supplier_name)
            zf.writestr(f"ordre_virement_{batch_id}_{safe_sup}.pdf", pdf_bytes)

    filename = f"ordres_virement_batch_{batch_id}.zip"
    return Response(
        content=buf.getvalue(),
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

@app.post("/reglements/{batch_id}/mark_paid")
def reglement_mark_paid(batch_id: int, session: Session = Depends(get_session)):
    batch = session.get(PaymentBatch, batch_id)
    if not batch:
        return RedirectResponse("/reglements", status_code=303)

    if (batch.status or "").upper() == "PAID":
        return RedirectResponse(f"/reglements/{batch_id}", status_code=303)

    lines = session.exec(select(PaymentLine).where(PaymentLine.batch_id == batch_id)).all()

    for ln in lines:
        inv = session.get(Invoice, ln.invoice_id)
        if not inv:
            continue

        total = float(inv.amount_ttc if inv.amount_ttc is not None else (inv.amount or 0.0))
        paid = float(getattr(inv, "amount_paid", 0.0) or 0.0)
        paid += float(ln.amount or 0.0)

        inv.amount_paid = paid
        inv.payment_date = batch.payment_date

        if total > 0 and paid >= total - 0.0001:
            inv.status = "PAYEE"
        else:
            inv.status = "PARTIEL"

        session.add(inv)

    batch.status = "PAID"
    session.add(batch)
    session.commit()
    return RedirectResponse(f"/reglements/{batch_id}", status_code=303)


# ---------------- DEBUG ----------------

@app.get("/debug/uploads", response_class=HTMLResponse)
def debug_uploads(request: Request):
    try:
        files = sorted(os.listdir(UPLOAD_DIR), reverse=True)
    except Exception:
        files = []

    rows = []
    for fn in files[:300]:
        href = f"/uploads/{fn}"
        rows.append(
            f"""
          <tr>
            <td style="padding:8px; border-bottom:1px solid #eee; font-family:monospace;">{fn}</td>
            <td style="padding:8px; border-bottom:1px solid #eee;">
              <a href="{href}" target="_blank" rel="noopener">ouvrir</a>
            </td>
          </tr>
        """
        )

    html = f"""
    <!doctype html>
    <html lang="fr">
    <head>
      <meta charset="utf-8"/>
      <meta name="viewport" content="width=device-width, initial-scale=1"/>
      <title>Debug Uploads</title>
      <link rel="stylesheet" href="/static/style.css"/>
    </head>
    <body>
      <div class="container">
        <div class="card section-gap">
          <h2 style="margin:0 0 6px;">Debug Uploads</h2>
          <div class="muted">UPLOAD_DIR = {UPLOAD_DIR}</div>
          <div class="muted">Nb fichiers (affiché max 300) = {min(len(files), 300)}</div>
          <div style="margin-top:10px;">
            <a class="btn" href="/factures">← Retour factures</a>
          </div>
        </div>

        <div class="card">
          <table style="width:100%; border-collapse:collapse;">
            <thead>
              <tr>
                <th style="text-align:left; padding:8px; border-bottom:1px solid #eee;">Fichier</th>
                <th style="text-align:left; padding:8px; border-bottom:1px solid #eee;">Lien</th>
              </tr>
            </thead>
            <tbody>
              {''.join(rows) if rows else '<tr><td colspan="2" style="padding:10px;" class="muted">Aucun fichier trouvé.</td></tr>'}
            </tbody>
          </table>
        </div>
      </div>
    </body>
    </html>
    """
    return HTMLResponse(content=html)




# ---------------- PARAMETRES : POSTES ----------------

@app.get("/postes", response_class=HTMLResponse)
def postes_page(request: Request, session: Session = Depends(get_session)):
    user = get_current_user(request, session)

    kind = (request.query_params.get("kind") or "BUDGET").upper()
    if kind not in ("BUDGET", "ANALYTIC"):
        kind = "BUDGET"

    posts = session.exec(
        select(Post).where(Post.kind == kind).order_by(Post.is_active.desc(), Post.code, Post.name)
    ).all()

    return templates.TemplateResponse(
        "postes.html",
        {"request": request, "user": user, "kind": kind, "posts": posts},
    )


@app.post("/postes/upsert")
def postes_upsert(
    request: Request,
    session: Session = Depends(get_session),
    kind: str = Form(...),
    code: Optional[str] = Form(None),
    name: str = Form(...),
    is_active: str = Form("1"),
):
    _ = get_current_user(request, session)

    kind = (kind or "").upper().strip()
    if kind not in ("BUDGET", "ANALYTIC"):
        kind = "BUDGET"

    p = Post(
        kind=kind,
        code=(code or "").strip()[:40] or None,
        name=(name or "").strip()[:120],
        is_active=(is_active == "1"),
    )
    session.add(p)
    session.commit()
    return RedirectResponse(f"/postes?kind={kind}", status_code=303)


@app.post("/postes/{post_id}/delete")
def postes_delete(post_id: int, request: Request, session: Session = Depends(get_session)):
    _ = get_current_user(request, session)
    p = session.get(Post, post_id)
    if p:
        kind = p.kind
        session.delete(p)
        session.commit()
        return RedirectResponse(f"/postes?kind={kind}", status_code=303)
    return RedirectResponse("/postes", status_code=303)


@app.get("/debug/routes", response_class=PlainTextResponse)
def debug_routes():
    lines = []
    for r in app.routes:
        name = getattr(r, "name", "")
        path = getattr(r, "path", "")
        methods = ",".join(sorted(getattr(r, "methods", []) or []))
        lines.append(f"{methods:20s} {path:35s}  {name}")
    return "\n".join(lines)

